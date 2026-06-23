#!/usr/bin/env python
# coding: utf-8
"""
09_matching_universe_v5.py
==========================
Matching Universe — AI 기반 채용 후보자 매칭 파이프라인 (v5)
Converted from: 09_matching_universe_v5.ipynb

실행 방법 (Windows 터미널 기준):
  python 09_matching_universe_v5.py

주요 설정은 "Cell 01 | 경로 및 실행 설정" 블록에서 변경하세요.
  - MANUAL_TARGET_JD_IDS  : 매칭할 JD 목록
  - TOP_N_CROSS_ENCODER   : Cross-Encoder 적용 상위 N명
  - USE_LLM_EVALUATION    : Gemini LLM 해설 on/off

Gemini API Key 환경변수 설정 (LLM 해설 사용 시):
  set GEMINI_API_KEY=발급받은_API_KEY   (Windows cmd)
  $env:GEMINI_API_KEY="..."             (Windows PowerShell)
"""

# coding: utf-8

# # 09 Matching Universe v5
# 
# ### v5 변경 사항
# 
# - 기존 v4 노트북을 기준으로 `_v5` 산출물명을 적용합니다.
# - `latest_matching_result_all_semantics` 추가 결과 파일을 생성합니다.
# - 추가 결과 파일은 `latest_matching_result_all`과 동일한 매칭 기준을 따르되, Gold/가상 CV를 제외합니다.
# - JD별 상위 N 제한 없이 JD별 전원 후보에 대해 맥락 해석 컬럼을 생성합니다.
# - 이전 855명 결과 누락 문제를 방지하기 위해 JD별 건수, Gold CV 제외 건수, 최종 semantics 건수를 검증하는 audit 셀을 포함합니다.
# 

# In[1]:


# Cell 00 | 라이브러리 로드
from pathlib import Path
import os
import sys
import re
import json
import time
import warnings
import numpy as np
import pandas as pd

from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

warnings.filterwarnings("ignore")
pd.set_option("display.max_columns", 160)
pd.set_option("display.max_colwidth", 220)

print("라이브러리 로드 완료")


# In[ ]:


# Cell 01 | 경로 및 실행 설정
# ------------------------------------------------------------
# Windows 실제 운영 경로를 기본값으로 사용합니다.
# ChatGPT 임시 업로드 파일명에는 괄호 표기를 쓰지 않습니다.
# 단, 임시 실행 환경에서 업로드명이 자동 변경된 경우 glob 탐색으로만 보정합니다.
# ------------------------------------------------------------

from pathlib import Path
import os
import sys
import glob

CURRENT_DIR = Path.cwd()
WINDOWS_PROJECT_ROOT = Path("D:/003.AIPJT")


def find_project_root(start_dir=None):
    """실행 위치와 관계없이 프로젝트 루트를 찾습니다."""
    start_dir = Path(start_dir or Path.cwd()).resolve()
    candidates = [
        WINDOWS_PROJECT_ROOT,
        WINDOWS_PROJECT_ROOT / "heading_project",
        start_dir,
        *start_dir.parents,
        start_dir / "heading_project",
        start_dir.parent / "heading_project" if start_dir.parent else start_dir / "heading_project",
    ]
    for cand in candidates:
        if (cand / "data").exists() or (cand / "src" / "normalization" / "taxonomy_text.py").exists():
            return cand
        if (cand / "heading_project" / "src" / "normalization" / "taxonomy_text.py").exists():
            return cand / "heading_project"
    return WINDOWS_PROJECT_ROOT if WINDOWS_PROJECT_ROOT.exists() else start_dir


PROJECT_ROOT = find_project_root(CURRENT_DIR)
DATA_DIR = PROJECT_ROOT / "data"

if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

# ------------------------------------------------------------
# 입력 경로: Cell 02에서 수정된 경로 기준
# ------------------------------------------------------------
CV_MASTER_PATH = DATA_DIR / "05_profiles/candidate_profile_jo_all_FINAL5.xlsx"
JD_MASTER_PATH = DATA_DIR / "10_jd/jd_dataset_613_filled.csv"

TAXONOMY_DIR = DATA_DIR / "11_taxonomy"
STOPWORDS_PATH = PROJECT_ROOT / "configs/taxonomy_stopwords.yaml"
LLM_OUTPUT_DIR = DATA_DIR / "06_llm_outputs"
EMBEDDING_DIR = DATA_DIR / "07_embeddings"
MATCHING_OUTPUT_DIR = DATA_DIR / "08_matching"
DASHBOARD_OUTPUT_DIR = DATA_DIR / "09_dashboard"
EVALUATION_DIR = DATA_DIR / "10_evaluation"

CV_EMBEDDING_PATH = EMBEDDING_DIR / "cv_embeddings_only_bge_m3_10.npz"
CV_EMBEDDING_META_PATH = EMBEDDING_DIR / "cv_embedding_meta_bge_m3_10.csv"
CV_EMBEDDING_PROGRESS_PATH = EMBEDDING_DIR / "cv_embedding_progress_bge_m3_10.json"
JD_EMBEDDING_PATH = EMBEDDING_DIR / "jd_embeddings_bge_m3.npy"

GROUND_TRUTH_PATH = EVALUATION_DIR / "ground_truth_matches.xlsx"
EVALUATION_REPORT_PATH = EVALUATION_DIR / "evaluation_report_latest.xlsx"
WEIGHT_VALIDATION_PATH = EVALUATION_DIR / "weight_config_validation_latest.xlsx"

# ChatGPT 임시 실행 fallback: 괄호 없는 표준 파일명만 명시합니다.
SANDBOX_DIR = Path("/mnt/data")
SANDBOX_STANDARD_FILES = {
    "CV_EMBEDDING_PATH": SANDBOX_DIR / "cv_embeddings_only_bge_m3_10.npz",
    "CV_EMBEDDING_META_PATH": SANDBOX_DIR / "cv_embedding_meta_bge_m3_10.csv",
    "CV_EMBEDDING_PROGRESS_PATH": SANDBOX_DIR / "cv_embedding_progress_bge_m3_10.json",
    "JD_EMBEDDING_PATH": SANDBOX_DIR / "jd_embeddings_bge_m3.npy",
}
SANDBOX_GLOB_PATTERNS = {
    "CV_EMBEDDING_PATH": "cv_embeddings_only_bge_m3_10*.npz",
    "CV_EMBEDDING_META_PATH": "cv_embedding_meta_bge_m3_10*.csv",
    "CV_EMBEDDING_PROGRESS_PATH": "cv_embedding_progress_bge_m3_10*.json",
    "JD_EMBEDDING_PATH": "jd_embeddings_bge_m3*.npy",
}


def resolve_input_path(path, standard_fallback=None, glob_pattern=None, required=True):
    """운영 경로 → 표준 임시 경로 → glob 탐색 순서로 파일을 찾습니다."""
    path = Path(path)
    if path.exists():
        return path
    if standard_fallback is not None and Path(standard_fallback).exists():
        return Path(standard_fallback)
    if glob_pattern and SANDBOX_DIR.exists():
        matches = sorted(SANDBOX_DIR.glob(glob_pattern))
        if matches:
            return matches[0]
    if required:
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {path}")
    return path


# 임베딩/보조 파일은 실행 환경에 따라 자동 resolve합니다.
CV_EMBEDDING_PATH = resolve_input_path(
    CV_EMBEDDING_PATH,
    SANDBOX_STANDARD_FILES["CV_EMBEDDING_PATH"],
    SANDBOX_GLOB_PATTERNS["CV_EMBEDDING_PATH"],
    required=True,
)
CV_EMBEDDING_META_PATH = resolve_input_path(
    CV_EMBEDDING_META_PATH,
    SANDBOX_STANDARD_FILES["CV_EMBEDDING_META_PATH"],
    SANDBOX_GLOB_PATTERNS["CV_EMBEDDING_META_PATH"],
    required=False,
)
CV_EMBEDDING_PROGRESS_PATH = resolve_input_path(
    CV_EMBEDDING_PROGRESS_PATH,
    SANDBOX_STANDARD_FILES["CV_EMBEDDING_PROGRESS_PATH"],
    SANDBOX_GLOB_PATTERNS["CV_EMBEDDING_PROGRESS_PATH"],
    required=False,
)
JD_EMBEDDING_PATH = resolve_input_path(
    JD_EMBEDDING_PATH,
    SANDBOX_STANDARD_FILES["JD_EMBEDDING_PATH"],
    SANDBOX_GLOB_PATTERNS["JD_EMBEDDING_PATH"],
    required=True,
)

# 출력 경로
for _dir in [LLM_OUTPUT_DIR, EMBEDDING_DIR, MATCHING_OUTPUT_DIR, DASHBOARD_OUTPUT_DIR, EVALUATION_DIR]:
    _dir.mkdir(parents=True, exist_ok=True)

OUTPUT_MATCHING_CSV = MATCHING_OUTPUT_DIR / "matching_result_dashboard_demo_A_jo.csv"
OUTPUT_MATCHING_XLSX = MATCHING_OUTPUT_DIR / "matching_result_dashboard_demo_A_jo.xlsx"
OUTPUT_DASHBOARD_SUMMARY_CSV = DASHBOARD_OUTPUT_DIR / "dashboard_summary_demo_A_jo.csv"
OUTPUT_DASHBOARD_HTML = DASHBOARD_OUTPUT_DIR / "dashboard_preview_demo_A_jo.html"
OUTPUT_LLM_INPUT_JSONL = LLM_OUTPUT_DIR / "llm_interpretation_input_demo_A_jo.jsonl"

# 새 CV/JD 임베딩을 최우선 후보로 사용합니다.
PRECOMPUTED_EMBEDDING_CANDIDATE_PATHS = [CV_EMBEDDING_PATH]
PRECOMPUTED_EMBEDDING_PATH = CV_EMBEDDING_PATH
SELECTED_PRECOMPUTED_EMBEDDING_PATH = None
PRECOMPUTED_EMBEDDING_METADATA_PATH = EMBEDDING_DIR / "bge_embedding_metadata.json"

# ------------------------------------------------------------
# 실행 대상 설정
# ------------------------------------------------------------
USE_GROUND_TRUTH_JD_IDS = True
MANUAL_TARGET_JD_IDS = [191, 236, 605, 166, 278, 409, 14, 321, 490, 424]


def load_ground_truth_target_jd_ids(path):
    """Ground Truth 평가셋에서 평가 대상 JD 목록을 읽습니다."""
    path = Path(path)
    if not path.exists():
        print(f"Ground Truth 파일 없음: {path}")
        return []
    sheet_name = "ground_truth_matches"
    try:
        available_sheets = pd.ExcelFile(path).sheet_names
        if sheet_name not in available_sheets:
            sheet_name = available_sheets[0]
        gt_preview = pd.read_excel(path, sheet_name=sheet_name, usecols=["jd_id"])
    except Exception as e:
        print(f"Ground Truth JD 목록 로드 실패: {type(e).__name__}: {e}")
        return []
    jd_ids = (
        gt_preview["jd_id"]
        .dropna()
        .astype(str)
        .str.replace(r"\.0$", "", regex=True)
        .drop_duplicates()
        .tolist()
    )
    return jd_ids


GROUND_TRUTH_TARGET_JD_IDS = load_ground_truth_target_jd_ids(GROUND_TRUTH_PATH)
TARGET_JD_IDS = GROUND_TRUTH_TARGET_JD_IDS if USE_GROUND_TRUTH_JD_IDS and GROUND_TRUTH_TARGET_JD_IDS else MANUAL_TARGET_JD_IDS

JD_SAMPLE_N = None
CV_SAMPLE_N = None
# v3: 전체 후보 final_score 랭킹 산출을 위해 retrieval 후보 pool을 전체 후보 수로 확대합니다.
# - CrossEncoder를 사용하는 경우 JD별 865명 전체에 대해 정밀 점수를 다시 계산합니다.
# - latest_matching_result_all은 TOP_N_FINAL=10을 유지하여 JD별 Top 10 제출용으로 저장합니다.
TOP_N_RETRIEVAL = 865
TOP_N_CROSS_ENCODER = 100
TOP_N_FINAL = 10
TOP_N_PER_JD = TOP_N_FINAL
CROSS_ENCODER_BATCH_SIZE = 16
MAX_TASK_CHARS = 800
MAX_DOMAIN_CHARS = 600
FULL_RANKING_PLUS10_OUTPUT_BASENAME = "latest_matching_result_855plus10_v5"
FULL_RANKING_855ONLY_OUTPUT_BASENAME = "latest_matching_result_855only_v5"
EXPECTED_CV_COUNT_FROM_REQUEST = 855
EXPECTED_CV_COUNT_FROM_NEW_FILES = 865

# v5 추가 산출물: Gold/가상 CV를 제외한 JD별 전원 맥락 해석 파일
SEMANTICS_OUTPUT_BASENAME = "latest_matching_result_all_semantics"
SEMANTICS_EXCLUDE_GOLD_CV = True
SEMANTICS_FULL_UNIVERSE_MODE = True

# ------------------------------------------------------------
# 모델 사용 설정
# ------------------------------------------------------------
USE_EMBEDDING_RETRIEVAL = True
USE_PRECOMPUTED_EMBEDDING = True
USE_CROSS_ENCODER_REQUESTED = True
USE_CROSS_ENCODER = USE_CROSS_ENCODER_REQUESTED
USE_LLM_EVALUATION = False                      # 기본값 False
USE_GEMINI = False                              # 기본값 False
USE_OPENAI_API = False                          # 기본값 False
USE_LLM_EXPLANATION = False                     # 기본값 False
USE_HARDSKILL_EVIDENCE_LLM = False              # 기본값 False
USE_SOFTSKILL_LLM = False                       # 기본값 False

EMBEDDING_MODEL_NAME = "BAAI/bge-m3"
CROSS_ENCODER_MODEL_NAME = "BAAI/bge-reranker-v2-m3"
GEMINI_MODEL = "gemini-2.5-flash"
LLM_EXPLAIN_TOP_N_PER_JD = 10
USE_CAREER_PATH_LLM = False
CAREER_PATH_TOP_N_PER_JD = 10

# ------------------------------------------------------------
# Final score 가중치
# ------------------------------------------------------------
RETRIEVAL_SCORE_WEIGHT = 0.25
KEYWORD_BASELINE_SCORE_WEIGHT = 0.15
COLUMN_SCORE_WEIGHT = 0.75

print("PROJECT_ROOT:", PROJECT_ROOT)
print("CV_MASTER_PATH:", CV_MASTER_PATH)
print("JD_MASTER_PATH:", JD_MASTER_PATH)
print("CV_EMBEDDING_PATH:", CV_EMBEDDING_PATH)
print("CV_EMBEDDING_META_PATH:", CV_EMBEDDING_META_PATH)
print("CV_EMBEDDING_PROGRESS_PATH:", CV_EMBEDDING_PROGRESS_PATH)
print("JD_EMBEDDING_PATH:", JD_EMBEDDING_PATH)
print("TAXONOMY_DIR:", TAXONOMY_DIR)
print("STOPWORDS_PATH:", STOPWORDS_PATH)
print("GROUND_TRUTH_PATH:", GROUND_TRUTH_PATH)
print("PRECOMPUTED_EMBEDDING_CANDIDATE_PATHS:")
for _p in PRECOMPUTED_EMBEDDING_CANDIDATE_PATHS:
    print(" -", _p)
print("USE_GROUND_TRUTH_JD_IDS:", USE_GROUND_TRUTH_JD_IDS)
print("TARGET_JD_IDS:", TARGET_JD_IDS)
print("TOP_N_RETRIEVAL:", TOP_N_RETRIEVAL)
print("TOP_N_CROSS_ENCODER:", TOP_N_CROSS_ENCODER)
print("CROSS_ENCODER_BATCH_SIZE:", CROSS_ENCODER_BATCH_SIZE)
print("MAX_TASK_CHARS:", MAX_TASK_CHARS)
print("MAX_DOMAIN_CHARS:", MAX_DOMAIN_CHARS)
print("TOP_N_FINAL:", TOP_N_FINAL)
print("FULL_RANKING_PLUS10_OUTPUT_BASENAME:", FULL_RANKING_PLUS10_OUTPUT_BASENAME)
print("FULL_RANKING_855ONLY_OUTPUT_BASENAME:", FULL_RANKING_855ONLY_OUTPUT_BASENAME)


# In[ ]:


# Cell 02 | API Key 입력부
# ------------------------------------------------------------
# 권장 방식 1: 환경변수 사용
#   export GEMINI_API_KEY="발급받은키"
#
# 권장 방식 2: 노트북에서 임시 입력
# os.environ["GEMINI_API_KEY"] = "API 값을 입력하세요"
#
# 주의: 노트북을 공유할 때 API Key가 저장되지 않도록 실행 후 지우세요.
# ------------------------------------------------------------

GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "")

if GEMINI_API_KEY:
    print("GEMINI_API_KEY 환경변수 확인 완료")
else:
    print("GEMINI_API_KEY가 비어 있습니다. LLM optional cell 실행 전 환경변수를 설정하세요.")


# In[4]:


# Cell 03 | CV/JD 임베딩 파일 형식 검증
# ------------------------------------------------------------
# CV: cv_embeddings_only_bge_m3_10.npz
# JD: jd_embeddings_bge_m3.npy
# BGE-M3 기준 1024차원 여부, row 수, key 구성을 먼저 확인합니다.
# ------------------------------------------------------------

cv_npz_check = np.load(CV_EMBEDDING_PATH, allow_pickle=True)
cv_npz_keys = list(cv_npz_check.files)
print("CV npz keys:", cv_npz_keys)

cv_embedding_key = None
for _key in ["embeddings", "cv_embeddings", "cv_embedding", "candidate_embeddings"]:
    if _key in cv_npz_keys:
        cv_embedding_key = _key
        break
if cv_embedding_key is None:
    raise KeyError(f"CV npz에서 embedding key를 찾지 못했습니다. keys={cv_npz_keys}")

cv_embedding_shape = tuple(cv_npz_check[cv_embedding_key].shape)
cv_embedding_dtype = str(cv_npz_check[cv_embedding_key].dtype)
cv_candidate_id_count = int(len(cv_npz_check["candidate_ids"])) if "candidate_ids" in cv_npz_keys else None

jd_embedding_check = np.load(JD_EMBEDDING_PATH, allow_pickle=True)
if jd_embedding_check.ndim != 2:
    raise ValueError(f"JD embedding은 2차원이어야 합니다. 현재 shape={jd_embedding_check.shape}")

embedding_file_report = {
    "cv_embedding_path": str(CV_EMBEDDING_PATH),
    "cv_npz_keys": cv_npz_keys,
    "cv_embedding_key": cv_embedding_key,
    "cv_shape": cv_embedding_shape,
    "cv_dtype": cv_embedding_dtype,
    "cv_candidate_id_count": cv_candidate_id_count,
    "cv_all_finite": bool(np.isfinite(cv_npz_check[cv_embedding_key]).all()),
    "jd_embedding_path": str(JD_EMBEDDING_PATH),
    "jd_shape": tuple(jd_embedding_check.shape),
    "jd_dtype": str(jd_embedding_check.dtype),
    "jd_all_finite": bool(np.isfinite(jd_embedding_check).all()),
}

if cv_embedding_shape[1] != jd_embedding_check.shape[1]:
    raise ValueError(f"CV/JD embedding 차원이 다릅니다. CV={cv_embedding_shape}, JD={jd_embedding_check.shape}")
if cv_embedding_shape[1] != 1024:
    print(f"주의: BGE-M3 예상 차원은 1024입니다. 현재 CV 차원={cv_embedding_shape[1]}")
if cv_embedding_shape[0] != EXPECTED_CV_COUNT_FROM_REQUEST:
    print(f"주의: 요청 파일명은 855 기준이지만 실제 CV embedding 행 수는 {cv_embedding_shape[0]}입니다.")
if cv_embedding_shape[0] == EXPECTED_CV_COUNT_FROM_NEW_FILES:
    print("확인: 새 CV 임베딩 진행 파일 기준 865건과 일치합니다.")

print(json.dumps(embedding_file_report, ensure_ascii=False, indent=2))


# In[5]:


# Cell 04 | CV / JD Master 로드

def read_table_auto(path):
    """csv/xlsx 파일을 확장자에 맞게 안전하게 로드합니다."""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"파일이 없습니다: {path}")

    suffix = path.suffix.lower()
    if suffix in [".xlsx", ".xlsm", ".xls"]:
        return pd.read_excel(path)

    if suffix == ".csv":
        last_error = None
        for enc in ["utf-8-sig", "utf-8", "cp949", "euc-kr", "latin1"]:
            try:
                return pd.read_csv(path, encoding=enc)
            except UnicodeDecodeError as e:
                last_error = e
                continue
        raise last_error

    raise ValueError(f"지원하지 않는 파일 확장자입니다: {path}")

cv_df = read_table_auto(CV_MASTER_PATH)
jd_df = read_table_auto(JD_MASTER_PATH)

# JD embedding .npy는 원본 JD row 순서 기준으로 생성되어 있으므로, 필터링 전 원본 row index를 보존합니다.
cv_df["__source_row_id"] = np.arange(len(cv_df))
jd_df["__source_row_id"] = np.arange(len(jd_df))

# 모든 컬럼을 문자열로 통일하면 taxonomy / text concat 단계에서 에러가 줄어듭니다.
for df in [cv_df, jd_df]:
    for col in df.columns:
        df[col] = df[col].fillna("").astype(str)

# ------------------------------------------------------------
# JD 선택 로직
# ------------------------------------------------------------
# 1) TARGET_JD_IDS가 있으면 해당 JD만 선택합니다.
# 2) TARGET_JD_IDS가 비어 있고 JD_SAMPLE_N이 있으면 head 샘플을 사용합니다.
# 3) CV_SAMPLE_N=None이면 855명 전체 CV를 사용합니다.
#
# v22.1 수정:
# - TARGET_JD_IDS=["191, 236, ..."]처럼 쉼표가 포함된 문자열 1개로 들어와도
#   ["191", "236", ...] 형태로 자동 분해합니다.
# - int / float / str / tuple / set / pandas Series 입력을 모두 안전하게 처리합니다.
# ------------------------------------------------------------
def normalize_target_jd_ids(raw_ids):
    """TARGET_JD_IDS를 JD 원본의 jd_id와 비교 가능한 문자열 리스트로 정규화합니다."""
    if raw_ids is None:
        return []

    if isinstance(raw_ids, (str, int, float)):
        raw_iter = [raw_ids]
    else:
        raw_iter = list(raw_ids)

    normalized = []
    for item in raw_iter:
        if item is None:
            continue
        text = str(item).strip()
        if not text or text.lower() == "nan":
            continue

        # "191, 236, 605" / "191 236 605" / "191\n236" 모두 허용
        parts = re.split(r"[,;\s]+", text)
        for part in parts:
            part = str(part).strip().strip("[](){}'\"")
            if not part or part.lower() == "nan":
                continue
            part = re.sub(r"\.0$", "", part)
            normalized.append(part)

    # 순서 보존 중복 제거
    return list(dict.fromkeys(normalized))


def normalize_id_series(series):
    return series.astype(str).str.strip().str.replace(r"\.0$", "", regex=True)


if CV_SAMPLE_N is not None:
    cv_df = cv_df.head(CV_SAMPLE_N).copy()

if 'TARGET_JD_IDS' in globals() and TARGET_JD_IDS:
    target_ids = normalize_target_jd_ids(TARGET_JD_IDS)
    before_n = len(jd_df)
    original_jd_ids = normalize_id_series(jd_df["jd_id"]) if "jd_id" in jd_df.columns else pd.Series([], dtype=str)

    jd_df = jd_df[original_jd_ids.isin(target_ids)].copy()
    if jd_df.empty:
        available = original_jd_ids.head(20).tolist()
        raise ValueError(
            f"TARGET_JD_IDS={target_ids}에 해당하는 JD가 없습니다. "
            f"JD 원본 jd_id 예시={available}. "
            "Cell 01의 MANUAL_TARGET_JD_IDS 또는 Ground Truth jd_id를 확인하세요."
        )
    missing_ids = [x for x in target_ids if x not in set(original_jd_ids)]
    if missing_ids:
        print(f"경고: JD 원본에 없는 TARGET_JD_IDS 제외: {missing_ids}")
    print(f"TARGET_JD_IDS 적용: {target_ids} | JD {before_n}건 중 {len(jd_df)}건 선택")
elif JD_SAMPLE_N is not None:
    jd_df = jd_df.head(JD_SAMPLE_N).copy()

print("CV columns:")
print(cv_df.columns.tolist())

print()
print("JD columns:")
print(jd_df.columns.tolist())

display(cv_df.head(2))
display(jd_df.head(2))


# ## 1. Taxonomy 설정
# 
# 이 노트북은 `data/11_taxonomy/*_taxonomy_jo.xlsx`를 active taxonomy source로 사용합니다.
# 
# 연결 우선순위:
# 
# ```text
# 1. data/11_taxonomy/*_taxonomy_jo.xlsx
# 2. 노트북 내부 기본 Dictionary fallback
# ```
# 
# 예전 `data/taxonomy/` 폴더와 백업본은 혼동을 줄이기 위해 `data/_archive/legacy_reference_tables/`에 보관합니다.
# 
# Taxonomy의 목표는 원문 컬럼을 매칭 가능한 표준어 컬럼으로 바꾸는 것입니다.
# 
# ```text
# CV Skill / Career_Description / Career / Position
# ↓
# cv_skill_standard / cv_task_standard / cv_domain_standard / cv_role_standard
# 
# JD required_skill / main_task / domain / role_signal
# ↓
# jd_skill_standard / jd_task_standard / jd_domain_standard / jd_role_standard
# ```
# 

# In[6]:


# Cell 05 | Taxonomy 로드 및 기본 Dictionary 정의
from src.normalization.taxonomy_text import load_stopwords, is_stopword, normalize_token

TAXONOMY_STOPWORDS, TAXONOMY_STOPWORD_PATTERNS = load_stopwords(STOPWORDS_PATH)


def _is_taxonomy_noise_term(value):
    token = normalize_token(value)
    if not token:
        return True
    parts = token.split()
    if is_stopword(token, TAXONOMY_STOPWORDS, TAXONOMY_STOPWORD_PATTERNS):
        return True
    if parts and all(is_stopword(part, TAXONOMY_STOPWORDS, TAXONOMY_STOPWORD_PATTERNS) for part in parts):
        return True
    return False


# ------------------------------------------------------------
# 연결 우선순위:
# 1) data/11_taxonomy/*_taxonomy_jo.xlsx
#    - raw_term, standard_term, taxonomy_type, category, use_yn 컬럼 사용
#    - use_yn이 있으면 Y/YES/TRUE/1만 사용
# 2) 외부 파일이 비었거나 읽기 실패하면 아래 기본 Dictionary 유지
# ------------------------------------------------------------

USE_EXTERNAL_TAXONOMY = True

DEFAULT_SKILL_TAXONOMY = {
    # language / data
    "파이썬": "Python", "python": "Python", "python3": "Python",
    "sql": "SQL", "mysql": "SQL", "postgresql": "SQL", "oracle": "SQL", "bigquery": "BigQuery",
    "pandas": "Pandas", "numpy": "NumPy", "scikit": "Scikit-learn", "sklearn": "Scikit-learn",
    # bi / dashboard
    "tableau": "Tableau", "looker": "Looker Studio", "looker studio": "Looker Studio",
    "power bi": "Power BI", "dashboard": "Dashboard", "대시보드": "Dashboard",
    # ml / ai
    "tensorflow": "TensorFlow", "pytorch": "PyTorch", "머신러닝": "Machine Learning",
    "딥러닝": "Deep Learning", "llm": "LLM", "rag": "RAG", "langchain": "LangChain",
    "faiss": "FAISS", "추천": "Recommendation",
    # cloud / data engineering
    "spark": "Spark", "aws": "AWS", "gcp": "GCP", "airflow": "Airflow",
    # analytics tools
    "google analytics": "Google Analytics", "ga4": "Google Analytics", "google tag manager": "Google Tag Manager",
    "excel": "Excel", "figma": "Figma", "qgis": "QGIS",
}

DEFAULT_TASK_TAXONOMY = {
    "고객 분석": "Customer Analytics", "고객": "Customer Analytics", "이탈": "Churn Analysis", "churn": "Churn Analysis",
    "세그먼트": "Customer Segmentation", "페르소나": "Persona Analysis", "퍼널": "Funnel Analysis",
    "추천": "Recommendation", "recommendation": "Recommendation",
    "예측": "Predictive Modeling", "모델": "Modeling", "모델링": "Modeling",
    "분석": "Data Analysis", "시각화": "Data Visualization", "대시보드": "Dashboard",
    "자동화": "Automation", "전처리": "Data Preprocessing", "정제": "Data Preprocessing",
    "a/b": "Experiment Analysis", "ab test": "Experiment Analysis", "실험": "Experiment Analysis",
    "seo": "SEO Analysis", "전환": "Conversion Analysis", "리포트": "Reporting", "보고서": "Reporting",
}

DEFAULT_DOMAIN_TAXONOMY = {
    "커머스": "Commerce", "이커머스": "Commerce", "쇼핑": "Commerce", "상품": "Commerce",
    "금융": "Finance", "은행": "Finance", "카드": "Finance", "보험": "Finance",
    "통신": "Telecom", "브로드밴드": "Telecom", "sk 브로드밴드": "Telecom",
    "게임": "Game", "제조": "Manufacturing", "공공": "Public", "공공기관": "Public",
    "의료": "Healthcare", "바이오": "Bio", "교육": "Education", "마케팅": "Marketing",
    "플랫폼": "Platform", "국방": "Defense",
    "엔터테인먼트": "Media/Contents", "미디어": "Media/Contents", "콘텐츠": "Media/Contents",
    "entertainment": "Media/Contents", "media": "Media/Contents", "content": "Media/Contents",
}

DEFAULT_ROLE_TAXONOMY = {
    "data analyst": "Data Analyst", "데이터 분석": "Data Analyst", "데이터분석": "Data Analyst",
    "bi analyst": "BI Analyst", "business analyst": "Business Analyst",
    "data scientist": "Data Scientist", "데이터 사이언티스트": "Data Scientist",
    "ml engineer": "ML Engineer", "머신러닝 엔지니어": "ML Engineer",
    "ai engineer": "AI Engineer", "개발자": "Developer", "developer": "Developer",
    "기획": "Product/Service Planner", "서비스 기획": "Product/Service Planner",
    "researcher": "Researcher", "연구": "Researcher",
}


def _empty_taxonomy_bundle():
    return {"skill": {}, "task": {}, "domain": {}, "role": {}}


def _taxonomy_counts(bundle):
    return {k: len(v) for k, v in bundle.items()}


def _merge_taxonomy_bundle(base, extra):
    merged = {k: dict(base.get(k, {})) for k in ["skill", "task", "domain", "role"]}
    for tax_type, mapping in extra.items():
        merged.setdefault(tax_type, {})
        merged[tax_type].update(mapping)
    return merged


def _clean_cell(value):
    if pd.isna(value):
        return ""
    return str(value).strip()


def _is_active_row(row):
    if "use_yn" not in row.index or _clean_cell(row.get("use_yn", "")) == "":
        return True
    return _clean_cell(row.get("use_yn", "")).upper() in {"Y", "YES", "TRUE", "1"}


def _add_alias(mapping, alias, canonical):
    alias = _clean_cell(alias)
    canonical = _clean_cell(canonical)
    if not alias or not canonical:
        return
    if _is_taxonomy_noise_term(alias) or _is_taxonomy_noise_term(canonical):
        return
    mapping[alias] = canonical


def _load_reviewed_taxonomy_files(taxonomy_dir):
    bundle = _empty_taxonomy_bundle()
    file_map = {
        "skill": taxonomy_dir / "skill_taxonomy_jo.xlsx",
        "task": taxonomy_dir / "task_taxonomy_jo.xlsx",
        "domain": taxonomy_dir / "domain_taxonomy_jo.xlsx",
        "role": taxonomy_dir / "role_taxonomy_jo.xlsx",
    }

    for tax_type, path in file_map.items():
        if not path.exists():
            print(f"검수 taxonomy 파일 없음: {path}")
            continue
        try:
            df = pd.read_excel(path)
        except Exception as e:
            print(f"검수 taxonomy 읽기 실패: {path} | {e}")
            continue

        required = {"raw_term", "standard_term"}
        if not required.issubset(df.columns):
            print(f"검수 taxonomy 컬럼 부족: {path} | columns={df.columns.tolist()}")
            continue

        for _, row in df.iterrows():
            if not _is_active_row(row):
                continue
            raw_term = _clean_cell(row.get("raw_term", ""))
            standard_term = _clean_cell(row.get("standard_term", "")) or raw_term
            _add_alias(bundle[tax_type], raw_term, standard_term)

    return bundle


def _load_keyword_template(path):
    bundle = _empty_taxonomy_bundle()
    if not path.exists():
        return bundle
    try:
        df = pd.read_excel(path, sheet_name="taxonomy_keywords")
    except Exception as e:
        print(f"taxonomy keyword template 읽기 실패: {path} | {e}")
        return bundle

    required = {"taxonomy_type", "canonical_name", "aliases"}
    if not required.issubset(df.columns):
        print(f"taxonomy keyword template 컬럼 부족: {df.columns.tolist()}")
        return bundle

    for _, row in df.iterrows():
        tax_type = _clean_cell(row.get("taxonomy_type", "")).lower()
        if tax_type not in bundle:
            continue
        canonical = _clean_cell(row.get("canonical_name", ""))
        aliases = [_clean_cell(x) for x in _clean_cell(row.get("aliases", "")).split("|")]
        aliases.append(canonical)
        for alias in aliases:
            _add_alias(bundle[tax_type], alias, canonical)

    return bundle


def _load_taxo_dic(path):
    bundle = _empty_taxonomy_bundle()
    if not path.exists():
        return bundle
    try:
        df = pd.read_excel(path, sheet_name=1)
    except Exception as e:
        print(f"taxo_dic 읽기 실패: {path} | {e}")
        return bundle

    required = {"canonical_term", "synonym", "term_type"}
    if not required.issubset(df.columns):
        print(f"taxo_dic 컬럼 부족: {df.columns.tolist()}")
        return bundle

    if "exclude_flag" in df.columns:
        df = df[df["exclude_flag"].fillna(0).astype(str).str.strip().isin(["", "0", "False", "false"])]
    if "apply_scope" in df.columns:
        scope = df["apply_scope"].fillna("").astype(str).str.upper()
        df = df[(scope == "") | (scope.str.contains("JD_CV", regex=False))]

    type_map = {
        "skill": "skill",
        "task": "task",
        "domain": "domain",
        "role": "role",
        "job_role": "role",
    }

    for _, row in df.iterrows():
        raw_type = _clean_cell(row.get("term_type", "")).lower()
        tax_type = type_map.get(raw_type)
        if tax_type not in bundle:
            continue
        canonical = _clean_cell(row.get("canonical_term", ""))
        synonym = _clean_cell(row.get("synonym", ""))
        normalized = _clean_cell(row.get("normalized_key", ""))
        _add_alias(bundle[tax_type], synonym, canonical)
        _add_alias(bundle[tax_type], normalized, canonical)
        _add_alias(bundle[tax_type], canonical, canonical)

    return bundle


base_bundle = {
    "skill": DEFAULT_SKILL_TAXONOMY,
    "task": DEFAULT_TASK_TAXONOMY,
    "domain": DEFAULT_DOMAIN_TAXONOMY,
    "role": DEFAULT_ROLE_TAXONOMY,
}

external_bundle = _empty_taxonomy_bundle()
TAXONOMY_SOURCE = "internal_default"

if USE_EXTERNAL_TAXONOMY:
    reviewed_bundle = _load_reviewed_taxonomy_files(TAXONOMY_DIR)
    print("검수 taxonomy 로드 수:", _taxonomy_counts(reviewed_bundle))

    # data/11_taxonomy 검수본만 active taxonomy source로 사용합니다.
    external_bundle = reviewed_bundle

    if sum(_taxonomy_counts(external_bundle).values()) > 0:
        TAXONOMY_SOURCE = "external_files"
        taxonomy_bundle = _merge_taxonomy_bundle(base_bundle, external_bundle)
    else:
        print("외부 taxonomy가 비어 있어 내부 기본 dictionary만 사용합니다.")
        taxonomy_bundle = base_bundle
else:
    taxonomy_bundle = base_bundle

SKILL_TAXONOMY = taxonomy_bundle["skill"]
TASK_TAXONOMY = taxonomy_bundle["task"]
DOMAIN_TAXONOMY = taxonomy_bundle["domain"]
ROLE_TAXONOMY = taxonomy_bundle["role"]

print("Taxonomy 로드 완료")
print("Taxonomy source:", TAXONOMY_SOURCE)
print("Skill terms:", len(SKILL_TAXONOMY))
print("Task terms:", len(TASK_TAXONOMY))
print("Domain terms:", len(DOMAIN_TAXONOMY))
print("Role terms:", len(ROLE_TAXONOMY))




# In[7]:


# Cell 06 | 컬럼별 가중치 설정
# ------------------------------------------------------------
# 가중치는 두 단계로 분리합니다.
#
# 1) RETRIEVAL_COLUMN_WEIGHTS
#    - 1차 후보 추림용 가중치입니다.
#    - Skill-only로 자르면 Career_Description에만 드러나는 RAG/MCP/Agent 업무 경험을 놓칠 수 있으므로
#      Skill-heavy + Task 보완 구조로 둡니다.
#
# 2) WEIGHTS
#    - Top-N 후보에 대한 정밀 점수(column_score) 내부 가중치입니다.
#    - 최종 final_score는 retrieval_score_norm과 column_score를 다시 섞습니다.
# ------------------------------------------------------------
USE_EXTERNAL_WEIGHTS = False

# 1차 Retrieval용: 전체 855명에서 Top-N 후보를 추리는 깔대기 단계
RETRIEVAL_COLUMN_WEIGHTS = {
    "skill": 0.60,
    "task": 0.25,
    "role": 0.10,
    "domain": 0.05,
}

# 최종 Ranking용: Top-N 후보의 skill/task/domain/role/AI tag 정밀 점수 합산
WEIGHTS = {
    "skill_score": 0.35,
    "task_score": 0.25,
    "domain_score": 0.10,
    "role_score": 0.15,
    "ai_skill_overlap_score": 0.15,
}

# ------------------------------------------------------------
# 추후 외부 가중치 Excel 연결 예시
# matching_weights_jo.xlsx 예시:
# score_component | weight
# skill_score     | 0.35
# task_score      | 0.25
# domain_score    | 0.10
# role_score      | 0.15
# ai_skill_overlap_score | 0.15
# ------------------------------------------------------------
# if USE_EXTERNAL_WEIGHTS:
#     weights_df = pd.read_excel("data/12_config/matching_weights_jo.xlsx")
#     WEIGHTS = dict(zip(weights_df["score_component"], weights_df["weight"]))

retrieval_weight_sum = sum(RETRIEVAL_COLUMN_WEIGHTS.values())
if abs(retrieval_weight_sum - 1.0) > 1e-6:
    print("경고: 1차 Retrieval 컬럼 가중치 합이 1이 아닙니다.", retrieval_weight_sum)
else:
    print("1차 Retrieval 컬럼 가중치 설정 완료:", RETRIEVAL_COLUMN_WEIGHTS)

weight_sum = sum(WEIGHTS.values())
if abs(weight_sum - 1.0) > 1e-6:
    print("경고: 컬럼 내부 가중치 합이 1이 아닙니다.", weight_sum)
else:
    print("컬럼 내부 가중치 설정 완료:", WEIGHTS)

if abs((RETRIEVAL_SCORE_WEIGHT + COLUMN_SCORE_WEIGHT) - 1.0) > 1e-6:
    print("경고: retrieval/column 최종 가중치 합이 1이 아닙니다.", RETRIEVAL_SCORE_WEIGHT + COLUMN_SCORE_WEIGHT)
else:
    print("최종 점수 가중치 설정 완료:", {
        "retrieval_score_norm": RETRIEVAL_SCORE_WEIGHT,
        "column_score": COLUMN_SCORE_WEIGHT,
    })


# In[8]:


# Cell 07 | 텍스트 정리 / Taxonomy 적용 함수

def clean_text(text):
    """텍스트 결측/공백/줄바꿈을 정리합니다."""
    text = "" if text is None else str(text)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def safe_get(row, col):
    """row에서 컬럼이 없거나 NaN이면 빈 문자열 반환."""
    return str(row.get(col, "") or "")


def _keyword_pattern(raw):
    raw = str(raw).strip().lower()
    if not raw:
        return None
    escaped = re.escape(raw)
    # 영어/숫자 키워드는 부분 문자열 오매칭을 줄이기 위해 경계를 둡니다.
    if re.fullmatch(r"[a-z0-9][a-z0-9\+\#\.\- ]*", raw):
        return rf"(?<![a-z0-9]){escaped}(?![a-z0-9])"
    return escaped


def apply_taxonomy(text, taxonomy_dict):
    """
    taxonomy alias를 찾아 표준어 목록으로 변환합니다.
    - 영어 키워드는 word-boundary 성격의 패턴으로 매칭합니다.
    - 긴 alias를 먼저 검사해 짧은 단어의 과매칭을 줄입니다.
    - 불용어 사전 기준 noise alias/standard는 결과에서 제외합니다.
    """
    text = clean_text(text)
    text_lower = text.lower()
    found = []
    for raw_term, standard_term in sorted(taxonomy_dict.items(), key=lambda x: len(str(x[0])), reverse=True):
        raw = str(raw_term).lower().strip()
        if _is_taxonomy_noise_term(raw):
            continue
        pattern = _keyword_pattern(raw)
        if pattern and re.search(pattern, text_lower):
            standard = str(standard_term).strip()
            if standard and not _is_taxonomy_noise_term(standard):
                found.append(standard)
    return ", ".join(sorted(set(found)))

print("함수 정의 완료")



# In[9]:


# Cell 08 | 필수 컬럼 보정
# ------------------------------------------------------------
# 업로드된 JD/CV 파일의 실제 컬럼명을 기준으로 설계했습니다.
# 컬럼이 없는 경우 빈 문자열 컬럼을 생성해 노트북이 중단되지 않게 합니다.
# ------------------------------------------------------------

jd_required_cols = [
    "jd_id", "company", "job_title", "career_requirement", "required_skill", "main_task",
    "role_signal", "job_family", "salary", "education_requirement", "qualification", "preferred",
    "domain", "deliverable_signal", "project_signal", "location", "employment_type", "raw_jd", "source", "clean_text"
]
cv_required_cols = [
    "candidate_id", "job_role", "resume_no", "file_name", "Education", "GPA", "Certificates",
    "Languages", "Paper", "Skill", "Overseas_Experience", "Training", "Awards", "Total_Career_Years",
    "Career", "Career_Description", "Position", "End_Date", "Employment_Status", "Salary",
    "Self_Introduction", "Portfolio", "Raw_Text"
]

for col in jd_required_cols:
    if col not in jd_df.columns:
        jd_df[col] = ""

for col in cv_required_cols:
    if col not in cv_df.columns:
        cv_df[col] = ""

# jd_dataset_613_filled.csv에는 clean_text가 없을 수 있으므로,
# raw_jd / qualification / preferred를 합쳐 매칭 본문을 보강합니다.
clean_text_blank = jd_df["clean_text"].fillna("").astype(str).str.strip().eq("")
jd_df.loc[clean_text_blank, "clean_text"] = (
    jd_df.loc[clean_text_blank, "raw_jd"].fillna("").astype(str)
    + "\n" +
    jd_df.loc[clean_text_blank, "qualification"].fillna("").astype(str)
    + "\n" +
    jd_df.loc[clean_text_blank, "preferred"].fillna("").astype(str)
).map(clean_text)

print("필수 컬럼 보정 완료")
print("JD clean_text non-empty:", int(jd_df["clean_text"].fillna("").astype(str).str.len().gt(0).sum()))



# In[10]:


# Cell 09 | CV/JD 표준화 컬럼 생성
# ------------------------------------------------------------
# 이 단계에서 이전 대화에서 말한 파생 컬럼을 자동 생성합니다.
# 수동으로 cv_skill_standard 등을 만들 필요 없습니다.
# ------------------------------------------------------------

# CV 기준
# - Skill 컬럼 → cv_skill_standard
# - Career + Career_Description → cv_task_standard
# - Career + Career_Description → cv_domain_standard
# - Position + Career → cv_role_standard
cv_df["cv_skill_standard"] = cv_df["Skill"].apply(lambda x: apply_taxonomy(x, SKILL_TAXONOMY))
cv_df["cv_task_standard"] = (cv_df["Career"] + "\n" + cv_df["Career_Description"]).apply(lambda x: apply_taxonomy(x, TASK_TAXONOMY))
cv_df["cv_domain_standard"] = (cv_df["Career"] + "\n" + cv_df["Career_Description"] + "\n" + cv_df["job_role"]).apply(lambda x: apply_taxonomy(x, DOMAIN_TAXONOMY))
cv_df["cv_role_standard"] = (cv_df["Position"] + "\n" + cv_df["Career"] + "\n" + cv_df["job_role"]).apply(lambda x: apply_taxonomy(x, ROLE_TAXONOMY))

# JD 기준
# - required_skill → jd_skill_standard
# - main_task + clean_text → jd_task_standard
# - domain + company + clean_text → jd_domain_standard
# - role_signal + job_title + job_family → jd_role_standard
jd_df["jd_skill_standard"] = jd_df["required_skill"].apply(lambda x: apply_taxonomy(x, SKILL_TAXONOMY))
jd_df["jd_task_standard"] = (jd_df["main_task"] + "\n" + jd_df["clean_text"]).apply(lambda x: apply_taxonomy(x, TASK_TAXONOMY))
jd_df["jd_domain_standard"] = (jd_df["domain"] + "\n" + jd_df["company"] + "\n" + jd_df["clean_text"]).apply(lambda x: apply_taxonomy(x, DOMAIN_TAXONOMY))
jd_df["jd_role_standard"] = (jd_df["role_signal"] + "\n" + jd_df["job_title"] + "\n" + jd_df["job_family"]).apply(lambda x: apply_taxonomy(x, ROLE_TAXONOMY))

standard_cols_cv = ["candidate_id", "Skill", "cv_skill_standard", "cv_task_standard", "cv_domain_standard", "cv_role_standard"]
standard_cols_jd = ["jd_id", "job_title", "required_skill", "main_task", "domain", "role_signal", "jd_skill_standard", "jd_task_standard", "jd_domain_standard", "jd_role_standard"]

display(cv_df[standard_cols_cv].head(5))
display(jd_df[standard_cols_jd].head(5))


# ## 2. 매칭용 텍스트 생성
# 
# ### Retrieval용 full_text
# 
# 1차 검색에서는 **CV/JD 대표 텍스트 1개**를 만들어 사용합니다.
# 
# CV full text에 들어가는 항목:
# 
# ```text
# Skill
# Career
# Career_Description
# Position
# cv_skill_standard
# cv_task_standard
# cv_domain_standard
# cv_role_standard
# ```
# 
# JD full text에 들어가는 항목:
# 
# ```text
# required_skill
# main_task
# domain
# role_signal
# job_title
# clean_text
# jd_skill_standard
# jd_task_standard
# jd_domain_standard
# jd_role_standard
# ```
# 
# ### 정밀 점수용 컬럼별 text
# 
# Top-N 후보에 대해서만 아래 pair를 비교합니다.
# 
# ```text
# Skill:  JD jd_skill_text   ↔ CV cv_skill_text
# Task:   JD jd_task_text    ↔ CV cv_task_text
# Domain: JD jd_domain_text  ↔ CV cv_domain_text
# Role:   JD jd_role_text    ↔ CV cv_role_text
# ```
# 
# Cross Encoder를 붙일 때도 위 pair가 그대로 input type이 됩니다.
# 

# In[11]:


# Cell 10 | 매칭용 텍스트 생성

def build_cv_texts(row):
    """
    CV 매칭 텍스트 구성.

    cv_full_text는 1차 Retrieval용입니다.
    cv_skill/task/domain/role_text는 Top-N 후보에 대한 정밀 점수용입니다.
    """
    return pd.Series({
        # Skill 정밀 비교용: 스킬 원문 + 표준화 스킬
        "cv_skill_text": clean_text(f"{safe_get(row, 'Skill')}\n{safe_get(row, 'cv_skill_standard')}"),

        # Task 정밀 비교용: 경력/프로젝트 설명 + task taxonomy
        "cv_task_text": clean_text(f"{safe_get(row, 'Career')}\n{safe_get(row, 'Career_Description')}\n{safe_get(row, 'cv_task_standard')}"),

        # Domain 정밀 비교용: 회사/경력/프로젝트 설명 + domain taxonomy
        "cv_domain_text": clean_text(f"{safe_get(row, 'Career')}\n{safe_get(row, 'Career_Description')}\n{safe_get(row, 'job_role')}\n{safe_get(row, 'cv_domain_standard')}"),

        # Role 정밀 비교용: 희망/현재 직무, 포지션, 경력 + role taxonomy
        "cv_role_text": clean_text(f"{safe_get(row, 'job_role')}\n{safe_get(row, 'Position')}\n{safe_get(row, 'Career')}\n{safe_get(row, 'cv_role_standard')}"),

        # 1차 Retrieval용 대표 텍스트
        "cv_full_text": clean_text(f"""
        [Skill]
        {safe_get(row, 'Skill')}
        {safe_get(row, 'cv_skill_standard')}

        [Career]
        {safe_get(row, 'Career')}

        [Career Description]
        {safe_get(row, 'Career_Description')}

        [Position]
        {safe_get(row, 'Position')}

        [Standard Taxonomy]
        Skill: {safe_get(row, 'cv_skill_standard')}
        Task: {safe_get(row, 'cv_task_standard')}
        Domain: {safe_get(row, 'cv_domain_standard')}
        Role: {safe_get(row, 'cv_role_standard')}
        """)
    })


def build_jd_texts(row):
    """
    JD 매칭 텍스트 구성.

    jd_full_text는 1차 Retrieval용입니다.
    jd_skill/task/domain/role_text는 Top-N 후보에 대한 정밀 점수용입니다.
    """
    return pd.Series({
        # Skill 정밀 비교용: JD 요구 스킬 + 표준화 스킬
        "jd_skill_text": clean_text(f"{safe_get(row, 'required_skill')}\n{safe_get(row, 'jd_skill_standard')}"),

        # Task 정밀 비교용: JD 주요 업무 + 정제 본문 + task taxonomy
        "jd_task_text": clean_text(f"{safe_get(row, 'main_task')}\n{safe_get(row, 'clean_text')}\n{safe_get(row, 'jd_task_standard')}"),

        # Domain 정밀 비교용: JD 도메인/회사/본문 + domain taxonomy
        "jd_domain_text": clean_text(f"{safe_get(row, 'domain')}\n{safe_get(row, 'company')}\n{safe_get(row, 'clean_text')}\n{safe_get(row, 'jd_domain_standard')}"),

        # Role 정밀 비교용: JD 타이틀/role signal/job family + role taxonomy
        "jd_role_text": clean_text(f"{safe_get(row, 'job_title')}\n{safe_get(row, 'role_signal')}\n{safe_get(row, 'job_family')}\n{safe_get(row, 'jd_role_standard')}"),

        # 1차 Retrieval용 대표 텍스트
        "jd_full_text": clean_text(f"""
        [Job Title]
        {safe_get(row, 'job_title')}

        [Required Skill]
        {safe_get(row, 'required_skill')}
        {safe_get(row, 'jd_skill_standard')}

        [Main Task]
        {safe_get(row, 'main_task')}
        {safe_get(row, 'jd_task_standard')}

        [Domain / Role]
        Domain: {safe_get(row, 'domain')}
        Role: {safe_get(row, 'role_signal')}
        Job Family: {safe_get(row, 'job_family')}
        Standard Domain: {safe_get(row, 'jd_domain_standard')}
        Standard Role: {safe_get(row, 'jd_role_standard')}

        [Clean JD Text]
        {safe_get(row, 'clean_text')}
        """)
    })

cv_text_df = cv_df.apply(build_cv_texts, axis=1)
jd_text_df = jd_df.apply(build_jd_texts, axis=1)

cv_df = pd.concat([cv_df, cv_text_df], axis=1)
jd_df = pd.concat([jd_df, jd_text_df], axis=1)

print("매칭용 텍스트 생성 완료")
display(cv_df[["candidate_id", "cv_full_text", "cv_skill_text", "cv_task_text"]].head(2))
display(jd_df[["jd_id", "jd_full_text", "jd_skill_text", "jd_task_text"]].head(2))

# ------------------------------------------------------------
# ==================================================
# AI/LLM/RAG 보정 태그 생성
# ==================================================

AI_SKILL_ALIASES = {
    "LLM": ["llm", "large language model", "gpt", "openai", "gemini", "claude", "생성형ai", "생성형 ai"],
    "RAG": ["rag", "retrieval augmented generation", "검색증강생성", "검색 증강"],
    "AI Agent": ["ai agent", "agent", "에이전트", "langgraph", "langchain agent"],
    "Machine Learning": ["machine learning", "머신러닝", "ml", "예측모델", "예측 모델"],
    "Deep Learning": ["deep learning", "딥러닝", "pytorch", "tensorflow"],
    "Backend/API": ["fastapi", "api", "restapi", "rest api", "서버", "백엔드", "backend"],
    "Cloud": ["aws", "gcp", "google cloud", "vertex ai", "cloud"],
    "MLOps": ["mlops", "docker", "kubernetes", "kubeflow", "airflow", "model serving", "모델 서빙"],
}

def normalize_ai_skill_tags(text):
    text = str(text).lower()
    tags = []
    for standard, aliases in AI_SKILL_ALIASES.items():
        if any(alias.lower() in text for alias in aliases):
            tags.append(standard)
    return ", ".join(sorted(set(tags)))

jd_tag_cols = [
    "required_skill",
    "main_task",
    "role_signal",
    "job_family",
    "domain",
    "deliverable_signal",
    "project_signal",
    "qualification",
    "preferred",
    "raw_jd",
]

cv_tag_cols = [
    "Skill",
    "Career",
    "Career_Description",
    "Position",
    "Self_Introduction",
    "Raw_Text",
]

jd_tag_cols = [c for c in jd_tag_cols if c in jd_df.columns]
cv_tag_cols = [c for c in cv_tag_cols if c in cv_df.columns]

jd_df["jd_ai_skill_tags"] = (
    jd_df[jd_tag_cols]
    .fillna("")
    .astype(str)
    .agg(" ".join, axis=1)
    .map(normalize_ai_skill_tags)
)

cv_df["cv_ai_skill_tags"] = (
    cv_df[cv_tag_cols]
    .fillna("")
    .astype(str)
    .agg(" ".join, axis=1)
    .map(normalize_ai_skill_tags)
)

print("AI/LLM/RAG 보정 태그 생성 완료")

display(jd_df[["jd_id", "job_title", "jd_ai_skill_tags"]])
display(cv_df[["candidate_id", "cv_ai_skill_tags"]].head(10))


# In[12]:


# Cell 11 | 1차 Retrieval - BM25 + Weighted TF-IDF keyword baseline
# ------------------------------------------------------------
# v22 핵심 변경:
# - 외부 패키지 없이 BM25를 직접 구현합니다.
# - TF-IDF 점수와 BM25 점수를 각각 계산합니다.
# - baseline은 두 점수의 평균인 keyword_baseline_score를 사용합니다.
# - embedding retrieval을 사용하더라도 비교 평가를 위해 KEYWORD_SCORE_DF를 보존합니다.
# ------------------------------------------------------------

import math
from collections import Counter


def bm25_tokenize(text):
    """BM25용 간단 토크나이저. 한글/영문/숫자 토큰을 보존합니다."""
    text = "" if pd.isna(text) else str(text).lower()
    return re.findall(r"[가-힣a-zA-Z0-9_+#./-]+", text)


def minmax_array(values):
    arr = np.asarray(values, dtype=float)
    if arr.size == 0:
        return arr
    vmin = np.nanmin(arr)
    vmax = np.nanmax(arr)
    if not np.isfinite(vmin) or not np.isfinite(vmax) or vmax == vmin:
        return np.zeros_like(arr, dtype=float)
    return (arr - vmin) / (vmax - vmin)


def bm25_scores_for_one_jd(jd_text, cv_texts, k1=1.5, b=0.75):
    """JD 1건과 CV 여러 건 사이의 BM25 점수를 계산합니다."""
    query_tokens = bm25_tokenize(jd_text)
    docs = [bm25_tokenize(x) for x in list(cv_texts)]
    n_docs = len(docs)
    if n_docs == 0 or not query_tokens:
        return np.zeros(n_docs, dtype=float)

    doc_lens = np.asarray([len(d) for d in docs], dtype=float)
    avgdl = float(doc_lens.mean()) if doc_lens.size else 0.0
    if avgdl <= 0:
        return np.zeros(n_docs, dtype=float)

    term_freqs = [Counter(d) for d in docs]
    doc_freq = Counter()
    for tf in term_freqs:
        for term in tf.keys():
            doc_freq[term] += 1

    scores = np.zeros(n_docs, dtype=float)
    for term in query_tokens:
        df = doc_freq.get(term, 0)
        if df == 0:
            continue
        # Robertson/Sparck Jones IDF with smoothing
        idf = math.log(1 + (n_docs - df + 0.5) / (df + 0.5))
        for i, tf in enumerate(term_freqs):
            freq = tf.get(term, 0)
            if freq == 0:
                continue
            denom = freq + k1 * (1 - b + b * doc_lens[i] / avgdl)
            scores[i] += idf * (freq * (k1 + 1)) / denom
    return scores


def tfidf_scores_for_one_jd(jd_text, cv_texts):
    corpus = ["" if pd.isna(jd_text) else str(jd_text)] + ["" if pd.isna(x) else str(x) for x in list(cv_texts)]
    vectorizer = TfidfVectorizer(lowercase=True, ngram_range=(1, 2), min_df=1, max_features=20000)
    mat = vectorizer.fit_transform(corpus)
    return cosine_similarity(mat[0:1], mat[1:]).flatten()


def weighted_keyword_scores_for_one_jd(jd_row, cv_source_df, weights=None):
    """JD 1건과 전체 CV를 skill/task/role/domain 축으로 비교해 TF-IDF, BM25, baseline 점수를 만듭니다."""
    if weights is None:
        weights = globals().get("RETRIEVAL_COLUMN_WEIGHTS", {
            "skill": 0.60,
            "task": 0.25,
            "role": 0.10,
            "domain": 0.05,
        })

    pair_specs = {
        "skill": ("jd_skill_text", "cv_skill_text"),
        "task": ("jd_task_text", "cv_task_text"),
        "role": ("jd_role_text", "cv_role_text"),
        "domain": ("jd_domain_text", "cv_domain_text"),
    }

    tfidf_final = np.zeros(len(cv_source_df), dtype=float)
    bm25_final = np.zeros(len(cv_source_df), dtype=float)
    score_map = {}

    for name, (jd_col, cv_col) in pair_specs.items():
        jd_text = jd_row.get(jd_col, "")
        cv_texts = cv_source_df[cv_col] if cv_col in cv_source_df.columns else pd.Series([""] * len(cv_source_df))

        tfidf_scores = tfidf_scores_for_one_jd(jd_text, cv_texts)
        bm25_scores = bm25_scores_for_one_jd(jd_text, cv_texts)

        # 축별 스케일 차이를 줄인 뒤 가중합합니다.
        tfidf_scores_norm = minmax_array(tfidf_scores)
        bm25_scores_norm = minmax_array(bm25_scores)

        w = float(weights.get(name, 0.0))
        tfidf_final += w * tfidf_scores_norm
        bm25_final += w * bm25_scores_norm

        score_map[f"retrieval_{name}_tfidf_score"] = tfidf_scores_norm
        score_map[f"retrieval_{name}_bm25_score"] = bm25_scores_norm

    keyword_baseline = (tfidf_final + bm25_final) / 2.0
    return keyword_baseline, tfidf_final, bm25_final, score_map


keyword_rows = []
retrieval_rows = []

print("BM25 + TF-IDF keyword baseline 계산 시작")
print("대상 JD 수:", len(jd_df), "| CV 수:", len(cv_df))

for jd_pos, (_, jd) in enumerate(jd_df.iterrows(), start=1):
    jd_id = str(jd["jd_id"])
    if jd_pos == 1 or jd_pos % 10 == 0 or jd_pos == len(jd_df):
        print(f"[{jd_pos}/{len(jd_df)}] keyword scoring JD={jd_id}")

    keyword_scores, tfidf_scores, bm25_scores, component_scores = weighted_keyword_scores_for_one_jd(
        jd,
        cv_df,
        RETRIEVAL_COLUMN_WEIGHTS
    )

    temp = pd.DataFrame({
        "jd_id": jd_id,
        "candidate_id": cv_df["candidate_id"].astype(str).values,
        "retrieval_score": keyword_scores,
        "keyword_baseline_score": keyword_scores,
        "retrieval_tfidf_score": tfidf_scores,
        "retrieval_bm25_score": bm25_scores,
        "retrieval_method": "bm25_tfidf_average_column_pairs",
    })

    for col, vals in component_scores.items():
        temp[col] = vals

    keyword_rows.append(temp)
    retrieval_rows.append(temp.sort_values("keyword_baseline_score", ascending=False).head(TOP_N_RETRIEVAL))

KEYWORD_SCORE_DF = pd.concat(keyword_rows, ignore_index=True)
retrieval_candidates_df = pd.concat(retrieval_rows, ignore_index=True)
KEYWORD_RETRIEVAL_CANDIDATES_DF = retrieval_candidates_df.copy()

print("KEYWORD_SCORE_DF:", KEYWORD_SCORE_DF.shape)
print("retrieval_candidates_df:", retrieval_candidates_df.shape)
print("RETRIEVAL_COLUMN_WEIGHTS:", RETRIEVAL_COLUMN_WEIGHTS)

display_cols = [
    "jd_id", "candidate_id",
    "keyword_baseline_score", "retrieval_tfidf_score", "retrieval_bm25_score",
    "retrieval_method",
]
display(retrieval_candidates_df[[c for c in display_cols if c in retrieval_candidates_df.columns]].head(10))


# In[13]:


# Cell 12 | 1차 Retrieval - Embedding optional
# ------------------------------------------------------------
# USE_EMBEDDING_RETRIEVAL=True이면 embedding retrieval을 사용합니다.
# v22 변경:
# - 새 CV embedding 파일 cv_embeddings_only_bge_m3_10.npz를 우선 사용합니다.
# - embedding 후보군에도 BM25/TF-IDF keyword 점수를 병합합니다.
# - retrieval_score는 embedding 점수로 유지하되, final_score에는 keyword_baseline_score도 별도 반영됩니다.
# ------------------------------------------------------------


def _l2_normalize(mat):
    mat = np.asarray(mat, dtype="float32")
    denom = np.linalg.norm(mat, axis=1, keepdims=True)
    denom[denom == 0] = 1.0
    return mat / denom


def _to_str_list(arr):
    try:
        return [str(x) for x in arr.tolist()]
    except Exception:
        return [str(x) for x in list(arr)]


def _load_precomputed_embeddings(path):
    emb = np.load(path, allow_pickle=True)
    files = list(emb.files)
    print("precomputed embedding file:", path)
    print("precomputed embedding keys:", files)

    cv_embeddings = emb["cv_embeddings"] if "cv_embeddings" in files else None
    jd_embeddings = emb["jd_embeddings"] if "jd_embeddings" in files else None

    # v2: CV npz에는 JD embedding이 없으므로 별도 JD .npy 파일을 연결합니다.
    jd_embedding_path = globals().get("JD_EMBEDDING_PATH", None)
    if jd_embeddings is None and jd_embedding_path is not None and Path(jd_embedding_path).exists():
        jd_embeddings = np.load(jd_embedding_path, allow_pickle=True)
        print("jd_embeddings 별도 .npy 사용:", jd_embedding_path)

    candidate_ids = _to_str_list(emb["candidate_ids"]) if "candidate_ids" in files else None
    jd_ids = _to_str_list(emb["jd_ids"]) if "jd_ids" in files else None

    # 일부 파일은 CV embedding만 다른 키 이름으로 저장됐을 수 있어 보수적으로 탐색합니다.
    if cv_embeddings is None:
        for key in ["embeddings", "cv_embedding", "candidate_embeddings"]:
            if key in files:
                cv_embeddings = emb[key]
                print(f"cv_embeddings 대체 key 사용: {key}")
                break

    if cv_embeddings is not None:
        cv_embeddings = _l2_normalize(cv_embeddings)
        print("cv_embeddings:", cv_embeddings.shape)
    if jd_embeddings is not None:
        jd_embeddings = _l2_normalize(jd_embeddings)
        print("jd_embeddings:", jd_embeddings.shape)

    return {
        "path": Path(path),
        "cv_embeddings": cv_embeddings,
        "jd_embeddings": jd_embeddings,
        "candidate_ids": candidate_ids,
        "jd_ids": jd_ids,
    }


def _get_cv_embedding_matrix_for_current_cv(pre):
    """
    현재 cv_df 순서에 맞춰 CV embedding matrix를 재정렬합니다.
    candidate_ids가 있으면 ID 기준으로 매핑하고,
    없으면 cv_df와 npz가 같은 순서라고 가정합니다.
    """
    cv_emb = pre["cv_embeddings"]
    if cv_emb is None:
        return None

    if pre["candidate_ids"] is None:
        if len(cv_emb) < len(cv_df):
            raise ValueError("candidate_ids가 없고, cv_embeddings 수가 현재 cv_df보다 적습니다.")
        print("candidate_ids가 없어 현재 cv_df 순서와 embedding 순서가 같다고 가정합니다.")
        return cv_emb[:len(cv_df)]

    id_to_pos = {str(cid): i for i, cid in enumerate(pre["candidate_ids"])}
    current_ids = cv_df["candidate_id"].astype(str).tolist()
    missing = [cid for cid in current_ids if cid not in id_to_pos]
    if missing:
        raise ValueError(f"precomputed embedding에 없는 candidate_id가 있습니다. 예: {missing[:5]}")

    positions = [id_to_pos[cid] for cid in current_ids]
    return cv_emb[positions]


def _get_precomputed_jd_vector(pre, jd_id, jd_position):
    jd_emb = pre["jd_embeddings"]
    if jd_emb is None:
        return None

    if pre["jd_ids"] is not None:
        id_to_pos = {str(jid): i for i, jid in enumerate(pre["jd_ids"])}
        pos = id_to_pos.get(str(jd_id))
        if pos is not None:
            return jd_emb[pos:pos+1]

    if len(jd_emb) == len(jd_df):
        return jd_emb[jd_position:jd_position+1]

    if len(jd_emb) == 1 and len(jd_df) == 1:
        return jd_emb[0:1]

    return None


def _select_precomputed_embedding():
    """후보 embedding 파일 중 현재 cv_df와 호환되는 첫 번째 파일을 선택합니다."""
    candidate_paths = globals().get("PRECOMPUTED_EMBEDDING_CANDIDATE_PATHS", [globals().get("PRECOMPUTED_EMBEDDING_PATH")])
    for path in candidate_paths:
        if path is None:
            continue
        path = Path(path)
        if not path.exists():
            print("embedding 후보 파일 없음:", path)
            continue
        try:
            pre = _load_precomputed_embeddings(path)
            cv_matrix = _get_cv_embedding_matrix_for_current_cv(pre)
            if cv_matrix is None:
                raise ValueError("CV embedding matrix가 없습니다.")
            print("선택된 embedding 파일:", path)
            return pre, cv_matrix, path
        except Exception as e:
            print("embedding 후보 파일 사용 불가:", path)
            print("에러:", e)
    return None, None, None


def _merge_keyword_scores(candidate_df):
    """embedding 후보군에 BM25/TF-IDF keyword 점수를 병합합니다."""
    if "KEYWORD_SCORE_DF" not in globals():
        print("KEYWORD_SCORE_DF가 없어 keyword 점수 병합을 건너뜁니다.")
        return candidate_df

    keyword_cols = [
        "jd_id", "candidate_id",
        "keyword_baseline_score", "retrieval_tfidf_score", "retrieval_bm25_score",
    ]
    optional_cols = [c for c in KEYWORD_SCORE_DF.columns if c.startswith("retrieval_") and (c.endswith("_tfidf_score") or c.endswith("_bm25_score"))]
    keyword_cols = list(dict.fromkeys(keyword_cols + optional_cols))

    out = candidate_df.copy()
    out["jd_id"] = out["jd_id"].astype(str)
    out["candidate_id"] = out["candidate_id"].astype(str)
    key_df = KEYWORD_SCORE_DF[[c for c in keyword_cols if c in KEYWORD_SCORE_DF.columns]].copy()
    key_df["jd_id"] = key_df["jd_id"].astype(str)
    key_df["candidate_id"] = key_df["candidate_id"].astype(str)

    # 동일 이름 컬럼이 있으면 candidate_df의 값을 보존하지 않고 keyword 계산값으로 갱신합니다.
    duplicate_cols = [c for c in key_df.columns if c in out.columns and c not in ["jd_id", "candidate_id"]]
    if duplicate_cols:
        out = out.drop(columns=duplicate_cols)

    out = out.merge(key_df, on=["jd_id", "candidate_id"], how="left")
    for c in ["keyword_baseline_score", "retrieval_tfidf_score", "retrieval_bm25_score"]:
        if c in out.columns:
            out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0.0)
    return out


if USE_EMBEDDING_RETRIEVAL:
    emb_rows = []
    used_embedding = False
    embedding_model = None

    pre = None
    cv_embedding_matrix = None
    selected_path = None

    if USE_PRECOMPUTED_EMBEDDING:
        pre, cv_embedding_matrix, selected_path = _select_precomputed_embedding()
        if cv_embedding_matrix is not None:
            used_embedding = True
            SELECTED_PRECOMPUTED_EMBEDDING_PATH = selected_path
            print("precomputed CV embedding 연결 완료:", cv_embedding_matrix.shape)

    # precomputed CV가 없으면 SentenceTransformer로 CV/JD 모두 새로 encode
    if cv_embedding_matrix is None:
        try:
            from sentence_transformers import SentenceTransformer
            embedding_model = SentenceTransformer(EMBEDDING_MODEL_NAME)
            print("SentenceTransformer CV embedding 신규 생성 시작:", len(cv_df))
            cv_embedding_matrix = embedding_model.encode(
                cv_df["cv_full_text"].tolist(),
                normalize_embeddings=True,
                show_progress_bar=True,
                batch_size=32,
            )
            cv_embedding_matrix = _l2_normalize(cv_embedding_matrix)
            used_embedding = True
            print("SentenceTransformer CV embedding 생성 완료:", cv_embedding_matrix.shape)
        except Exception as e:
            print("Embedding retrieval 사용 실패. BM25+TF-IDF keyword baseline retrieval 결과를 유지합니다.")
            print("설치 예시: pip install sentence-transformers")
            print("에러:", e)
            used_embedding = False

    # JD별 embedding retrieval
    if used_embedding and cv_embedding_matrix is not None:
        print("Embedding retrieval 시작")
        print("대상 JD 수:", len(jd_df), "| CV embedding shape:", cv_embedding_matrix.shape)
        for j_pos, jd in jd_df.reset_index(drop=True).iterrows():
            jd_id = str(jd["jd_id"])
            print(f"[{j_pos + 1}/{len(jd_df)}] embedding retrieval JD={jd_id}")
            jd_vec = _get_precomputed_jd_vector(pre, jd_id, j_pos) if pre is not None else None

            # precomputed JD embedding이 현재 JD를 커버하지 못하면 JD만 새로 encode
            if jd_vec is None:
                try:
                    if embedding_model is None:
                        from sentence_transformers import SentenceTransformer
                        embedding_model = SentenceTransformer(EMBEDDING_MODEL_NAME)
                    jd_vec = embedding_model.encode(
                        [jd["jd_full_text"]],
                        normalize_embeddings=True,
                        show_progress_bar=False,
                        batch_size=1,
                    )
                    jd_vec = _l2_normalize(jd_vec)
                    print(f"{jd_id}: JD embedding 신규 생성")
                except Exception as e:
                    print(f"{jd_id}: JD embedding 없음 + 신규 생성 실패 → BM25+TF-IDF fallback")
                    scores, tfidf_scores, bm25_scores, component_scores = weighted_keyword_scores_for_one_jd(
                        jd,
                        cv_df,
                        RETRIEVAL_COLUMN_WEIGHTS
                    )
                    temp = pd.DataFrame({
                        "jd_id": jd_id,
                        "candidate_id": cv_df["candidate_id"].astype(str).values,
                        "retrieval_score": scores,
                        "keyword_baseline_score": scores,
                        "retrieval_tfidf_score": tfidf_scores,
                        "retrieval_bm25_score": bm25_scores,
                        "retrieval_method": "bm25_tfidf_average_column_fallback",
                    })
                    for col, vals in component_scores.items():
                        temp[col] = vals
                    emb_rows.append(temp.sort_values("retrieval_score", ascending=False).head(TOP_N_RETRIEVAL))
                    continue

            scores = (jd_vec @ cv_embedding_matrix.T).flatten()
            temp = pd.DataFrame({
                "jd_id": jd_id,
                "candidate_id": cv_df["candidate_id"].astype(str).values,
                "retrieval_score": scores,
                "embedding_retrieval_score": scores,
                "retrieval_method": "embedding_full_text_precomputed_or_bge",
            })
            emb_rows.append(temp.sort_values("retrieval_score", ascending=False).head(TOP_N_RETRIEVAL))

        retrieval_candidates_df = pd.concat(emb_rows, ignore_index=True)
        retrieval_candidates_df = _merge_keyword_scores(retrieval_candidates_df)
        print("Embedding retrieval로 retrieval_candidates_df 업데이트 완료")
        print("SELECTED_PRECOMPUTED_EMBEDDING_PATH:", SELECTED_PRECOMPUTED_EMBEDDING_PATH)
        display(retrieval_candidates_df.head(20))
else:
    print("USE_EMBEDDING_RETRIEVAL=False: BM25+TF-IDF keyword baseline retrieval 결과를 사용합니다.")


# In[14]:


# Cell 13 | Top-N 후보 pair table 생성
# ------------------------------------------------------------
# 이제부터는 모든 CV를 비교하지 않고, 1차 Retrieval에서 뽑힌 후보만 비교합니다.
# 이 pair_df가 Cross Encoder 또는 컬럼별 정밀 점수 계산의 입력 테이블입니다.
# ------------------------------------------------------------

pair_df = (
    retrieval_candidates_df
    .merge(jd_df, on="jd_id", how="left")
    .merge(cv_df, on="candidate_id", how="left", suffixes=("_jd", "_cv"))
)

print("pair_df:", pair_df.shape)
pair_display_cols = [
    "jd_id", "candidate_id", "retrieval_score", "embedding_retrieval_score",
    "keyword_baseline_score", "retrieval_tfidf_score", "retrieval_bm25_score",
    "job_title", "job_role", "required_skill", "Skill"
]
pair_display_cols = [c for c in pair_display_cols if c in pair_df.columns]
display(pair_df[pair_display_cols].head(10))


# In[15]:


# Cell 14 | 정밀 점수 함수 - TF-IDF pair scoring
# ------------------------------------------------------------
# Cross Encoder를 쓰지 않는 경우, 컬럼별 pair에 대해 TF-IDF cosine similarity로 대체합니다.
# ------------------------------------------------------------

def tfidf_pair_score(a, b):
    try:
        corpus = [clean_text(a), clean_text(b)]
        if not corpus[0] or not corpus[1]:
            return 0.0
        vectorizer = TfidfVectorizer(lowercase=True, ngram_range=(1, 2), min_df=1)
        mat = vectorizer.fit_transform(corpus)
        return float(cosine_similarity(mat[0:1], mat[1:2]).flatten()[0])
    except Exception:
        return 0.0

print("TF-IDF pair scoring 함수 정의 완료")


# In[16]:


# Cell 15 | Cross Encoder / TF-IDF 비교용 텍스트 컬럼 생성
# ------------------------------------------------------------
# 실제 원본 컬럼명 기준:
#
# JD:
# jd_id, company, job_title, required_skill, main_task,
# domain, role_signal, raw_jd
#
# CV:
# candidate_id, job_role, Skill, Career, Career_Description,
# Position, Raw_Text
#
# 이 셀은 원본 컬럼명을 기준으로 Cross Encoder가 사용할
# jd_skill_text / cv_skill_text 등 비교용 컬럼을 생성합니다.
# ------------------------------------------------------------

def safe_col(df, col):
    if col in df.columns:
        return df[col].fillna("").astype(str)
    else:
        return pd.Series([""] * len(df), index=df.index)


def join_cols(df, cols, sep="\n"):
    available = [c for c in cols if c in df.columns]
    if not available:
        return pd.Series([""] * len(df), index=df.index)
    return df[available].fillna("").astype(str).agg(sep.join, axis=1)


# ==================================================
# 1. JD 전체 텍스트
# ==================================================

jd_df["jd_full_text"] = join_cols(
    jd_df,
    [
        "job_title",
        "required_skill",
        "main_task",
        "domain",
        "role_signal",
        "raw_jd",
    ],
    sep="\n"
)


# ==================================================
# 2. CV 전체 텍스트
# ==================================================

cv_df["cv_full_text"] = join_cols(
    cv_df,
    [
        "job_role",
        "Skill",
        "Career",
        "Career_Description",
        "Position",
        "Raw_Text",
    ],
    sep="\n"
)


# ==================================================
# 3. JD 비교용 컬럼
# ==================================================
# required_skill 원문 + taxonomy/AI tag 결과를 같이 넣습니다.
# 단, 컬럼이 없으면 빈 문자열로 처리합니다.

jd_df["jd_skill_text"] = (
    safe_col(jd_df, "required_skill") + "\n" +
    safe_col(jd_df, "jd_skill_standard") + "\n" +
    safe_col(jd_df, "jd_ai_skill_tags")
).str.strip()

jd_df["jd_task_text"] = (
    safe_col(jd_df, "main_task") + "\n" +
    safe_col(jd_df, "jd_task_standard")
).str.strip()

jd_df["jd_domain_text"] = (
    safe_col(jd_df, "domain") + "\n" +
    safe_col(jd_df, "jd_domain_standard")
).str.strip()

jd_df["jd_role_text"] = (
    safe_col(jd_df, "role_signal") + "\n" +
    safe_col(jd_df, "job_title")
).str.strip()


# ==================================================
# 4. CV 비교용 컬럼
# ==================================================
# CV는 실제 컬럼명 Raw_Text를 사용합니다.
# Skill 원문 + taxonomy/AI tag 결과를 같이 넣습니다.

cv_df["cv_skill_text"] = (
    safe_col(cv_df, "Skill") + "\n" +
    safe_col(cv_df, "cv_skill_standard") + "\n" +
    safe_col(cv_df, "cv_ai_skill_tags")
).str.strip()

cv_df["cv_task_text"] = (
    safe_col(cv_df, "Career") + "\n" +
    safe_col(cv_df, "Career_Description") + "\n" +
    safe_col(cv_df, "cv_task_standard")
).str.strip()

cv_df["cv_domain_text"] = (
    safe_col(cv_df, "job_role") + "\n" +
    safe_col(cv_df, "Position") + "\n" +
    safe_col(cv_df, "cv_domain_standard")
).str.strip()

cv_df["cv_role_text"] = (
    safe_col(cv_df, "job_role") + "\n" +
    safe_col(cv_df, "Position")
).str.strip()


# ==================================================
# 5. 점검
# ==================================================

print("JD 비교용 컬럼 생성 완료")
print("CV 비교용 컬럼 생성 완료")

print("\nJD columns check:")
for c in [
    "jd_full_text",
    "jd_skill_text",
    "jd_task_text",
    "jd_domain_text",
    "jd_role_text",
]:
    print(c, ":", c in jd_df.columns)

print("\nCV columns check:")
for c in [
    "cv_full_text",
    "cv_skill_text",
    "cv_task_text",
    "cv_domain_text",
    "cv_role_text",
]:
    print(c, ":", c in cv_df.columns)

display(
    jd_df[
        [
            "jd_id",
            "job_title",
            "required_skill",
            "main_task",
            "domain",
            "role_signal",
            "jd_skill_text",
            "jd_task_text",
            "jd_domain_text",
            "jd_role_text",
        ]
    ].head(3)
)

display(
    cv_df[
        [
            "candidate_id",
            "job_role",
            "Skill",
            "Career",
            "Career_Description",
            "Position",
            "Raw_Text",
            "cv_skill_text",
            "cv_task_text",
            "cv_domain_text",
            "cv_role_text",
        ]
    ].head(3)
)


# In[17]:


# Cell 16 | 정밀 점수 계산 - CrossEncoder top100 4영역 + task/domain 길이 제한
# ------------------------------------------------------------
# 목적:
# - 전체 후보 865명은 pair_df에 유지합니다.
# - 최종 전체 랭킹 파일 855+10 / 855only 생성을 위해 전체 후보 점수는 모두 필요합니다.
# - 단, CPU 환경에서 865명 전체에 CrossEncoder를 적용하면 시간이 과도하게 걸리므로
#   CrossEncoder는 JD별 상위 TOP_N_CROSS_ENCODER명에게만 적용합니다.
#
# 적용 방식:
# - 전체 8,650 rows: TF-IDF fallback score 먼저 계산
# - JD별 상위 100명, 총 1,000 rows: CrossEncoder로 4영역 재점수
#   1) skill
#   2) task
#   3) domain
#   4) role
# - 나머지 7,650 rows: fallback score 유지
#
# 속도 안정화:
# - task 텍스트는 길어서 CrossEncoder 속도가 급격히 느려지므로 앞 800자만 사용
# - domain 텍스트도 안전하게 앞 600자만 사용
# - skill/role은 보통 짧으므로 제한 없이 사용
# ------------------------------------------------------------

import gc
import traceback
import numpy as np
import pandas as pd

cross_encoder_requested = bool(
    globals().get("USE_CROSS_ENCODER_REQUESTED", globals().get("USE_CROSS_ENCODER", False))
)

TOP_N_CROSS_ENCODER = int(globals().get("TOP_N_CROSS_ENCODER", 100))
ce_batch_size = int(globals().get("CROSS_ENCODER_BATCH_SIZE", 4))

MAX_TASK_CHARS = int(globals().get("MAX_TASK_CHARS", 800))
MAX_DOMAIN_CHARS = int(globals().get("MAX_DOMAIN_CHARS", 600))

print("=" * 80)
print("Cross Encoder requested:", cross_encoder_requested)
print("Model:", CROSS_ENCODER_MODEL_NAME)
print("pair_df shape:", pair_df.shape)
print("TOP_N_CROSS_ENCODER:", TOP_N_CROSS_ENCODER)
print("CROSS_ENCODER_BATCH_SIZE:", ce_batch_size)
print("MAX_TASK_CHARS:", MAX_TASK_CHARS)
print("MAX_DOMAIN_CHARS:", MAX_DOMAIN_CHARS)
print("=" * 80)


# ------------------------------------------------------------
# 1. 필수 비교 컬럼 검증
# ------------------------------------------------------------

required_text_cols = [
    "jd_skill_text", "cv_skill_text",
    "jd_task_text", "cv_task_text",
    "jd_domain_text", "cv_domain_text",
    "jd_role_text", "cv_role_text",
]

missing_text_cols = [col for col in required_text_cols if col not in pair_df.columns]

if missing_text_cols:
    raise KeyError(f"Cell 16 실행에 필요한 비교 텍스트 컬럼이 없습니다: {missing_text_cols}")


# ------------------------------------------------------------
# 2. 전체 후보 TF-IDF fallback score 계산
# ------------------------------------------------------------

print("전체 후보 TF-IDF fallback scoring 시작")

pair_df["skill_score_raw"] = pair_df.apply(
    lambda r: tfidf_pair_score(r["jd_skill_text"], r["cv_skill_text"]),
    axis=1
)

pair_df["task_score_raw"] = pair_df.apply(
    lambda r: tfidf_pair_score(r["jd_task_text"], r["cv_task_text"]),
    axis=1
)

pair_df["domain_score_raw"] = pair_df.apply(
    lambda r: tfidf_pair_score(r["jd_domain_text"], r["cv_domain_text"]),
    axis=1
)

pair_df["role_score_raw"] = pair_df.apply(
    lambda r: tfidf_pair_score(r["jd_role_text"], r["cv_role_text"]),
    axis=1
)

pair_df["scoring_method"] = "tfidf_fallback_all_candidates"

print("전체 후보 TF-IDF fallback scoring 완료")


# ------------------------------------------------------------
# 3. CrossEncoder 대상 후보 선택
# ------------------------------------------------------------

if "rank" in pair_df.columns:
    pair_df["_ce_rank"] = pd.to_numeric(pair_df["rank"], errors="coerce")
else:
    sort_score_col = "retrieval_score"

    if sort_score_col not in pair_df.columns:
        sort_score_col = "keyword_baseline_score"

    pair_df = pair_df.sort_values(
        ["jd_id", sort_score_col],
        ascending=[True, False]
    ).copy()

    pair_df["_ce_rank"] = (
        pair_df
        .groupby("jd_id")
        .cumcount()
        + 1
    )

ce_mask = pair_df["_ce_rank"] <= TOP_N_CROSS_ENCODER
ce_target_idx = pair_df.index[ce_mask]
ce_target_df = pair_df.loc[ce_target_idx].copy()

print(f"CrossEncoder 대상 pair 수: {len(ce_target_df):,} / 전체 pair 수: {len(pair_df):,}")

display(
    ce_target_df
    .groupby("jd_id")["candidate_id"]
    .nunique()
    .rename("ce_candidate_count")
    .reset_index()
)


# ------------------------------------------------------------
# 4. CrossEncoder 유틸 함수
# ------------------------------------------------------------

def _safe_text_series(df, col, max_chars=None):
    if col not in df.columns:
        raise KeyError(f"필수 비교 컬럼이 없습니다: {col}")

    s = df[col].fillna("").astype(str)

    if max_chars is not None:
        s = s.str.slice(0, max_chars)

    return s


def _build_pairs(df, jd_col, cv_col, max_chars=None):
    jd_texts = _safe_text_series(df, jd_col, max_chars=max_chars).tolist()
    cv_texts = _safe_text_series(df, cv_col, max_chars=max_chars).tolist()
    return list(zip(jd_texts, cv_texts))


def _predict_cross_encoder(
    cross_encoder,
    target_df,
    jd_col,
    cv_col,
    output_col,
    label,
    batch_size,
    max_chars=None,
):
    pairs = _build_pairs(
        target_df,
        jd_col,
        cv_col,
        max_chars=max_chars
    )

    print(
        f"Cross Encoder {label} scoring 시작: {len(pairs)} pairs "
        f"| batch_size={batch_size}"
        + (f" | max_chars={max_chars}" if max_chars is not None else "")
    )

    scores = cross_encoder.predict(
        pairs,
        batch_size=batch_size,
        show_progress_bar=True
    )

    result = pd.Series(
        np.asarray(scores, dtype="float32"),
        index=target_df.index,
        name=output_col
    )

    del pairs, scores
    gc.collect()

    print(f"Cross Encoder {label} scoring 완료")

    return result


# ------------------------------------------------------------
# 5. CrossEncoder 실행
# ------------------------------------------------------------

cross_encoder_success = False

if cross_encoder_requested and len(ce_target_df) > 0:
    try:
        from sentence_transformers import CrossEncoder

        cross_encoder = CrossEncoder(CROSS_ENCODER_MODEL_NAME)

        ce_results = {}

        ce_results["skill_score_raw"] = _predict_cross_encoder(
            cross_encoder=cross_encoder,
            target_df=ce_target_df,
            jd_col="jd_skill_text",
            cv_col="cv_skill_text",
            output_col="skill_score_raw",
            label="skill",
            batch_size=ce_batch_size,
            max_chars=None,
        )

        ce_results["task_score_raw"] = _predict_cross_encoder(
            cross_encoder=cross_encoder,
            target_df=ce_target_df,
            jd_col="jd_task_text",
            cv_col="cv_task_text",
            output_col="task_score_raw",
            label="task",
            batch_size=ce_batch_size,
            max_chars=MAX_TASK_CHARS,
        )

        ce_results["domain_score_raw"] = _predict_cross_encoder(
            cross_encoder=cross_encoder,
            target_df=ce_target_df,
            jd_col="jd_domain_text",
            cv_col="cv_domain_text",
            output_col="domain_score_raw",
            label="domain",
            batch_size=ce_batch_size,
            max_chars=MAX_DOMAIN_CHARS,
        )

        ce_results["role_score_raw"] = _predict_cross_encoder(
            cross_encoder=cross_encoder,
            target_df=ce_target_df,
            jd_col="jd_role_text",
            cv_col="cv_role_text",
            output_col="role_score_raw",
            label="role",
            batch_size=ce_batch_size,
            max_chars=None,
        )

        for col, series in ce_results.items():
            pair_df.loc[series.index, col] = series

        pair_df.loc[ce_target_idx, "scoring_method"] = "cross_encoder_top100_4areas_limited_text"
        cross_encoder_success = True

        del ce_results, cross_encoder
        gc.collect()

        print("Cross Encoder top100 4영역 scoring 완료")

    except Exception as e:
        print("=" * 80)
        print("Cross Encoder 사용 실패. 전체 후보는 TF-IDF fallback score를 유지합니다.")
        print("에러 타입:", type(e).__name__)
        print("에러 메시지:", e)
        print("상세 traceback:")
        traceback.print_exc()
        print("=" * 80)

        cross_encoder_success = False
else:
    print("CrossEncoder requested=False 또는 대상 후보 없음. 전체 후보는 TF-IDF fallback score를 사용합니다.")


# ------------------------------------------------------------
# 6. 실행 상태 기록 및 검증
# ------------------------------------------------------------

USE_CROSS_ENCODER = cross_encoder_success

score_cols_raw = [
    "skill_score_raw",
    "task_score_raw",
    "domain_score_raw",
    "role_score_raw",
]

print("scoring_method counts:")
print(pair_df["scoring_method"].value_counts(dropna=False))

print("raw score 결측 확인:")
print(pair_df[score_cols_raw].isna().sum())

print("CrossEncoder 적용 후보 수:")
print((pair_df["scoring_method"] == "cross_encoder_top100_4areas_limited_text").sum())

score_display_cols = [
    "jd_id",
    "candidate_id",
    "_ce_rank",
    "retrieval_score",
    "embedding_retrieval_score",
    "keyword_baseline_score",
    "scoring_method",
] + score_cols_raw

score_display_cols = [c for c in score_display_cols if c in pair_df.columns]

display(pair_df[score_display_cols].head(20))


# In[18]:


# Cell 17 | AI 태그 overlap score 함수
# ------------------------------------------------------------
# "LLM, RAG, Cloud" 같은 태그 문자열끼리 얼마나 겹치는지 계산합니다.
# JD 태그 중 CV 태그가 몇 개 포함되는지를 비율로 계산합니다.
# 예: JD = LLM, RAG, Cloud / CV = LLM, RAG → 2/3 = 0.67
# ------------------------------------------------------------

def overlap_score_from_tag_text(jd_tags, cv_tags):
    jd_set = set([
        x.strip()
        for x in str(jd_tags).split(",")
        if x.strip()
    ])
    cv_set = set([
        x.strip()
        for x in str(cv_tags).split(",")
        if x.strip()
    ])

    if not jd_set:
        return 0.0

    return len(jd_set & cv_set) / len(jd_set)


# In[19]:


# Cell 18 | Engineer JD에서 Researcher/Scientist 역할 패널티- 추가
def role_mismatch_penalty(row):
    jd_role_text = (
        str(row.get("jd_role_text", "")) + " " +
        str(row.get("jd_title", "")) + " " +
        str(row.get("jd_role_signal", ""))
    ).lower()

    cv_role_text = (
        str(row.get("cv_role_text", "")) + " " +
        str(row.get("job_role", "")) + " " +
        str(row.get("cv_job_role", ""))
    ).lower()

    is_engineer_jd = any(x in jd_role_text for x in [
        "engineer", "엔지니어", "developer", "개발자"
    ])

    is_researcher_cv = any(x in cv_role_text for x in [
        "researcher", "scientist", "research", "리서처", "연구원", "연구자"
    ])

    is_engineer_cv = any(x in cv_role_text for x in [
        "engineer", "developer", "개발자", "엔지니어", "backend", "frontend", "fullstack"
    ])

    if is_engineer_jd and is_researcher_cv and not is_engineer_cv:
        return 0.75   # 25% 감점

    return 1.0


pair_df["role_mismatch_penalty"] = pair_df.apply(role_mismatch_penalty, axis=1)


# In[20]:


# Cell 19 | 점수 정규화 및 final_score 계산
# ------------------------------------------------------------
# raw score는 모델/방식마다 스케일이 다르므로 JD별 min-max normalization을 적용합니다.
# 같은 JD 안에서 후보 간 상대 비교가 가능하도록 만듭니다.
#
# 수정 사항:
# - retrieval_score도 retrieval_score_norm으로 정규화합니다.
# - v22: retrieval_bm25_score / retrieval_tfidf_score / keyword_baseline_score도 JD별 정규화합니다.
# - skill/task/domain/role + ai_skill_overlap_score를 column_score로 합산합니다.
# - AI/LLM/RAG 계열 JD에서 skill과 ai tag가 모두 낮은 후보는 패널티를 적용합니다.
# - Engineer JD에서 Researcher/Scientist 후보가 1등으로 올라오는 문제를 강하게 보정합니다.
# - rank는 모든 패널티 적용 이후 JD별 내림차순 정렬 후 마지막에 다시 생성합니다.
# ------------------------------------------------------------

def minmax_by_group(df, group_col, score_col, output_col):
    df[output_col] = 0.0

    for key, idx in df.groupby(group_col).groups.items():
        s = pd.to_numeric(df.loc[idx, score_col], errors="coerce").fillna(0.0).astype(float)

        if s.max() == s.min():
            # 모두 같은 값이면 순위에 영향을 주지 않도록 0.0으로 둡니다.
            df.loc[idx, output_col] = 0.0
        else:
            df.loc[idx, output_col] = (s - s.min()) / (s.max() - s.min())

    return df


def safe_text(row, cols):
    texts = []
    for col in cols:
        if col in row.index:
            val = row.get(col, "")
            if pd.notna(val):
                texts.append(str(val))
    return " ".join(texts).lower()


def get_role_mismatch_penalty(row):
    """
    역할 불일치 보정 multiplier입니다.

    주의:
    - 1.0 = 감점 없음
    - 0.80~0.95 = 약한 감점
    - JD191 정답지에는 ai_scientist 후보도 포함되므로,
      researcher/scientist를 일괄 강감점하지 않습니다.
    - 최종 순위는 skill/task/domain/role/experience/soft skill 및
      Cross Encoder 점수와 함께 평가합니다.
    """

    jd_text = safe_text(row, [
        "jd_title",
        "job_title",
        "jd_role_signal",
        "role_signal",
        "jd_role_text",
        "jd_task_text",
        "main_task",
    ])

    cv_text = safe_text(row, [
        "candidate_id",
        "job_role",
        "cv_job_role",
        "Position",
        "cv_role_text",
        "cv_task_text",
        "Career",
        "Career_Description",
    ])

    is_engineer_jd = any(k in jd_text for k in [
        "product engineer",
        "engineer",
        "developer",
        "fullstack",
        "backend",
        "frontend",
        "엔지니어",
        "개발자",
        "풀스택",
        "백엔드",
        "프론트엔드",
    ])

    role_source_text = safe_text(row, [
        "candidate_id",
        "job_role",
        "cv_job_role",
    ])

    is_research_source_cv = any(k in role_source_text for k in [
        "ai_researcher",
        "ai_scientist",
        "researcher",
        "scientist",
        "리서처",
    ])

    is_research_cv = is_research_source_cv or any(k in cv_text for k in [
        "researcher",
        "scientist",
        "research",
        "ai_researcher",
        "ai_scientist",
        "리서처",
        "연구원",
        "연구자",
        "연구",
    ])

    is_engineer_cv = any(k in cv_text for k in [
        "engineer",
        "developer",
        "fullstack",
        "backend",
        "frontend",
        "software engineer",
        "product engineer",
        "ai engineer",
        "엔지니어",
        "개발자",
        "풀스택",
        "백엔드",
        "프론트엔드",
    ])

    candidate_id_text = safe_text(row, ["candidate_id"])
    is_researcher_id = "ai_researcher" in candidate_id_text or "researcher" in candidate_id_text
    is_scientist_id = "ai_scientist" in candidate_id_text or "scientist" in candidate_id_text

    # Engineer/Product Engineer JD에서 순수 researcher가 1위로 올라오는 것은 명시적으로 막습니다.
    # Scientist는 정답셋에 포함될 수 있어 researcher보다 약하게 보정하되,
    # engineer evidence가 없으면 상위 독주를 막습니다.
    if is_engineer_jd and is_researcher_id:
        return 0.55

    if is_engineer_jd and is_scientist_id and not is_engineer_cv:
        return 0.75

    if is_engineer_jd and is_scientist_id and is_engineer_cv:
        return 0.92

    if is_engineer_jd and is_research_source_cv and not is_engineer_cv:
        return 0.75

    if is_engineer_jd and is_research_cv and not is_engineer_cv:
        return 0.70

    if is_engineer_jd and is_research_cv and is_engineer_cv:
        return 0.92

    return 1.0


# ============================================================
# 1) Retrieval score 정규화
# ============================================================

pair_df = minmax_by_group(
    pair_df,
    "jd_id",
    "retrieval_score",
    "retrieval_score_norm"
)

# v22: keyword baseline component 정규화
for _raw_col, _norm_col in [
    ("retrieval_bm25_score", "retrieval_bm25_score_norm"),
    ("retrieval_tfidf_score", "retrieval_tfidf_score_norm"),
    ("keyword_baseline_score", "keyword_baseline_score_norm"),
    ("embedding_retrieval_score", "embedding_retrieval_score_norm"),
]:
    if _raw_col not in pair_df.columns:
        pair_df[_raw_col] = 0.0
    pair_df = minmax_by_group(pair_df, "jd_id", _raw_col, _norm_col)


# ============================================================
# 2) 컬럼별 raw score 정규화
# ============================================================

for raw_col, out_col in [
    ("skill_score_raw", "skill_score"),
    ("task_score_raw", "task_score"),
    ("domain_score_raw", "domain_score"),
    ("role_score_raw", "role_score"),
]:
    if raw_col not in pair_df.columns:
        pair_df[raw_col] = 0.0

    pair_df = minmax_by_group(
        pair_df,
        "jd_id",
        raw_col,
        out_col
    )


# ============================================================
# 3) AI/LLM/RAG 보정 태그 overlap 점수
# ============================================================

if "jd_ai_skill_tags" not in pair_df.columns:
    pair_df["jd_ai_skill_tags"] = ""

if "cv_ai_skill_tags" not in pair_df.columns:
    pair_df["cv_ai_skill_tags"] = ""

pair_df["ai_skill_overlap_score"] = pair_df.apply(
    lambda r: overlap_score_from_tag_text(
        r.get("jd_ai_skill_tags", ""),
        r.get("cv_ai_skill_tags", "")
    ),
    axis=1
)


# ============================================================
# 4) 컬럼 점수 합산
# ============================================================

pair_df["column_score"] = (
    WEIGHTS.get("skill_score", 0) * pair_df["skill_score"] +
    WEIGHTS.get("task_score", 0) * pair_df["task_score"] +
    WEIGHTS.get("domain_score", 0) * pair_df["domain_score"] +
    WEIGHTS.get("role_score", 0) * pair_df["role_score"] +
    WEIGHTS.get("ai_skill_overlap_score", 0) * pair_df["ai_skill_overlap_score"]
)


# ============================================================
# 5) Skill Risk 패널티
# ============================================================
# JD에 AI 핵심 태그가 있는데,
# 후보의 skill_score와 ai_skill_overlap_score가 둘 다 낮으면 감점합니다.
# ------------------------------------------------------------

pair_df["skill_risk_flag"] = pair_df.apply(
    lambda r:
        bool(str(r.get("jd_ai_skill_tags", "")).strip())
        and float(r.get("skill_score", 0)) < 0.10
        and float(r.get("ai_skill_overlap_score", 0)) < 0.20,
    axis=1
)

pair_df["skill_risk_penalty"] = 1.0
pair_df.loc[pair_df["skill_risk_flag"], "skill_risk_penalty"] = 0.65

pair_df["column_score_before_skill_penalty"] = pair_df["column_score"]
pair_df["column_score"] = pair_df["column_score"] * pair_df["skill_risk_penalty"]


# ============================================================
# 6) Role mismatch 패널티
# ============================================================
# AI Product Engineer JD에서 researcher/scientist가 과도하게 올라오는 문제를 방지합니다.
# ------------------------------------------------------------

pair_df["role_mismatch_penalty"] = pair_df.apply(
    get_role_mismatch_penalty,
    axis=1
)


# ============================================================
# 7) 최종 점수 계산
# ============================================================

# v22: BM25+TF-IDF baseline을 final score에 별도 반영합니다.
_retrieval_w = float(globals().get("RETRIEVAL_SCORE_WEIGHT", 0.25))
_keyword_w = float(globals().get("KEYWORD_BASELINE_SCORE_WEIGHT", 0.15))
_column_w = float(globals().get("COLUMN_SCORE_WEIGHT", 0.75))
_weight_denom = _retrieval_w + _keyword_w + _column_w
if _weight_denom <= 0:
    _retrieval_w, _keyword_w, _column_w = 0.25, 0.15, 0.75
    _weight_denom = _retrieval_w + _keyword_w + _column_w

pair_df["final_score_before_role_penalty"] = (
    _retrieval_w * pair_df["retrieval_score_norm"] +
    _keyword_w * pair_df["keyword_baseline_score_norm"] +
    _column_w * pair_df["column_score"]
) / _weight_denom

pair_df["final_score"] = (
    pair_df["final_score_before_role_penalty"] *
    pair_df["role_mismatch_penalty"]
)


# ============================================================
# 8) 랭크 재생성
# ============================================================

pair_df = pair_df.sort_values(
    ["jd_id", "final_score"],
    ascending=[True, False]
).copy()

pair_df["rank"] = pair_df.groupby("jd_id").cumcount() + 1

top_match_df = pair_df[pair_df["rank"] <= TOP_N_FINAL].copy()


# ============================================================
# 9) 점검 출력
# ============================================================

print("top_match_df:", top_match_df.shape)

debug_cols = [
    "jd_id",
    "candidate_id",
    "job_role",
    "cv_job_role",
    "rank",
    "retrieval_score",
    "retrieval_score_norm",
    "embedding_retrieval_score_norm",
    "retrieval_bm25_score",
    "retrieval_bm25_score_norm",
    "retrieval_tfidf_score",
    "retrieval_tfidf_score_norm",
    "keyword_baseline_score",
    "keyword_baseline_score_norm",
    "skill_score",
    "task_score",
    "domain_score",
    "role_score",
    "ai_skill_overlap_score",
    "column_score_before_skill_penalty",
    "skill_risk_penalty",
    "column_score",
    "final_score_before_role_penalty",
    "role_mismatch_penalty",
    "final_score",
    "jd_ai_skill_tags",
    "cv_ai_skill_tags",
    "skill_risk_flag",
]

debug_cols = [c for c in debug_cols if c in top_match_df.columns]

display(
    top_match_df[debug_cols].head(20)
)

# researcher/scientist 후보가 실제로 감점됐는지 별도 확인
role_check_cols = [
    "candidate_id",
    "job_role",
    "cv_job_role",
    "rank",
    "final_score_before_role_penalty",
    "role_mismatch_penalty",
    "final_score",
]

role_check_cols = [c for c in role_check_cols if c in pair_df.columns]

display(
    pair_df[
        pair_df["candidate_id"].astype(str).str.contains(
            "researcher|scientist",
            case=False,
            na=False
        )
    ][role_check_cols].head(20)
)


# In[21]:


# Cell 20 | 대시보드 육각형 지표 생성 - 공통 6축
# ------------------------------------------------------------
# 육각형 차트는 JD191 전용 신호가 아니라 모든 JD/CV에 공통 적용 가능한 축으로 구성합니다.
#
# Radar 6축:
# 1) Skill       : 기술 스택 적합도
# 2) Task        : 실제 수행 업무 적합도
# 3) Domain      : 산업/도메인 적합도
# 4) Role        : 직무 역할 적합도
# 5) Experience  : 경력 수준/요구경력 충족도
# 6) Soft Skill  : 자기주도, 협업, 리더십, 커뮤니케이션, 임팩트 신호
#
# 주의:
# - radar_* 컬럼은 시각화/해석용입니다.
# - final_score 랭킹 계산과는 분리합니다.
# ------------------------------------------------------------

import re
import numpy as np
import pandas as pd


def _get_text_from_row(row, cols):
    values = []
    for col in cols:
        if col in row.index:
            val = row.get(col, "")
            if pd.notna(val) and str(val).strip():
                values.append(str(val))
    return " ".join(values)


def extract_years_from_text(text):
    """'총 3년 1개월', '2년 이상', '18년 4개월' 등에서 년/개월을 float year로 변환."""
    text = "" if pd.isna(text) else str(text)

    years = 0.0
    months = 0.0

    y = re.search(r"(\d+(?:\.\d+)?)\s*년", text)
    m = re.search(r"(\d+(?:\.\d+)?)\s*개월", text)

    if y:
        years = float(y.group(1))
    if m:
        months = float(m.group(1))

    if not y:
        y2 = re.search(r"(\d+(?:\.\d+)?)\s*(?:yrs|years|year)", text, flags=re.I)
        if y2:
            years = float(y2.group(1))

    return years + months / 12.0


def experience_score_from_row(row):
    cv_years = extract_years_from_text(_get_text_from_row(row, [
        "Total_Career_Years",
        "cv_total_career_years",
        "Career",
    ]))

    req_years = extract_years_from_text(_get_text_from_row(row, [
        "career_requirement",
        "jd_career_requirement",
    ]))

    position_text = _get_text_from_row(row, [
        "Position",
        "cv_position",
        "Career",
    ]).lower()

    if req_years <= 0:
        base = min(cv_years / 5.0, 1.0) if cv_years > 0 else 0.5
    else:
        base = min(cv_years / req_years, 1.0) if cv_years > 0 else 0.0

    seniority_bonus = 0.0
    if any(k in position_text for k in ["대리", "과장", "차장", "부장", "책임", "수석", "팀장", "lead", "principal", "senior"]):
        seniority_bonus = 0.08

    return float(np.clip(base + seniority_bonus, 0.0, 1.0))


SOFT_SKILL_PATTERNS = {
    "ownership": [
        "주도", "리드", "lead", "owner", "ownership", "책임", "직접", "설계", "기획",
        "문제 정의", "개선", "고도화"
    ],
    "collaboration": [
        "협업", "커뮤니케이션", "조율", "po", "pm", "디자이너", "기획자", "고객사",
        "stakeholder", "cross-functional"
    ],
    "leadership": [
        "팀장", "리더", "lead", "멘토", "관리", "총괄", "architect", "pm", "프로젝트 관리"
    ],
    "impact": [
        "성과", "개선", "향상", "감소", "단축", "절감", "증가", "%", "배", "달성",
        "운영", "배포", "production", "프로덕션", "고객", "납품"
    ],
    "communication": [
        "문서화", "발표", "공유", "교육", "보고", "가이드", "세미나", "블로그",
        "리포트", "대시보드"
    ],
}


def soft_skill_score_from_row(row):
    text = _get_text_from_row(row, [
        "Career_Description",
        "cv_career_description",
        "Career",
        "cv_career",
        "Self_Introduction",
        "cv_self_introduction",
        "Portfolio",
        "cv_portfolio",
    ]).lower()

    if not text.strip():
        return 0.0

    matched_dimensions = 0
    for _, keywords in SOFT_SKILL_PATTERNS.items():
        if any(k.lower() in text for k in keywords):
            matched_dimensions += 1

    return matched_dimensions / len(SOFT_SKILL_PATTERNS)


top_match_df["radar_skill_score"] = pd.to_numeric(top_match_df.get("skill_score", 0), errors="coerce").fillna(0.0)
top_match_df["radar_task_score"] = pd.to_numeric(top_match_df.get("task_score", 0), errors="coerce").fillna(0.0)
top_match_df["radar_domain_score"] = pd.to_numeric(top_match_df.get("domain_score", 0), errors="coerce").fillna(0.0)
top_match_df["radar_role_score"] = pd.to_numeric(top_match_df.get("role_score", 0), errors="coerce").fillna(0.0)

top_match_df["experience_score"] = top_match_df.apply(experience_score_from_row, axis=1)
top_match_df["soft_skill_score"] = top_match_df.apply(soft_skill_score_from_row, axis=1)

top_match_df["radar_experience_score"] = top_match_df["experience_score"]
top_match_df["radar_soft_skill_score"] = top_match_df["soft_skill_score"]

RADAR_AXIS_COLUMNS = [
    "radar_skill_score",
    "radar_task_score",
    "radar_domain_score",
    "radar_role_score",
    "radar_experience_score",
    "radar_soft_skill_score",
]

for col in RADAR_AXIS_COLUMNS + ["experience_score", "soft_skill_score"]:
    top_match_df[col + "_100"] = (pd.to_numeric(top_match_df[col], errors="coerce").fillna(0.0) * 100).round(1)

print("Radar metric columns added:")
display(top_match_df[[
    "jd_id", "candidate_id", "rank",
    "radar_skill_score_100", "radar_task_score_100", "radar_domain_score_100",
    "radar_role_score_100", "radar_experience_score_100", "radar_soft_skill_score_100"
]].head(20))


# In[22]:


# Cell 21 | Ground Truth 평가셋 로드 + Ranking Metrics
# ------------------------------------------------------------
# 목적:
# - data/10_evaluation/ground_truth_matches.xlsx를 유일한 정답 기준으로 사용합니다.
# - 정답셋의 jd_id/candidate_id/relevance_grade/gold_rank 기준으로 ranking metric을 계산합니다.
# - 파일이 없으면 샘플 정답을 만들지 않고 템플릿만 생성한 뒤 중단합니다.
#
# 평가 기준:
# - relevance_grade >= 2: positive candidate
# - NDCG@K: relevance_grade의 graded relevance를 그대로 사용
# - Precision/Recall/MRR: positive candidate 기준
# ------------------------------------------------------------

import math
import json
from pathlib import Path
import numpy as np
import pandas as pd

# Cell 01에서 만든 경로를 우선 사용하고, 단독 실행 시 fallback합니다.
EVALUATION_DIR = globals().get("EVALUATION_DIR", PROJECT_ROOT / "data" / "10_evaluation" if "PROJECT_ROOT" in globals() else Path("data/10_evaluation"))
EVALUATION_DIR.mkdir(parents=True, exist_ok=True)

GROUND_TRUTH_PATH = globals().get("GROUND_TRUTH_PATH", EVALUATION_DIR / "ground_truth_matches.xlsx")
EVALUATION_REPORT_PATH = globals().get("EVALUATION_REPORT_PATH", EVALUATION_DIR / "evaluation_report_latest.xlsx")
WEIGHT_VALIDATION_PATH = globals().get("WEIGHT_VALIDATION_PATH", EVALUATION_DIR / "weight_config_validation_latest.xlsx")

GROUND_TRUTH_REQUIRED_COLUMNS = ["jd_id", "candidate_id", "relevance_grade"]
GROUND_TRUTH_OPTIONAL_COLUMNS = [
    "jd_title", "company", "gold_rank", "final_score_text", "final_score_value",
    "final_score_max", "career_years_text", "reason", "soft_skill",
    "tenure_pattern", "source_sheet", "evaluator", "created_at",
]


def create_ground_truth_template(path):
    template_cols = GROUND_TRUTH_REQUIRED_COLUMNS + GROUND_TRUTH_OPTIONAL_COLUMNS
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame(columns=template_cols).to_excel(writer, sheet_name="ground_truth_matches", index=False)
        pd.DataFrame({
            "field": GROUND_TRUTH_REQUIRED_COLUMNS,
            "description": [
                "JD identifier. Must match result_df/pair_df jd_id.",
                "Candidate identifier. Must match candidate_id in CV master.",
                "0=negative, 1=reserve, 2=interview, 3=top candidate.",
            ],
        }).to_excel(writer, sheet_name="README", index=False)
    return path


def read_ground_truth(path):
    if not path.exists():
        create_ground_truth_template(path)
        raise FileNotFoundError(
            f"Ground Truth 파일이 없어 템플릿만 생성했습니다: {path}\n"
            "정답 후보를 입력한 뒤 이 셀부터 다시 실행하세요."
        )

    sheet_names = pd.ExcelFile(path).sheet_names
    sheet_name = "ground_truth_matches" if "ground_truth_matches" in sheet_names else sheet_names[0]
    gt = pd.read_excel(path, sheet_name=sheet_name)
    gt.columns = [str(c).strip() for c in gt.columns]

    missing = [c for c in GROUND_TRUTH_REQUIRED_COLUMNS if c not in gt.columns]
    if missing:
        raise ValueError(f"Ground Truth 필수 컬럼이 없습니다: {missing} | file={path} | sheet={sheet_name}")

    gt = gt.copy()
    gt["jd_id"] = gt["jd_id"].astype(str).str.replace(r"\.0$", "", regex=True).str.strip()
    gt["candidate_id"] = gt["candidate_id"].astype(str).str.strip()
    gt["relevance_grade"] = pd.to_numeric(gt["relevance_grade"], errors="coerce").fillna(0).astype(int)

    if "gold_rank" in gt.columns:
        gt["gold_rank"] = pd.to_numeric(gt["gold_rank"], errors="coerce")
    else:
        gt["gold_rank"] = np.nan

    # 완전히 빈 행 제거 + 중복 정답 제거. 중복은 더 높은 grade/더 낮은 gold_rank를 우선합니다.
    gt = gt[(gt["jd_id"] != "") & (gt["candidate_id"] != "")]
    gt = gt.sort_values(["jd_id", "candidate_id", "relevance_grade", "gold_rank"], ascending=[True, True, False, True])
    gt = gt.drop_duplicates(["jd_id", "candidate_id"], keep="first").reset_index(drop=True)

    if gt.empty:
        raise ValueError(f"Ground Truth가 비어 있습니다: {path}")

    return gt, sheet_name


GROUND_TRUTH_DF, GROUND_TRUTH_SHEET_NAME = read_ground_truth(GROUND_TRUTH_PATH)
GROUND_TRUTH_JD_IDS = sorted(GROUND_TRUTH_DF["jd_id"].astype(str).unique().tolist(), key=lambda x: int(x) if x.isdigit() else x)
PREDICTED_JD_IDS = sorted(pair_df["jd_id"].astype(str).unique().tolist(), key=lambda x: int(x) if x.isdigit() else x)
MISSING_PREDICTION_JD_IDS = [jd for jd in GROUND_TRUTH_JD_IDS if jd not in set(PREDICTED_JD_IDS)]

print(f"Ground Truth loaded: {GROUND_TRUTH_PATH}")
print(f"Ground Truth sheet: {GROUND_TRUTH_SHEET_NAME}")
print(f"Ground Truth rows: {len(GROUND_TRUTH_DF)} | JD count: {len(GROUND_TRUTH_JD_IDS)}")
print("Ground Truth JD IDs:", GROUND_TRUTH_JD_IDS)
print("Prediction JD IDs:", PREDICTED_JD_IDS)
if MISSING_PREDICTION_JD_IDS:
    print("주의: 이번 실행 결과에 없는 Ground Truth JD IDs:", MISSING_PREDICTION_JD_IDS)
    print("Cell 01의 USE_GROUND_TRUTH_JD_IDS=True 또는 TARGET_JD_IDS 설정을 확인하세요.")

display(GROUND_TRUTH_DF.head(30))

# ------------------------------------------------------------
# Ranking Metrics
# ------------------------------------------------------------
POSITIVE_RELEVANCE_THRESHOLD = 2
EVALUATION_K_VALUES = (3, 5, 10, 20, 50)


def _dcg(relevances):
    return sum((2 ** rel - 1) / math.log2(i + 2) for i, rel in enumerate(relevances))


def evaluate_one_jd(pred_df, gt_df, jd_id, score_col="final_score", k=10, positive_threshold=2):
    jd_id = str(jd_id)
    pred = pred_df[pred_df["jd_id"].astype(str) == jd_id].copy()
    gt = gt_df[gt_df["jd_id"].astype(str) == jd_id].copy()

    if pred.empty or gt.empty:
        return {
            "jd_id": jd_id,
            "k": k,
            "pred_count": len(pred),
            "gt_count": len(gt),
            "positive_gt_count": int((gt["relevance_grade"] >= positive_threshold).sum()) if not gt.empty else 0,
            "precision_at_k": np.nan,
            "recall_at_k": np.nan,
            "ndcg_at_k": np.nan,
            "mrr_at_k": np.nan,
            "hit_count_at_k": 0,
            "hit_ids_at_k": "",
            "missing_prediction": bool(pred.empty and not gt.empty),
        }

    if score_col in pred.columns:
        pred = pred.sort_values(score_col, ascending=False)
    elif "rank" in pred.columns:
        pred = pred.sort_values("rank", ascending=True)
    else:
        raise ValueError(f"{score_col} 또는 rank 컬럼이 필요합니다.")

    pred_ids = pred["candidate_id"].astype(str).head(k).tolist()
    rel_map = dict(zip(gt["candidate_id"].astype(str), gt["relevance_grade"].astype(int)))

    positive_ids = set(gt.loc[gt["relevance_grade"] >= positive_threshold, "candidate_id"].astype(str))
    hit_ids = [cid for cid in pred_ids if cid in positive_ids]

    precision = len(hit_ids) / min(k, len(pred_ids)) if pred_ids else np.nan
    recall = len(hit_ids) / len(positive_ids) if positive_ids else np.nan

    ranked_rels = [rel_map.get(cid, 0) for cid in pred_ids]
    actual_dcg = _dcg(ranked_rels)
    ideal_rels = sorted(gt["relevance_grade"].astype(int).tolist(), reverse=True)[:k]
    ideal_dcg = _dcg(ideal_rels)
    ndcg = actual_dcg / ideal_dcg if ideal_dcg > 0 else np.nan

    mrr = 0.0
    for i, cid in enumerate(pred_ids):
        if cid in positive_ids:
            mrr = 1 / (i + 1)
            break

    return {
        "jd_id": jd_id,
        "k": k,
        "pred_count": len(pred),
        "gt_count": len(gt),
        "positive_gt_count": len(positive_ids),
        "precision_at_k": round(precision, 4) if not pd.isna(precision) else np.nan,
        "recall_at_k": round(recall, 4) if not pd.isna(recall) else np.nan,
        "ndcg_at_k": round(ndcg, 4) if not pd.isna(ndcg) else np.nan,
        "mrr_at_k": round(mrr, 4),
        "hit_count_at_k": len(hit_ids),
        "hit_ids_at_k": ", ".join(hit_ids),
        "missing_prediction": False,
    }


def evaluate_ranking(pred_df, gt_df, score_col="final_score", k_values=EVALUATION_K_VALUES, positive_threshold=2):
    rows = []
    for jd_id in sorted(gt_df["jd_id"].astype(str).unique(), key=lambda x: int(x) if str(x).isdigit() else str(x)):
        for k in k_values:
            rows.append(evaluate_one_jd(pred_df, gt_df, jd_id, score_col=score_col, k=k, positive_threshold=positive_threshold))

    detail = pd.DataFrame(rows)
    summary = (
        detail.groupby("k", dropna=False)[["precision_at_k", "recall_at_k", "ndcg_at_k", "mrr_at_k"]]
        .mean(numeric_only=True)
        .reset_index()
        .rename(columns={
            "precision_at_k": "macro_precision_at_k",
            "recall_at_k": "macro_recall_at_k",
            "ndcg_at_k": "macro_ndcg_at_k",
            "mrr_at_k": "macro_mrr_at_k",
        })
    )
    coverage = pd.DataFrame({
        "ground_truth_jd_count": [gt_df["jd_id"].astype(str).nunique()],
        "predicted_jd_count": [pred_df["jd_id"].astype(str).nunique()],
        "missing_prediction_jd_count": [len(MISSING_PREDICTION_JD_IDS)],
        "missing_prediction_jd_ids": [", ".join(MISSING_PREDICTION_JD_IDS)],
        "positive_threshold": [positive_threshold],
        "k_values": [", ".join(map(str, k_values))],
    })
    return detail, summary, coverage


EVALUATION_DETAIL_DF, EVALUATION_SUMMARY_DF, EVALUATION_COVERAGE_DF = evaluate_ranking(
    pair_df,
    GROUND_TRUTH_DF,
    score_col="final_score",
    k_values=EVALUATION_K_VALUES,
    positive_threshold=POSITIVE_RELEVANCE_THRESHOLD,
)



# ------------------------------------------------------------
# v22 | 모델/점수별 비교 평가
# ------------------------------------------------------------
# Keyword analysis는 가능하면 KEYWORD_SCORE_DF 전체 후보군 기준으로 평가합니다.
# Rule-based / embedding / cross-encoder 결과는 pair_df 후보군 기준으로 평가합니다.
# ------------------------------------------------------------

def evaluate_multiple_model_specs(model_specs, gt_df, k_values=EVALUATION_K_VALUES, positive_threshold=2):
    summary_rows = []
    detail_frames = []

    for spec in model_specs:
        model_name = spec["model_name"]
        score_col = spec["score_col"]
        group_name = spec["group_name"]
        pred_df = spec["pred_df"]
        candidate_scope = spec.get("candidate_scope", "pair_df")

        if pred_df is None or len(pred_df) == 0:
            print(f"평가 제외: {model_name} | 예측 데이터프레임이 비어 있습니다.")
            continue
        if score_col not in pred_df.columns:
            print(f"평가 제외: {model_name} | 없는 컬럼: {score_col}")
            continue

        detail_df, summary_df, _ = evaluate_ranking(
            pred_df,
            gt_df,
            score_col=score_col,
            k_values=k_values,
            positive_threshold=positive_threshold,
        )
        detail_df.insert(0, "model_name", model_name)
        detail_df.insert(1, "score_col", score_col)
        detail_df.insert(2, "group_name", group_name)
        detail_df.insert(3, "candidate_scope", candidate_scope)
        detail_frames.append(detail_df)

        for _, row in summary_df.iterrows():
            summary_rows.append({
                "group_name": group_name,
                "model_name": model_name,
                "score_col": score_col,
                "candidate_scope": candidate_scope,
                "k": int(row["k"]),
                "macro_precision_at_k": row["macro_precision_at_k"],
                "macro_recall_at_k": row["macro_recall_at_k"],
                "macro_ndcg_at_k": row["macro_ndcg_at_k"],
                "macro_mrr_at_k": row["macro_mrr_at_k"],
            })

    summary = pd.DataFrame(summary_rows)
    if not summary.empty:
        summary = summary.sort_values(
            ["k", "macro_ndcg_at_k", "macro_mrr_at_k", "macro_recall_at_k"],
            ascending=[True, False, False, False]
        ).reset_index(drop=True)
    detail = pd.concat(detail_frames, ignore_index=True) if detail_frames else pd.DataFrame()
    return summary, detail


keyword_eval_df = globals().get("KEYWORD_SCORE_DF", pair_df).copy()
if "KEYWORD_SCORE_DF" in globals():
    print("Keyword analysis 평가는 KEYWORD_SCORE_DF 전체 후보군 기준으로 수행합니다:", keyword_eval_df.shape)
else:
    print("KEYWORD_SCORE_DF가 없어 pair_df 기준으로 keyword analysis 평가를 수행합니다.")

MODEL_SCORE_SPECS = [
    {
        "group_name": "Keyword analysis",
        "model_name": "Keyword BM25",
        "score_col": "retrieval_bm25_score",
        "pred_df": keyword_eval_df,
        "candidate_scope": "KEYWORD_SCORE_DF_all_candidates",
    },
    {
        "group_name": "Keyword analysis",
        "model_name": "Keyword TF-IDF",
        "score_col": "retrieval_tfidf_score",
        "pred_df": keyword_eval_df,
        "candidate_scope": "KEYWORD_SCORE_DF_all_candidates",
    },
    {
        "group_name": "Keyword analysis",
        "model_name": "Keyword BM25+TF-IDF baseline",
        "score_col": "keyword_baseline_score",
        "pred_df": keyword_eval_df,
        "candidate_scope": "KEYWORD_SCORE_DF_all_candidates",
    },
    {
        "group_name": "Rule-based / retrieval",
        "model_name": "Embedding retrieval",
        "score_col": "embedding_retrieval_score_norm",
        "pred_df": pair_df,
        "candidate_scope": "pair_df_top_candidates",
    },
    {
        "group_name": "Rule-based / retrieval",
        "model_name": "Retrieval active score",
        "score_col": "retrieval_score_norm",
        "pred_df": pair_df,
        "candidate_scope": "pair_df_top_candidates",
    },
    {
        "group_name": "Rule-based / CE or TF-IDF",
        "model_name": "Column rule-based score",
        "score_col": "column_score",
        "pred_df": pair_df,
        "candidate_scope": "pair_df_top_candidates",
    },
    {
        "group_name": "Rule-based / full pipeline",
        "model_name": "Final score",
        "score_col": "final_score",
        "pred_df": pair_df,
        "candidate_scope": "pair_df_top_candidates",
    },
]

MODEL_COMPARISON_SUMMARY_DF, MODEL_COMPARISON_DETAIL_DF = evaluate_multiple_model_specs(
    MODEL_SCORE_SPECS,
    GROUND_TRUTH_DF,
    k_values=EVALUATION_K_VALUES,
    positive_threshold=POSITIVE_RELEVANCE_THRESHOLD,
)

with pd.ExcelWriter(EVALUATION_REPORT_PATH, engine="openpyxl") as writer:
    EVALUATION_SUMMARY_DF.to_excel(writer, sheet_name="summary", index=False)
    EVALUATION_DETAIL_DF.to_excel(writer, sheet_name="detail", index=False)
    EVALUATION_COVERAGE_DF.to_excel(writer, sheet_name="coverage", index=False)
    MODEL_COMPARISON_SUMMARY_DF.to_excel(writer, sheet_name="model_comparison_summary", index=False)
    MODEL_COMPARISON_DETAIL_DF.to_excel(writer, sheet_name="model_comparison_detail", index=False)
    GROUND_TRUTH_DF.to_excel(writer, sheet_name="ground_truth", index=False)

print(f"Evaluation report saved: {EVALUATION_REPORT_PATH}")
display(EVALUATION_COVERAGE_DF)
print("Final score evaluation summary")
display(EVALUATION_SUMMARY_DF)
print("BM25 / TF-IDF / Rule-based 모델 비교 summary")
display(MODEL_COMPARISON_SUMMARY_DF)
display(EVALUATION_DETAIL_DF)


# In[23]:


# Cell 22 | 모델 가중치 검증 - Ground Truth 기반 설정 비교
# ------------------------------------------------------------
# 목적:
# - 가중치를 감으로 바꾸지 않고, Ground Truth 평가셋 기준으로 비교합니다.
# - JD가 여러 개 쌓이면 JD 단위 macro-average가 됩니다.
# - 현재 평가셋이 JD 1개뿐이면 엄밀한 k-fold CV가 아니라 "JD191 holdout validation"입니다.
#
# 평가 대상:
# - 기존 retrieval 후보군(pair_df) 안에서 final_score 구성 가중치만 바꿔 비교합니다.
# - Retrieval 자체의 후보군을 바꾸는 실험은 Cell 10/11부터 다시 실행해야 합니다.
# ------------------------------------------------------------

MODEL_WEIGHT_CONFIGS = {
    "current": {
        "retrieval_score_norm": RETRIEVAL_SCORE_WEIGHT,
        "keyword_baseline_score_norm": KEYWORD_BASELINE_SCORE_WEIGHT,
        "column_score": COLUMN_SCORE_WEIGHT,
        "skill_score": WEIGHTS.get("skill_score", 0.35),
        "task_score": WEIGHTS.get("task_score", 0.25),
        "domain_score": WEIGHTS.get("domain_score", 0.10),
        "role_score": WEIGHTS.get("role_score", 0.15),
        "ai_skill_overlap_score": WEIGHTS.get("ai_skill_overlap_score", 0.15),
    },
    "user_skill_task_55301005": {
        "retrieval_score_norm": RETRIEVAL_SCORE_WEIGHT,
        "column_score": COLUMN_SCORE_WEIGHT,
        "skill_score": 0.55,
        "task_score": 0.30,
        "role_score": 0.10,
        "domain_score": 0.05,
        "ai_skill_overlap_score": 0.00,
    },
    "skill_heavy": {
        "retrieval_score_norm": 0.25,
        "column_score": 0.75,
        "skill_score": 0.45,
        "task_score": 0.25,
        "domain_score": 0.10,
        "role_score": 0.10,
        "ai_skill_overlap_score": 0.10,
    },
    "skill_task_balanced": {
        "retrieval_score_norm": 0.25,
        "column_score": 0.75,
        "skill_score": 0.35,
        "task_score": 0.35,
        "domain_score": 0.10,
        "role_score": 0.10,
        "ai_skill_overlap_score": 0.10,
    },
    "role_stronger": {
        "retrieval_score_norm": 0.25,
        "column_score": 0.75,
        "skill_score": 0.35,
        "task_score": 0.25,
        "domain_score": 0.10,
        "role_score": 0.20,
        "ai_skill_overlap_score": 0.10,
    },
    "task_role_stronger": {
        "retrieval_score_norm": 0.25,
        "column_score": 0.75,
        "skill_score": 0.30,
        "task_score": 0.30,
        "domain_score": 0.10,
        "role_score": 0.20,
        "ai_skill_overlap_score": 0.10,
    },
    "column_heavy": {
        "retrieval_score_norm": 0.15,
        "column_score": 0.85,
        "skill_score": 0.35,
        "task_score": 0.25,
        "domain_score": 0.10,
        "role_score": 0.15,
        "ai_skill_overlap_score": 0.15,
    },
}


def normalize_component_weights(config):
    component_keys = ["skill_score", "task_score", "domain_score", "role_score", "ai_skill_overlap_score"]
    total = sum(float(config.get(k, 0.0)) for k in component_keys)
    if total <= 0:
        raise ValueError("component weights sum must be positive")
    return {k: float(config.get(k, 0.0)) / total for k in component_keys}


def score_with_config(df, config, output_col):
    out = df.copy()
    component_weights = normalize_component_weights(config)

    config_column_score = np.zeros(len(out), dtype=float)
    for col, w in component_weights.items():
        if col not in out.columns:
            out[col] = 0.0
        config_column_score += w * pd.to_numeric(out[col], errors="coerce").fillna(0.0).values

    # 기존과 동일하게 skill risk와 role mismatch multiplier를 적용합니다.
    skill_penalty = pd.to_numeric(out.get("skill_risk_penalty", 1.0), errors="coerce").fillna(1.0).values
    role_penalty = pd.to_numeric(out.get("role_mismatch_penalty", 1.0), errors="coerce").fillna(1.0).values

    config_column_score = config_column_score * skill_penalty

    retrieval_w = float(config.get("retrieval_score_norm", RETRIEVAL_SCORE_WEIGHT))
    keyword_w = float(config.get("keyword_baseline_score_norm", globals().get("KEYWORD_BASELINE_SCORE_WEIGHT", 0.15)))
    column_w = float(config.get("column_score", COLUMN_SCORE_WEIGHT))
    denom = retrieval_w + keyword_w + column_w
    if denom <= 0:
        retrieval_w, keyword_w, column_w = RETRIEVAL_SCORE_WEIGHT, globals().get("KEYWORD_BASELINE_SCORE_WEIGHT", 0.15), COLUMN_SCORE_WEIGHT
        denom = retrieval_w + keyword_w + column_w

    retrieval_w, keyword_w, column_w = retrieval_w / denom, keyword_w / denom, column_w / denom

    retrieval_component = pd.to_numeric(out.get("retrieval_score_norm", 0.0), errors="coerce").fillna(0.0).values
    keyword_component = pd.to_numeric(out.get("keyword_baseline_score_norm", 0.0), errors="coerce").fillna(0.0).values
    out[output_col] = (retrieval_w * retrieval_component + keyword_w * keyword_component + column_w * config_column_score) * role_penalty
    return out

validation_rows = []
validation_details = []

for config_name, config in MODEL_WEIGHT_CONFIGS.items():
    score_col = f"score__{config_name}"
    scored_df = score_with_config(pair_df, config, score_col)

    detail_df, summary_df, _coverage_df = evaluate_ranking(
        scored_df,
        GROUND_TRUTH_DF,
        score_col=score_col,
        k_values=globals().get("EVALUATION_K_VALUES", (3, 5, 10, 20, 50)),
        positive_threshold=POSITIVE_RELEVANCE_THRESHOLD,
    )
    detail_df.insert(0, "config_name", config_name)
    validation_details.append(detail_df)

    for _, row in summary_df.iterrows():
        validation_rows.append({
            "config_name": config_name,
            "k": int(row["k"]),
            "macro_precision_at_k": row["macro_precision_at_k"],
            "macro_recall_at_k": row["macro_recall_at_k"],
            "macro_ndcg_at_k": row["macro_ndcg_at_k"],
            "macro_mrr_at_k": row["macro_mrr_at_k"],
            "weights": json.dumps(config, ensure_ascii=False),
        })

WEIGHT_CONFIG_VALIDATION_DF = pd.DataFrame(validation_rows).sort_values(
    ["k", "macro_ndcg_at_k", "macro_recall_at_k", "macro_precision_at_k"],
    ascending=[True, False, False, False]
)
WEIGHT_CONFIG_DETAIL_DF = pd.concat(validation_details, ignore_index=True) if validation_details else pd.DataFrame()

with pd.ExcelWriter(WEIGHT_VALIDATION_PATH, engine="openpyxl") as writer:
    WEIGHT_CONFIG_VALIDATION_DF.to_excel(writer, sheet_name="summary", index=False)
    WEIGHT_CONFIG_DETAIL_DF.to_excel(writer, sheet_name="detail", index=False)
    pd.DataFrame(MODEL_WEIGHT_CONFIGS).T.to_excel(writer, sheet_name="configs")

print(f"Weight config validation saved: {WEIGHT_VALIDATION_PATH}")
print(f"Ground Truth JD {GROUND_TRUTH_DF['jd_id'].astype(str).nunique()}개 기준으로 가중치 설정을 비교했습니다.")
display(WEIGHT_CONFIG_VALIDATION_DF)


# In[24]:


# Cell 23 | Grid Search - Hybrid score optimization

import itertools
import json
import numpy as np
import pandas as pd

GRID_SEARCH_REPORT_PATH = EVALUATION_DIR / "grid_search_hybrid_v26.xlsx"
GRID_SEARCH_K_VALUES = tuple(globals().get("EVALUATION_K_VALUES", (3, 5, 10, 20, 50)))
GRID_OBJECTIVE_K_MAIN = 3
GRID_OBJECTIVE_K_SUB = 5
HYBRID_EQUAL_SCORE_COL = "hybrid_equal_sum_score"
HYBRID_GRID_SCORE_COL = "hybrid_grid_best_score"


def _numeric_series(df, col, default=0.0):
    if col not in df.columns:
        return pd.Series(default, index=df.index, dtype=float)
    return pd.to_numeric(df[col], errors="coerce").fillna(default).astype(float)


def _safe_metric(summary_df, k, metric_col):
    s = summary_df.loc[summary_df["k"].astype(int) == int(k), metric_col]
    if s.empty or pd.isna(s.iloc[0]):
        return 0.0
    return float(s.iloc[0])


HYBRID_GRID_BASE_DF = pair_df.copy()
HYBRID_GRID_BASE_DF["_keyword_component"] = _numeric_series(HYBRID_GRID_BASE_DF, "keyword_baseline_score_norm")
HYBRID_GRID_BASE_DF["_embedding_component"] = _numeric_series(HYBRID_GRID_BASE_DF, "embedding_retrieval_score_norm")
HYBRID_GRID_BASE_DF["_cross_encoder_component"] = _numeric_series(HYBRID_GRID_BASE_DF, "column_score")
HYBRID_GRID_BASE_DF[HYBRID_EQUAL_SCORE_COL] = HYBRID_GRID_BASE_DF[["_keyword_component", "_embedding_component", "_cross_encoder_component"]].mean(axis=1)

GRID_KEYWORD_WEIGHTS = [0.05, 0.10, 0.15, 0.20, 0.25]
GRID_EMBEDDING_WEIGHTS = [0.55, 0.60, 0.65, 0.70, 0.75, 0.80]
GRID_CE_MIN = 0.10
GRID_CE_MAX = 0.35

_grid_summary_rows = []
_grid_detail_frames = []
_grid_scored_frames = {}
_grid_id = 0

for keyword_w, embedding_w in itertools.product(GRID_KEYWORD_WEIGHTS, GRID_EMBEDDING_WEIGHTS):
    cross_encoder_w = round(1.0 - keyword_w - embedding_w, 10)
    if cross_encoder_w < GRID_CE_MIN or cross_encoder_w > GRID_CE_MAX:
        continue

    _grid_id += 1
    grid_name = f"grid_{_grid_id:03d}"
    score_col = f"hybrid_grid_score_{_grid_id:03d}"

    scored_df = HYBRID_GRID_BASE_DF.copy()
    scored_df[score_col] = (
        keyword_w * scored_df["_keyword_component"] +
        embedding_w * scored_df["_embedding_component"] +
        cross_encoder_w * scored_df["_cross_encoder_component"]
    )

    detail_df, summary_df, _ = evaluate_ranking(
        scored_df,
        GROUND_TRUTH_DF,
        score_col=score_col,
        k_values=GRID_SEARCH_K_VALUES,
        positive_threshold=POSITIVE_RELEVANCE_THRESHOLD,
    )

    objective_score = (
        0.45 * _safe_metric(summary_df, GRID_OBJECTIVE_K_MAIN, "macro_ndcg_at_k") +
        0.35 * _safe_metric(summary_df, GRID_OBJECTIVE_K_MAIN, "macro_mrr_at_k") +
        0.15 * _safe_metric(summary_df, GRID_OBJECTIVE_K_SUB, "macro_ndcg_at_k") +
        0.05 * _safe_metric(summary_df, GRID_OBJECTIVE_K_SUB, "macro_mrr_at_k")
    )

    detail_df.insert(0, "grid_id", _grid_id)
    detail_df.insert(1, "grid_name", grid_name)
    _grid_detail_frames.append(detail_df)
    _grid_scored_frames[_grid_id] = scored_df[["jd_id", "candidate_id", score_col]].copy()

    for _, row in summary_df.iterrows():
        _grid_summary_rows.append({
            "grid_id": _grid_id,
            "grid_name": grid_name,
            "k": int(row["k"]),
            "keyword_weight": keyword_w,
            "embedding_weight": embedding_w,
            "cross_encoder_weight": cross_encoder_w,
            "objective_score": objective_score,
            "macro_precision_at_k": row["macro_precision_at_k"],
            "macro_recall_at_k": row["macro_recall_at_k"],
            "macro_ndcg_at_k": row["macro_ndcg_at_k"],
            "macro_mrr_at_k": row["macro_mrr_at_k"],
        })

GRID_SEARCH_FULL_DF = pd.DataFrame(_grid_summary_rows)
GRID_SEARCH_DETAIL_DF = pd.concat(_grid_detail_frames, ignore_index=True) if _grid_detail_frames else pd.DataFrame()

if GRID_SEARCH_FULL_DF.empty:
    raise ValueError("Grid search 결과가 없습니다. weight 범위를 확인하세요.")

GRID_SEARCH_BEST_CONFIG_DF = (
    GRID_SEARCH_FULL_DF[["grid_id", "grid_name", "keyword_weight", "embedding_weight", "cross_encoder_weight", "objective_score"]]
    .drop_duplicates()
    .sort_values(["objective_score", "embedding_weight", "cross_encoder_weight"], ascending=[False, False, False])
    .reset_index(drop=True)
)

BEST_GRID_SEARCH_CONFIG = GRID_SEARCH_BEST_CONFIG_DF.iloc[0].to_dict()
BEST_GRID_SEARCH_ID = int(BEST_GRID_SEARCH_CONFIG["grid_id"])
BEST_GRID_SEARCH_SCORE_SOURCE_COL = f"hybrid_grid_score_{BEST_GRID_SEARCH_ID:03d}"

HYBRID_GRID_BEST_DF = HYBRID_GRID_BASE_DF.copy()
HYBRID_GRID_BEST_DF[HYBRID_GRID_SCORE_COL] = _grid_scored_frames[BEST_GRID_SEARCH_ID][BEST_GRID_SEARCH_SCORE_SOURCE_COL].values
HYBRID_GRID_BEST_DF = HYBRID_GRID_BEST_DF.sort_values(["jd_id", HYBRID_GRID_SCORE_COL], ascending=[True, False]).copy()
HYBRID_GRID_BEST_DF["hybrid_grid_rank"] = HYBRID_GRID_BEST_DF.groupby("jd_id").cumcount() + 1


# In[25]:


# Cell 24 | Keyword baseline vs Hybrid 성능 비교

from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment

keyword_eval_for_hybrid_df = globals().get("KEYWORD_SCORE_DF", keyword_eval_df if "keyword_eval_df" in globals() else pair_df).copy()
HYBRID_EQUAL_DF = HYBRID_GRID_BASE_DF.copy()

HYBRID_MODEL_SCORE_SPECS = [
    {
        "group_name": "Keyword baseline",
        "model_name": "1) BM25+TF-IDF avg",
        "score_col": "keyword_baseline_score",
        "pred_df": keyword_eval_for_hybrid_df,
        "candidate_scope": "KEYWORD_SCORE_DF_all_candidates" if "KEYWORD_SCORE_DF" in globals() else "pair_df_top_candidates",
    },
    {
        "group_name": "Hybrid before grid",
        "model_name": "2) BM25+TF-IDF+Embedding+CrossEncoder equal avg",
        "score_col": HYBRID_EQUAL_SCORE_COL,
        "pred_df": HYBRID_EQUAL_DF,
        "candidate_scope": "pair_df_top_candidates",
    },
    {
        "group_name": "Hybrid before grid",
        "model_name": "Final score before grid",
        "score_col": "final_score",
        "pred_df": pair_df,
        "candidate_scope": "pair_df_top_candidates",
    },
    {
        "group_name": "Hybrid after grid",
        "model_name": "2) BM25+TF-IDF+Embedding+CrossEncoder grid best",
        "score_col": HYBRID_GRID_SCORE_COL,
        "pred_df": HYBRID_GRID_BEST_DF,
        "candidate_scope": "pair_df_top_candidates",
    },
]

HYBRID_MODEL_COMPARISON_SUMMARY_DF, HYBRID_MODEL_COMPARISON_DETAIL_DF = evaluate_multiple_model_specs(
    HYBRID_MODEL_SCORE_SPECS,
    GROUND_TRUTH_DF,
    k_values=GRID_SEARCH_K_VALUES,
    positive_threshold=POSITIVE_RELEVANCE_THRESHOLD,
)

baseline_summary = (
    HYBRID_MODEL_COMPARISON_SUMMARY_DF[HYBRID_MODEL_COMPARISON_SUMMARY_DF["model_name"] == "1) BM25+TF-IDF avg"]
    [["k", "macro_precision_at_k", "macro_recall_at_k", "macro_ndcg_at_k", "macro_mrr_at_k"]]
    .rename(columns={
        "macro_precision_at_k": "baseline_precision_at_k",
        "macro_recall_at_k": "baseline_recall_at_k",
        "macro_ndcg_at_k": "baseline_ndcg_at_k",
        "macro_mrr_at_k": "baseline_mrr_at_k",
    })
)

HYBRID_COMPARISON_VIEW_DF = HYBRID_MODEL_COMPARISON_SUMMARY_DF.merge(baseline_summary, on="k", how="left")
for metric in ["precision", "recall", "ndcg", "mrr"]:
    HYBRID_COMPARISON_VIEW_DF[f"delta_{metric}_vs_keyword"] = (
        HYBRID_COMPARISON_VIEW_DF[f"macro_{metric}_at_k"] -
        HYBRID_COMPARISON_VIEW_DF[f"baseline_{metric}_at_k"]
    )

HYBRID_COMPARISON_VIEW_DF = HYBRID_COMPARISON_VIEW_DF[[
    "group_name", "model_name", "k",
    "macro_precision_at_k", "macro_recall_at_k", "macro_ndcg_at_k", "macro_mrr_at_k",
    "delta_precision_vs_keyword", "delta_recall_vs_keyword", "delta_ndcg_vs_keyword", "delta_mrr_vs_keyword",
]].sort_values(["k", "group_name", "model_name"]).reset_index(drop=True)

BEST_GRID_ROW_DF = pd.DataFrame([BEST_GRID_SEARCH_CONFIG])

chart_metric_sources = {
    "precision": "macro_precision_at_k",
    "recall": "macro_recall_at_k",
    "ndcg": "macro_ndcg_at_k",
    "mrr": "macro_mrr_at_k",
}

chart_wide_frames = {}
for metric_name, metric_col in chart_metric_sources.items():
    chart_wide_frames[metric_name] = (
        HYBRID_MODEL_COMPARISON_SUMMARY_DF
        .pivot_table(index="k", columns="model_name", values=metric_col, aggfunc="mean")
        .reset_index()
        .sort_values("k")
    )

with pd.ExcelWriter(GRID_SEARCH_REPORT_PATH, engine="openpyxl") as writer:
    HYBRID_COMPARISON_VIEW_DF.to_excel(writer, sheet_name="comparison", index=False)
    GRID_SEARCH_FULL_DF.to_excel(writer, sheet_name="grid_full", index=False)
    GRID_SEARCH_DETAIL_DF.to_excel(writer, sheet_name="grid_detail", index=False)
    GRID_SEARCH_BEST_CONFIG_DF.to_excel(writer, sheet_name="best_config", index=False)
    HYBRID_MODEL_COMPARISON_SUMMARY_DF.to_excel(writer, sheet_name="model_summary", index=False)
    HYBRID_MODEL_COMPARISON_DETAIL_DF.to_excel(writer, sheet_name="model_detail", index=False)
    for metric_name, metric_df in chart_wide_frames.items():
        metric_df.to_excel(writer, sheet_name=f"chart_{metric_name}", index=False)

wb = load_workbook(GRID_SEARCH_REPORT_PATH)
for ws in wb.worksheets:
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 42)

chart_ws = wb.create_sheet("charts") if "charts" not in wb.sheetnames else wb["charts"]
chart_positions = {
    "precision": "A1",
    "recall": "A18",
    "ndcg": "J1",
    "mrr": "J18",
}
chart_titles = {
    "precision": "Precision@K",
    "recall": "Recall@K",
    "ndcg": "NDCG@K",
    "mrr": "MRR@K",
}

for metric_name, pos in chart_positions.items():
    ws = wb[f"chart_{metric_name}"]
    chart = LineChart()
    chart.title = chart_titles[metric_name]
    chart.y_axis.title = chart_titles[metric_name]
    chart.x_axis.title = "K"
    chart.height = 8
    chart.width = 16
    data = Reference(ws, min_col=2, max_col=ws.max_column, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend.position = "b"
    chart_ws.add_chart(chart, pos)

wb.save(GRID_SEARCH_REPORT_PATH)

_display_cols = [
    "group_name", "model_name", "k",
    "macro_precision_at_k", "macro_recall_at_k", "macro_ndcg_at_k", "macro_mrr_at_k",
    "delta_ndcg_vs_keyword", "delta_mrr_vs_keyword",
]
display(HYBRID_COMPARISON_VIEW_DF[_display_cols])


# In[26]:


# Cell 25 | LLM/API 역할 Proxy Rerank 비교
# ------------------------------------------------------------
# 실제 API 키 없이, LLM이 추가로 볼 법한 의미 해석 기준을 rule/proxy score로 흉내냅니다.
# Ground Truth는 점수 산정에 사용하지 않고, 마지막 평가에만 사용합니다.
#
# Proxy 해석 기준:
# - 기존 final_score를 기본 신호로 둡니다.
# - Skill/Task/Role/AI overlap/Experience/Soft signal을 다시 가중합합니다.
# - Engineer JD에서 researcher/scientist-only 후보가 상위로 치고 올라오는 경우를 더 강하게 감점합니다.
# - production/API/RAG/agent/AWS/impact 같은 실무 evidence 텍스트가 있으면 소폭 가산합니다.
# ------------------------------------------------------------

LLM_PROXY_REPORT_PATH = EVALUATION_DIR / "llm_proxy_comparison_latest.xlsx"


def _text_has_any(text, keywords):
    text = "" if pd.isna(text) else str(text).lower()
    return any(k.lower() in text for k in keywords)


PRODUCTION_EVIDENCE_KEYWORDS = [
    "production", "프로덕션", "운영", "배포", "api", "fastapi", "django", "next.js", "react",
    "aws", "cloudwatch", "eks", "ci/cd", "cicd", "docker", "kubernetes",
    "rag", "langgraph", "langchain", "agent", "mcp", "vector", "pgvector", "qdrant", "milvus",
    "성능", "개선", "%", "절감", "단축", "향상", "고객", "납품", "poc",
]


def llm_proxy_semantic_score(row):
    base = float(row.get("final_score", 0) or 0)
    semantic = (
        0.25 * float(row.get("skill_score", 0) or 0) +
        0.25 * float(row.get("task_score", 0) or 0) +
        0.15 * float(row.get("role_score", 0) or 0) +
        0.15 * float(row.get("ai_skill_overlap_score", 0) or 0) +
        0.10 * float(row.get("experience_score", 0) or 0) +
        0.10 * float(row.get("soft_skill_score", 0) or 0)
    )

    evidence_text = "\n".join(str(row.get(c, "")) for c in [
        "cv_skill", "cv_career", "cv_career_description", "cv_position", "cv_self_introduction",
        "jd_required_skill", "jd_main_task", "jd_preferred", "jd_clean_text",
    ])
    evidence_bonus = 0.04 if _text_has_any(evidence_text, PRODUCTION_EVIDENCE_KEYWORDS) else 0.0

    penalty = 1.0
    role_penalty = float(row.get("role_mismatch_penalty", 1.0) or 1.0)
    if role_penalty <= 0.60:
        penalty *= 0.72
    elif role_penalty < 0.80:
        penalty *= 0.82
    elif role_penalty < 0.90:
        penalty *= 0.90
    elif role_penalty < 1.0:
        penalty *= 0.96

    if bool(row.get("skill_risk_flag", False)):
        penalty *= 0.88

    proxy = (0.55 * base + 0.45 * semantic + evidence_bonus) * penalty
    return float(np.clip(proxy, 0.0, 1.0))


LLM_PROXY_DF = pair_df.copy()
LLM_PROXY_DF["llm_proxy_score"] = LLM_PROXY_DF.apply(llm_proxy_semantic_score, axis=1)
LLM_PROXY_DF = LLM_PROXY_DF.sort_values(["jd_id", "llm_proxy_score"], ascending=[True, False]).copy()
LLM_PROXY_DF["llm_proxy_rank"] = LLM_PROXY_DF.groupby("jd_id").cumcount() + 1

LLM_PROXY_DETAIL_DF, LLM_PROXY_SUMMARY_DF, LLM_PROXY_COVERAGE_DF = evaluate_ranking(
    LLM_PROXY_DF,
    GROUND_TRUTH_DF,
    score_col="llm_proxy_score",
    k_values=EVALUATION_K_VALUES,
    positive_threshold=POSITIVE_RELEVANCE_THRESHOLD,
)

BASELINE_FOR_PROXY_DETAIL_DF, BASELINE_FOR_PROXY_SUMMARY_DF, _ = evaluate_ranking(
    pair_df,
    GROUND_TRUTH_DF,
    score_col="final_score",
    k_values=EVALUATION_K_VALUES,
    positive_threshold=POSITIVE_RELEVANCE_THRESHOLD,
)

LLM_PROXY_COMPARISON_DF = BASELINE_FOR_PROXY_SUMMARY_DF.merge(
    LLM_PROXY_SUMMARY_DF,
    on="k",
    suffixes=("_baseline", "_llm_proxy"),
)
for metric in ["macro_precision_at_k", "macro_recall_at_k", "macro_ndcg_at_k", "macro_mrr_at_k"]:
    LLM_PROXY_COMPARISON_DF[f"delta_{metric}"] = (
        LLM_PROXY_COMPARISON_DF[f"{metric}_llm_proxy"] -
        LLM_PROXY_COMPARISON_DF[f"{metric}_baseline"]
    )

with pd.ExcelWriter(LLM_PROXY_REPORT_PATH, engine="openpyxl") as writer:
    LLM_PROXY_COMPARISON_DF.to_excel(writer, sheet_name="comparison", index=False)
    LLM_PROXY_SUMMARY_DF.to_excel(writer, sheet_name="llm_proxy_summary", index=False)
    LLM_PROXY_DETAIL_DF.to_excel(writer, sheet_name="llm_proxy_detail", index=False)
    LLM_PROXY_DF.sort_values(["jd_id", "llm_proxy_rank"]).head(300).to_excel(writer, sheet_name="llm_proxy_top_candidates", index=False)
    BASELINE_FOR_PROXY_SUMMARY_DF.to_excel(writer, sheet_name="baseline_summary", index=False)

print(f"LLM proxy comparison saved: {LLM_PROXY_REPORT_PATH}")
display(LLM_PROXY_COMPARISON_DF)
display(LLM_PROXY_DF.sort_values(["jd_id", "llm_proxy_rank"])[["jd_id", "candidate_id", "llm_proxy_rank", "llm_proxy_score", "final_score", "skill_score", "task_score", "role_score", "ai_skill_overlap_score"]].head(30))


# In[27]:


# Cell 26 | 대시보드용 결과 테이블 생성
# ------------------------------------------------------------
# 이 result_df가 BI 대시보드/Tableau/Looker/Streamlit 등에 연결할 기본 fact table입니다.
# JD 내용과 CV 내용, 점수, 표준화 컬럼, 해석 컬럼이 함께 들어갑니다.
# ------------------------------------------------------------

def pick_col(df, col, default=""):
    if col in df.columns:
        return df[col].fillna("").astype(str)
    return pd.Series([default] * len(df), index=df.index)

result_df = pd.DataFrame({
    "match_id": top_match_df["jd_id"].astype(str) + "__" + top_match_df["candidate_id"].astype(str),
    "jd_id": top_match_df["jd_id"],
    "candidate_id": top_match_df["candidate_id"],
    "rank": top_match_df["rank"],

    # score
    "retrieval_score": top_match_df["retrieval_score"],
    "retrieval_score_norm": top_match_df["retrieval_score_norm"],
    "column_score": top_match_df["column_score"],
    "final_score": top_match_df["final_score"],
    "skill_score": top_match_df["skill_score"],
    "task_score": top_match_df["task_score"],
    "domain_score": top_match_df["domain_score"],
    "role_score": top_match_df["role_score"],
    "ai_skill_overlap_score": top_match_df["ai_skill_overlap_score"],
    "experience_score": top_match_df.get("experience_score", pd.Series([0.0] * len(top_match_df), index=top_match_df.index)),
    "soft_skill_score": top_match_df.get("soft_skill_score", pd.Series([0.0] * len(top_match_df), index=top_match_df.index)),
    "radar_skill_score": top_match_df.get("radar_skill_score", top_match_df["skill_score"]),
    "radar_task_score": top_match_df.get("radar_task_score", top_match_df["task_score"]),
    "radar_domain_score": top_match_df.get("radar_domain_score", top_match_df["domain_score"]),
    "radar_role_score": top_match_df.get("radar_role_score", top_match_df["role_score"]),
    "radar_experience_score": top_match_df.get("radar_experience_score", pd.Series([0.0] * len(top_match_df), index=top_match_df.index)),
    "radar_soft_skill_score": top_match_df.get("radar_soft_skill_score", pd.Series([0.0] * len(top_match_df), index=top_match_df.index)),
    "skill_risk_flag": top_match_df["skill_risk_flag"],
    "scoring_method": top_match_df["scoring_method"],
    "retrieval_method": top_match_df["retrieval_method"],

    # JD structured info
    "jd_company": pick_col(top_match_df, "company"),
    "jd_title": pick_col(top_match_df, "job_title"),
    "jd_career_requirement": pick_col(top_match_df, "career_requirement"),
    "jd_required_skill": pick_col(top_match_df, "required_skill"),
    "jd_main_task": pick_col(top_match_df, "main_task"),
    "jd_role_signal": pick_col(top_match_df, "role_signal"),
    "jd_job_family": pick_col(top_match_df, "job_family"),
    "jd_domain": pick_col(top_match_df, "domain"),
    "jd_deliverable_signal": pick_col(top_match_df, "deliverable_signal"),
    "jd_project_signal": pick_col(top_match_df, "project_signal"),
    "jd_qualification": pick_col(top_match_df, "qualification"),
    "jd_preferred": pick_col(top_match_df, "preferred"),
    "jd_clean_text": pick_col(top_match_df, "clean_text"),

    # JD taxonomy
    "jd_skill_standard": pick_col(top_match_df, "jd_skill_standard"),
    "jd_task_standard": pick_col(top_match_df, "jd_task_standard"),
    "jd_domain_standard": pick_col(top_match_df, "jd_domain_standard"),
    "jd_role_standard": pick_col(top_match_df, "jd_role_standard"),
    "jd_ai_skill_tags": pick_col(top_match_df, "jd_ai_skill_tags"),

    # CV structured info
    "cv_job_role": pick_col(top_match_df, "job_role"),
    "cv_file_name": pick_col(top_match_df, "file_name"),
    "cv_education": pick_col(top_match_df, "Education"),
    "cv_gpa": pick_col(top_match_df, "GPA"),
    "cv_certificates": pick_col(top_match_df, "Certificates"),
    "cv_languages": pick_col(top_match_df, "Languages"),
    "cv_paper": pick_col(top_match_df, "Paper"),
    "cv_skill": pick_col(top_match_df, "Skill"),
    "cv_overseas_experience": pick_col(top_match_df, "Overseas_Experience"),
    "cv_training": pick_col(top_match_df, "Training"),
    "cv_awards": pick_col(top_match_df, "Awards"),
    "cv_total_career_years": pick_col(top_match_df, "Total_Career_Years"),
    "cv_career": pick_col(top_match_df, "Career"),
    "cv_career_description": pick_col(top_match_df, "Career_Description"),
    "cv_position": pick_col(top_match_df, "Position"),
    "cv_employment_status": pick_col(top_match_df, "Employment_Status"),
    "cv_salary": pick_col(top_match_df, "Salary"),
    "cv_self_introduction": pick_col(top_match_df, "Self_Introduction"),
    "cv_portfolio": pick_col(top_match_df, "Portfolio"),
    "cv_raw_text": pick_col(top_match_df, "Raw_Text"),

    # CV taxonomy
    "cv_skill_standard": pick_col(top_match_df, "cv_skill_standard"),
    "cv_task_standard": pick_col(top_match_df, "cv_task_standard"),
    "cv_domain_standard": pick_col(top_match_df, "cv_domain_standard"),
    "cv_role_standard": pick_col(top_match_df, "cv_role_standard"),
    "cv_ai_skill_tags": pick_col(top_match_df, "cv_ai_skill_tags"),
})

# 대시보드에서 바로 보기 편하도록 점수는 0~100도 함께 생성
score_100_cols = [
    "final_score", "column_score",
    "skill_score", "task_score", "domain_score", "role_score",
    "ai_skill_overlap_score", "experience_score", "soft_skill_score",
    "radar_skill_score", "radar_task_score", "radar_domain_score",
    "radar_role_score", "radar_experience_score", "radar_soft_skill_score",
    "retrieval_score", "retrieval_score_norm",
]

for col in score_100_cols:
    if col in result_df.columns:
        result_df[col + "_100"] = (pd.to_numeric(result_df[col], errors="coerce").fillna(0.0) * 100).round(1)

print("result_df:", result_df.shape)
display(result_df.head(10))


# In[28]:


# Cell 27 | Rule-based 추천 근거 / Gap / 면접 질문 생성
# ------------------------------------------------------------
# LLM을 쓰기 전에도 대시보드에 최소한의 설명이 나오도록 규칙 기반 문장을 생성합니다.
# 이후 LLM 해석이 붙으면 이 컬럼은 baseline explanation 역할을 합니다.
# ------------------------------------------------------------

def split_terms(x):
    return set([t.strip() for t in str(x).split(",") if t.strip()])


def overlap_terms(a, b):
    return sorted(split_terms(a).intersection(split_terms(b)))


def build_rule_based_reason(row):
    reasons = []
    skill_overlap = overlap_terms(row.get("jd_skill_standard", ""), row.get("cv_skill_standard", ""))
    task_overlap = overlap_terms(row.get("jd_task_standard", ""), row.get("cv_task_standard", ""))
    domain_overlap = overlap_terms(row.get("jd_domain_standard", ""), row.get("cv_domain_standard", ""))
    role_overlap = overlap_terms(row.get("jd_role_standard", ""), row.get("cv_role_standard", ""))

    if skill_overlap:
        reasons.append(f"Skill 일치: {', '.join(skill_overlap)}")
    elif row.get("cv_skill_standard"):
        reasons.append(f"Skill 후보: {row.get('cv_skill_standard')}")

    ai_overlap = overlap_terms(row.get("jd_ai_skill_tags", ""), row.get("cv_ai_skill_tags", ""))
    if ai_overlap:
        reasons.append(f"AI 핵심 태그 일치: {', '.join(ai_overlap)}")

    if task_overlap:
        reasons.append(f"Task 일치: {', '.join(task_overlap)}")
    elif row.get("cv_task_standard"):
        reasons.append(f"Task 관련 경험: {row.get('cv_task_standard')}")

    if domain_overlap:
        reasons.append(f"Domain 일치: {', '.join(domain_overlap)}")
    if role_overlap:
        reasons.append(f"Role 일치: {', '.join(role_overlap)}")

    if row.get("cv_career_description"):
        reasons.append("CV 근거 일부: " + str(row.get("cv_career_description"))[:220])

    return "\n".join(reasons) if reasons else "규칙 기반으로 명확한 근거가 적어 LLM 또는 면접 확인이 필요합니다."


def build_gap(row):
    gaps = []
    jd_skill = split_terms(row.get("jd_skill_standard", ""))
    cv_skill = split_terms(row.get("cv_skill_standard", ""))
    missing_skill = sorted(jd_skill - cv_skill)
    if missing_skill:
        gaps.append("JD 요구 스킬 중 CV 표준화 결과에서 명시 확인이 약한 항목: " + ", ".join(missing_skill))

    jd_ai = split_terms(row.get("jd_ai_skill_tags", ""))
    cv_ai = split_terms(row.get("cv_ai_skill_tags", ""))
    missing_ai = sorted(jd_ai - cv_ai)
    if missing_ai:
        gaps.append("AI/LLM 핵심 태그 중 CV에서 약하게 보이는 항목: " + ", ".join(missing_ai))

    if bool(row.get("skill_risk_flag", False)):
        gaps.append("AI Engineer 계열 JD인데 핵심 기술 태그와 Skill 점수가 모두 낮아 우선순위 감점됨")

    if row.get("domain_score", 0) < 0.3:
        gaps.append("도메인 유사도가 낮아 실제 산업/서비스 맥락 확인 필요")
    if row.get("role_score", 0) < 0.3:
        gaps.append("직무 타이틀/역할 적합도 추가 확인 필요")
    return "\n".join(gaps) if gaps else "큰 Gap은 규칙 기반에서 확인되지 않음"


def build_interview_question(row):
    questions = []
    if row.get("skill_score", 0) < 0.5 or row.get("ai_skill_overlap_score", 0) < 0.5:
        questions.append("JD 핵심 요구 스킬/AI 태그를 실제 프로젝트에서 어느 수준까지 사용했는지 구체적인 산출물 기준으로 설명해 주세요.")
    if row.get("task_score", 0) < 0.5:
        questions.append("JD 주요업무와 유사한 과제를 수행한 경험이 있다면 본인 역할과 결과물을 설명해 주세요.")
    if row.get("domain_score", 0) < 0.5:
        questions.append("해당 도메인 또는 유사 산업에서의 문제 해결 경험이 있는지 확인해 주세요.")
    if not questions:
        questions.append("가장 유사한 프로젝트에서 본인이 직접 담당한 범위와 성과 수치를 확인해 주세요.")
    return "\n".join([f"- {q}" for q in questions])

result_df["match_reason_rule_based"] = result_df.apply(build_rule_based_reason, axis=1)
result_df["match_gap_rule_based"] = result_df.apply(build_gap, axis=1)
result_df["interview_question_rule_based"] = result_df.apply(build_interview_question, axis=1)

summary_cols = [
    "jd_id", "jd_title", "candidate_id", "rank", "final_score_100", "skill_score_100", "ai_skill_overlap_score_100", "task_score_100", "domain_score_100", "role_score_100",
    "match_reason_rule_based", "match_gap_rule_based", "interview_question_rule_based"
]
display(result_df[summary_cols].head(10))


# In[29]:


# Cell 28 | 하드스킬 Evidence Level 시스템 프롬프트 정의
# ------------------------------------------------------------
# 이 프롬프트는 규칙 기반 매칭으로 추출된 hard skill 후보와 CV 원문을 받아
# 각 skill의 V1 / V2 Evidence Level을 판정하도록 설계되었습니다.
# ------------------------------------------------------------

HARDSKILL_EVIDENCE_SYSTEM_PROMPT = """
당신은 CV에서 하드스킬의 Evidence Level을 판정하는 전문 평가자입니다.
규칙 기반 매칭으로 추출된 스킬 후보 목록과 CV 원문을 받아,
각 스킬의 V1 / V2 여부를 판정하고 반드시 CV 원문을 근거로 인용하세요.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
[Evidence Level 정의]
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

미인정 (점수 없음 / 매칭 취소)
  조건: 스킬명 나열, 또는 프로젝트·경력 맥락 자체가 없음
  기준: 쉼표·공백으로 구분된 스킬명이 나열된 행
       또는 스킬 이름만 있고 어떤 업무 문장도 없는 경우
  처리: 해당 스킬은 판정 대상에서 제외 (결과에 포함하지 않음)

V1 — 적용 검증 (계수 × 0.70)
  조건: 본인 역할(내가 직접 한 행위) + 산출물(결과물의 이름·형태) 이 함께 명시됨
  판정 질문: "이 문장에서 누가 무엇을 만들었는가?"
             → 주어가 본인이고 + 결과물 이름이 있으면 V1

V2 — 성과 검증 (계수 × 1.00)
  조건: V1 조건 + 아래 성과 유형 중 하나 이상 확인됨

  [정량 성과]
  - 수치: %, 배수, 건수, 금액, 시간 단축 등 숫자가 포함된 결과

  [정성 성과]
  - 운영·배포: 실제 서비스에 배포, 운영 중, 프로덕션 반영
  - 고객·외부 적용: 고객사 납품, 외부 기관 제출, 파일럿 배포
  - 공식 인정: 인증 취득, 특허 출원, 수상, 정책 회의 채택

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
[공통 미인정 처리 — 아래 패턴은 V1 이상을 주지 않는다]
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

① 스킬명 나열
   "Python SQL R Pandas Numpy"
   "PyTorch TensorFlow Keras LangChain"
   → 판정 제외

② 단순 경험 선언
   "LangChain 활용 경험 보유"
   "RAG 관련 업무 담당"
   → 판정 제외

③ 참여 서술 (본인 역할 불명확)
   "RAG 프로젝트 참여"
   "팀에서 RAG 시스템을 구축했습니다"
   → 판정 제외

④ 학습·수료 (실무 적용 없음)
   "Python 교육 과정 수료"
   "LLM 실습 과정 이수"
   → 판정 제외

⑤ 의향·계획
   "LangChain을 활용할 예정입니다"
   → 판정 제외

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
[소프트스킬 문장 공유 기준]
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

소프트스킬(S1~S5) 판정 근거 문장이 특정 하드스킬을 직접 언급하고 있으면,
그 문장을 해당 하드스킬의 V1/V2 근거로 사용할 수 있습니다.

인정 예시:
  "Tableau 대시보드를 구축해 월간 보고서 공수를 1.5일→0.5일로 단축"
  → S5(비즈니스 임팩트) 근거로도 사용 가능
  → S035(Dashboard) V2 근거로도 사용 가능  ← Tableau가 직접 언급됨

불인정 예시:
  "탈회율 15% 급증을 발견하고 지시 없이 원인분석 착수, 익월 정상화"
  → S1(자기주도) 근거로 사용 가능
  → 특정 하드스킬이 언급되지 않음 → 하드스킬 V 판정에 사용 불가

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
[인정 예시 — V1]
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

예시 1 — V1 (RAG)
  "LangChain과 FAISS를 활용해 사내 문서 검색 RAG API를 구현"
  → 역할(본인 구현) + 산출물(RAG API) → V1
  → 수치·배포·채택 없으므로 V2 아님

예시 2 — V1 (Python)
  "Python으로 고객 세그먼트 분석 리포트를 작성하여 마케팅 팀에 제공"
  → 역할(본인 작성) + 산출물(분석 리포트) → V1

예시 3 — V1 (Figma)
  "Figma로 모바일 앱 주요 화면 와이어프레임을 설계하여 개발팀에 전달"
  → 역할(본인 설계) + 산출물(와이어프레임) → V1

예시 4 — V1 (Tableau)
  "Tableau 기반 매출 대시보드를 구축하여 주간 보고에 활용"
  → 역할(본인 구축) + 산출물(대시보드) → V1
  → "주간 보고에 활용"은 성과 아님 → V2 아님

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
[인정 예시 — V2]
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

예시 1 — V2 정량 (RAG)
  "RAG 기반 고객지원 챗봇을 구축하고 검색 정확도를 20% 개선, 고객사 PoC에 배포"
  → 역할 + 산출물 + 수치(20%) + 고객사 배포 → V2

예시 2 — V2 정량 (Tableau)
  "Tableau 대시보드를 구축해 월간 보고서 작성 공수를 1.5일에서 0.5일로 단축"
  → 역할 + 산출물 + 시간 단축 수치 → V2

예시 3 — V2 정량 (Python)
  "Python으로 거래 복귀 확률 예측 모델을 개발하여 캠페인 전환율 2.44배 달성"
  → 역할 + 산출물 + 배수 수치 → V2

예시 4 — V2 정량 (SQL)
  "SQL 쿼리 자동화로 오지급 오류 200건을 방지하고 월간 검증 시간 2시간 단축"
  → 역할 + 산출물 + 건수·시간 수치 → V2

예시 5 — V2 정성 (납품)
  "Computer Vision 기반 생체정보 측정 시스템을 개발해 공공기관 42대 납품"
  → 역할 + 산출물 + 납품(외부 적용) → V2

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
[불인정 예시]
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

불인정 1 — 미인정 (LangChain)
  "LangChain 활용 경험 보유"
  → 스킬명 + 경험 선언 → 판정 제외

불인정 2 — 미인정 (PyTorch)
  "딥러닝 프로젝트에 PyTorch를 사용했습니다"
  → 어떤 모델을 어떻게 만들었는지 없음 → 판정 제외

불인정 3 — 미인정 (AI Agent)
  "AI Agent 시스템 개발 참여"
  → "참여"는 본인 역할 불명확 → 판정 제외

불인정 4 — V1이지만 V2 불인정 (Tableau)
  "Tableau 기반 매출 대시보드를 구축하여 주간 보고에 활용"
  → 역할·산출물 있음 → V1
  → "주간 보고에 활용"은 수치·배포·채택 없음 → V2 불인정

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
[인터뷰 플래그 조건]
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

아래 조건이면 interview_flag: true 를 세운다:

- 팀 단위 서술에서 본인 기여 범위가 불명확할 때
  질문: "이 시스템에서 본인이 직접 설계하거나 구현한 부분이 어디인가요?"

- 산출물 이름은 있는데 본인 역할이 생략됐을 때
  질문: "이 프로젝트에서 본인의 담당 업무가 구체적으로 무엇이었나요?"

- V2 성과 수치가 있는데 맥락이 불충분할 때
  질문: "이 수치는 어떤 모델/지표를 기준으로 측정한 것인가요?"

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
[출력 형식 — JSON만 반환, 다른 텍스트 없음]
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

반드시 아래 JSON 형식으로만 응답하라. 다른 텍스트를 추가하지 말 것.
미인정 스킬은 skill_evidence 배열에 포함하지 않고 excluded_skills에 넣는다.

{
  "skill_evidence": [
    {
      "skill_id": "S003",
      "skill_name": "RAG",
      "V_level": "V1 또는 V2",
      "V_coeff": 0.70 또는 1.00,
      "evidence_text": "CV 원문 직접 인용 (최대 2문장)",
      "reason": "판정 이유 한 줄 — 역할/산출물/성과 근거 명시",
      "interview_flag": true 또는 false,
      "interview_question": "플래그가 true일 때만 작성, 아니면 null"
    }
  ],
  "excluded_skills": [
    {
      "skill_id": "S002",
      "skill_name": "Python",
      "reason": "스킬명 나열 패턴으로 판정 제외"
    }
  ]
}
"""

print("하드스킬 Evidence 시스템 프롬프트 정의 완료")
print(f"프롬프트 길이: {len(HARDSKILL_EVIDENCE_SYSTEM_PROMPT):,}자")


# In[30]:


# Cell 29 | 소프트스킬 시스템 프롬프트 정의
# ------------------------------------------------------------
# CV 원문을 받아 S1~S5를 0/1로 판정하고 evidence를 인용하는 프롬프트입니다.
# ------------------------------------------------------------

SOFTSKILL_SYSTEM_PROMPT = """
당신은 CV에서 소프트스킬을 추출하는 전문 평가자입니다.
아래 가이드라인에 따라 S1~S5 각 항목을 0 또는 1로 판정하고,
판정 근거로 반드시 CV 원문을 직접 인용하세요.

=== 핵심 원칙 ===
증거 인정 조건: 행동 주체(본인) + 구체적 산출물 또는 결과 — 둘 다 있어야 1점.

공통 0점 처리 — 아래 패턴은 어느 항목에서도 1점을 주지 않는다:
- 형용사·부사로만 서술된 태도: "주도적으로", "능동적으로", "적극적으로", "성실하게"
- 동사 없는 역할 나열: "팀장", "PL", "리드", "총괄" — 직함만 있고 행동 없음
- 인과 없는 결과 선언: "매출 향상에 기여", "성과를 냈습니다" — 무엇을 해서인지 없음
- 미래 의향 서술: "~하고 싶습니다", "~하겠습니다"
- 교육·자격증 나열: 수료 사실만 있고 실제 프로젝트 적용 없음

=== S1 · 자기주도·오너십 ===
판정 질문: CV에 "요청 없이 본인이 문제를 발견하고 착수해서 결과를 만든" 서술이 있는가?
구조: 배경(문제 발견) → 본인이 착수 → 결과 도출
또는: 없던 기준·프로세스를 본인이 "정의·설계·수립"한 서술 (운영·관리·지원은 해당 없음)

인정 예시:
- "탈회율 15% 급증을 월초 데이터 정리 중 발견 → 지시 없이 원인분석 착수 → 가설기각 → 재분석 → 익월 탈회율 정상화"
  → 발견(월초 정리 중) + 착수(지시 없이) + 결과(정상화) 세 요소 완비
- "내부 데이터 단독 활용 포기 결정 → 공시지가+경매 데이터 직접 결합해 모형 설계 완성"
  → 방향 전환 결정이 본인 판단임이 명시 + 결과(모형 완성)
- "챌린지 신청자가 미신청자 대비 완강률 4배 높다는 데이터에 착안 → 게이미피케이션 설계"
  → 데이터에서 문제 발굴 + 서비스 기획으로 연결

불인정 예시:
- "프로젝트를 주도적으로 이끌었습니다" — 착수 배경·결과 없음
- "PL로서 팀원을 관리하였습니다" — 직함만, 본인이 문제 정의한 서술 없음
- PL·부장 직함이지만 모든 서술이 "고객사 요청 → 수행" 구조 → 0점

=== S2 · 협업 실행 ===
판정 질문: 이해관계가 다른 직군의 경계를 넘어 만든 공동 산출물이 있는가?
핵심: 기술↔비기술, 내부↔외부 파트너와 상이한 요구를 조율하여 프로세스 개선·성과를 만든 사례.

인정 예시:
- "IT/심사 파트와 협업하여 시스템 로직 개선 및 심사 가이드라인 재정립"
  → 데이터 분석가가 IT팀과 직군 경계를 넘어 시스템 기준까지 공동 재정립. 산출물(가이드라인) 명시.
- "다변량 회귀 수식을 항목별로 분해 → Java 개발자가 실제 시스템에 구현할 수 있도록 엑셀 시뮬레이터로 핸드오프"
  → 분석↔개발 경계를 넘어 상대방이 쓸 수 있는 형태로 변환. 공동 산출물.
- "영화사×카드사×통신사 3사 가명결합 데이터 분석 — 3개 이해관계자 조율 후 산출물 완성"
  → 서로 다른 이해관계를 가진 3개 조직 간 경계를 넘어 공동 데이터셋 완성
- "Figma·FigJam 활용 설계 문서 작성 → 내부 기획·개발·디자인 + 외부 의료기관·협력업체와 협업"
  → 내외부 동시 협업 산출물(설계 문서) 명시

불인정 예시:
- "유관부서와 협업하여 업무를 진행하였습니다" — 어느 직군인지, 산출물이 없음
- "팀원들과 긴밀히 소통하였습니다" — 같은 직군 내부
- "영업팀·마케팅팀에 체험고객 데이터 전달 및 콜영업 시나리오 공유" — 공유만 있고 공동 산출물 없음
- "구매팀·물류팀에 협조 요청하여 재고보유 수량 공유" — 협조 요청이지 공동 산출물 없음

=== S3 · 커뮤니케이션 전환력 ===
판정 질문: 기술↔비기술, 실무↔경영진 등 다른 언어 체계의 청중을 대상으로 한 커뮤니케이션 산출물이나 성과가 있는가?
핵심: 동일 내용을 다른 언어로 번역한 경험. 같은 직군 내부 소통은 해당 없음.

인정 예시:
- "제품 기능을 '정책 언어'로 번역해 B2G 영업전략 설계 → 제안서 영업조직 내부 채택"
  → 제품 언어 → 정부 조달 언어 번역. 산출물(제안서) + 채택.
- "사용자 DB 표준 가이드 작성·배포 → VOC 월 30% 감소"
  → 기술 기준 → 사용자 언어 번역. 산출물(가이드) + 수치 결과.
- "Figma/FigJam으로 시스템 구조 시각화 → 외부 의료기관 회의 기반 자료로 활용"
  → 기술 문서 → 비기술 의료기관 청중용 번역

불인정 예시:
- "커뮤니케이션 능력이 뛰어납니다" — 근거 없음
- "다양한 이해관계자와 소통하였습니다" — 청중이 누구인지, 산출물이 없음
- 동일 직군 내부에서 기술 문서를 공유한 것

=== S4 · 빠른 학습·적응력 ===
판정 질문: 이전에 없던 기술·도메인·직군으로 전환하면서도 각 단계에서 실제 산출물을 만든 이력이 있는가?
핵심: 교육 수료·자격증은 증거가 아님. 전환 후 프로젝트 적용과 산출물까지 있어야 인정.
비선형 커리어(직군·도메인을 바꿔가면서도 각 단계에서 산출물을 만들어온 이력)도 인정.

인정 예시:
- "통계학 → 금융AI(로보어드바이저) → 관세청 예측 → 부정탐지 → LLM 검색 — 5회 기술 확장, 매회 실제 배포"
  → 단순 학습이 아니라 매번 실제 시스템으로 완성. 5회 전환 모두 실적 연결.
- "웹 개발자 → 서비스 기획 → UX 전략 기획 — 전환 후 각 단계에서 수치 성과"
  → 직군 전환(개발→기획) + 각 단계 성과(이용률 5%↑, UX 문서 채택)

불인정 예시:
- "빠르게 습득하는 편입니다" — 근거 없음
- AI 과정 768시간 수료 + 자격증 3개 — 프로젝트 적용 없음
- 25년 경력이지만 Java/Oracle 동일 기술 스택 반복 → 전환 없음

=== S5 · 비즈니스 임팩트 지향 ===
판정 질문: 비즈니스 목적 + 결과가 함께 서술됐는가?
결과 인정 기준: 수치(%), 또는 출시·채택·구조변경의 명칭 — 이름이 있어야 함.
"기여했습니다"처럼 결과가 뭉뚱그려지면 불인정.

인정 예시:
- "KPI 70%대 후반 → 82%로 상향, 연간 목표 7%p 초과 달성"
  → 목적(KPI 개선) + 수치(7%p) 세트 명시
- "영업이익 6억 증대 — 소스 수명 예측으로 판매보증비 절감"
  → 분석→비즈니스 의사결정→재무 결과 연결. 수치(6억) + 경로 모두 명시.
- "UX 기준문서 정책 회의에서 채택 → 개발 스펙 반영"
  → 수치 없지만 '기준문서'라는 이름 + 채택이라는 결과 명시
- "공공기관 42대 납품, KCL 인증(맥박 96%, 혈압 97%)"
  → 납품이라는 명칭 + 수치 인증 결과

불인정 예시:
- "성과 중심으로 일합니다" — 선언
- "매출 향상에 기여하였습니다" — 수치도 없고, 상품 명칭도 없음
- "업무 효율성이 향상되었습니다" — 수치도 명칭도 없음

=== 인터뷰 플래그 ===
다음 조건에 해당하면 해당 항목에 인터뷰 플래그를 세운다:
- S1: "리드했다", "주도했다", "총괄했다" — 하지만 착수 배경과 결과물이 없을 때
  → 질문: "이 프로젝트에서 본인이 직접 문제를 정의하고 시작한 건지, 아니면 지시받은 과제를 이끈 건지요? 그 결과물이 구체적으로 무엇인가요?"
- S2: 협업 언급은 있는데 공동 산출물이 명시 안 됨
  → 질문: "OO팀과 협업했다고 하셨는데, 그 결과물이 구체적으로 무엇인가요? 본인 기여 범위는 어디까지인가요?"
- S4: 비선형 커리어인데 각 단계의 산출물이 CV에 빠져 있거나, 전환이 자발적인지 불명확
  → 질문: "OO에서 OO로 이동하셨는데, 각 단계에서 실제로 만드신 결과물이 있으면 말씀해 주세요. 그 전환이 본인이 선택한 건지, 상황에 의한 건지도 말씀해 주시겠어요?"
- S5: 프로젝트 수행 서술은 충분한데 결과(수치 또는 출시·채택 명칭)가 없을 때
  → 질문: "이 프로젝트가 실제 서비스나 의사결정에 어떻게 반영됐나요? 수치가 있다면 말씀해 주시고, 없다면 어떤 이름으로 출시됐거나 어떤 변화로 이어졌는지 말씀해 주세요."

=== 출력 형식 ===
반드시 아래 JSON 형식으로만 응답하라. 다른 텍스트를 추가하지 말 것.

{
  "scores": {
    "S1": {"score": 0 또는 1, "evidence": "CV 원문 직접 인용 또는 null", "reason": "판정 이유 한 줄"},
    "S2": {"score": 0 또는 1, "evidence": "CV 원문 직접 인용 또는 null", "reason": "판정 이유 한 줄"},
    "S3": {"score": 0 또는 1, "evidence": "CV 원문 직접 인용 또는 null", "reason": "판정 이유 한 줄"},
    "S4": {"score": 0 또는 1, "evidence": "CV 원문 직접 인용 또는 null", "reason": "판정 이유 한 줄"},
    "S5": {"score": 0 또는 1, "evidence": "CV 원문 직접 인용 또는 null", "reason": "판정 이유 한 줄"}
  },
  "raw_total": 0~5,
  "soft_score_30": 0~30,
  "interview_flags": [
    {"item": "S1", "condition": "플래그 발동 이유", "question": "헤드헌터가 물어볼 구체적 질문"}
  ],
  "s7_note": "평판조회에서 확인이 필요한 특이사항 (없으면 null)"
}
"""

print("소프트스킬 시스템 프롬프트 정의 완료")
print(f"프롬프트 길이: {len(SOFTSKILL_SYSTEM_PROMPT):,}자")


# In[31]:


# Cell 30 | JD-CV 매칭 해석용 시스템 프롬프트 정의
# ------------------------------------------------------------
# 최종 대시보드에는 JD 내용과 함께 LLM 해석이 들어갈 수 있도록 구성합니다.
# 이 프롬프트는 최종 result_df 한 행을 받아 추천 요약/강점/Gap/리스크/면접질문을 생성합니다.
# ------------------------------------------------------------

MATCH_EXPLANATION_SYSTEM_PROMPT = """
당신은 헤드헌팅 JD-CV 매칭 결과를 설명하는 전문 분석가입니다.
반드시 제공된 JD, CV, 점수, taxonomy 정보 안에서만 판단하세요.
추측하지 말고, CV 원문에 근거가 부족하면 확인 필요로 표시하세요.

출력은 반드시 JSON만 반환하세요.

출력 형식:
{
  "llm_match_summary": "한 문장 추천 요약",
  "llm_strength": "JD와 연결되는 주요 강점",
  "llm_gap": "부족하거나 확인 필요한 점",
  "llm_risk": "매칭 리스크",
  "llm_interview_question": ["질문1", "질문2", "질문3"],
  "llm_evidence": "판단에 사용한 CV/JD 근거 요약"
}
"""

print("JD-CV 매칭 해석 시스템 프롬프트 정의 완료")


# In[32]:


# Cell 31 | LLM 입력 JSONL 생성

# Cell 21 | LLM 입력 JSONL 생성 - JSON 직렬화 오류 방지 버전

import json
import math
from pathlib import Path

import numpy as np
import pandas as pd


def make_json_safe(obj):
    """
    json.dumps()에서 오류 나는 numpy/pandas 타입을
    Python 기본 타입으로 변환합니다.
    - np.float32, np.float64 -> float
    - np.int64, np.int32 -> int
    - np.ndarray -> list
    - NaN, inf -> None
    - pd.Timestamp -> str
    """

    if obj is None:
        return None

    if isinstance(obj, dict):
        return {str(k): make_json_safe(v) for k, v in obj.items()}

    if isinstance(obj, list):
        return [make_json_safe(v) for v in obj]

    if isinstance(obj, tuple):
        return [make_json_safe(v) for v in obj]

    if isinstance(obj, set):
        return [make_json_safe(v) for v in obj]

    if isinstance(obj, np.ndarray):
        return make_json_safe(obj.tolist())

    if isinstance(obj, np.integer):
        return int(obj)

    if isinstance(obj, np.floating):
        value = float(obj)
        if math.isnan(value) or math.isinf(value):
            return None
        return value

    if isinstance(obj, np.bool_):
        return bool(obj)

    if isinstance(obj, pd.Timestamp):
        return obj.isoformat()

    if isinstance(obj, float):
        if math.isnan(obj) or math.isinf(obj):
            return None
        return obj

    try:
        if pd.isna(obj):
            return None
    except Exception:
        pass

    return obj


# 출력 경로 확인
if "OUTPUT_LLM_INPUT_JSONL" not in globals():
    OUTPUT_LLM_INPUT_JSONL = OUTPUT_DIR / "llm_match_explanation_input.jsonl"

OUTPUT_LLM_INPUT_JSONL = Path(OUTPUT_LLM_INPUT_JSONL)
OUTPUT_LLM_INPUT_JSONL.parent.mkdir(parents=True, exist_ok=True)


# 필수 객체 확인
if "result_df" not in globals():
    raise NameError("result_df가 없습니다. 이전 셀에서 result_df를 먼저 생성하세요.")

if result_df.empty:
    raise ValueError("result_df가 비어 있습니다. 이전 매칭 결과를 확인하세요.")

### 누락 방지 추가
# build_match_explanation_user_payload 누락 방지

def build_match_explanation_user_payload(row):
    def get_value(key, default=""):
        try:
            value = row.get(key, default)
        except Exception:
            value = default

        if value is None:
            return ""

        try:
            import pandas as pd
            if pd.isna(value):
                return ""
        except Exception:
            pass

        return value

    payload = {
        "jd": {
            "jd_id": str(get_value("jd_id", "")),
            "jd_title": str(get_value("jd_title", get_value("title", ""))),
            "jd_position": str(get_value("jd_position", "")),
            "jd_required_skills": get_value("jd_required_skills", get_value("required_skills", "")),
            "jd_tasks": get_value("jd_tasks", get_value("tasks", "")),
            "jd_domain": get_value("jd_domain", get_value("domain", "")),
            "jd_text": str(get_value("jd_text", get_value("jd_full_text", "")))[:4000],
        },
        "candidate": {
            "candidate_id": str(get_value("candidate_id", "")),
            "cv_name": str(get_value("cv_name", get_value("name", ""))),
            "cv_position": str(get_value("cv_position", "")),
            "cv_skills": get_value("cv_skills", get_value("skills", "")),
            "cv_experience": get_value("cv_experience", get_value("experience", "")),
            "cv_text": str(get_value("cv_text", get_value("cv_full_text", "")))[:4000],
        },
        "scores": {
            "rank": get_value("rank", ""),
            "final_score_100": get_value("final_score_100", ""),
            "final_score": get_value("final_score", ""),
            "retrieval_score_norm": get_value("retrieval_score_norm", ""),
            "keyword_baseline_score_norm": get_value("keyword_baseline_score_norm", ""),
            "column_score": get_value("column_score", ""),
            "skill_score": get_value("skill_score", ""),
            "task_score": get_value("task_score", ""),
            "domain_score": get_value("domain_score", ""),
            "role_score": get_value("role_score", ""),
            "ai_skill_overlap_score": get_value("ai_skill_overlap_score", ""),
            "cross_encoder_score": get_value("cross_encoder_score", ""),
            "embedding_score": get_value("embedding_score", ""),
            "retrieval_bm25_score": get_value("retrieval_bm25_score", ""),
            "retrieval_tfidf_score": get_value("retrieval_tfidf_score", ""),
        },
        "rule_based_reason": str(get_value("match_reason_rule_based", ""))[:2000],
    }

    return payload


if "build_match_explanation_user_payload" not in globals():
    raise NameError("build_match_explanation_user_payload 함수가 없습니다. 이전 셀 정의를 확인하세요.")


# JSONL 저장
saved_count = 0

with open(OUTPUT_LLM_INPUT_JSONL, "w", encoding="utf-8") as f:
    for _, row in result_df.iterrows():
        payload = build_match_explanation_user_payload(row)
        payload = make_json_safe(payload)

        f.write(json.dumps(payload, ensure_ascii=False) + "\n")
        saved_count += 1


print("LLM 입력 JSONL 저장 완료:", OUTPUT_LLM_INPUT_JSONL)
print("저장 건수:", saved_count)
print()
print("예시 payload:")

example_payload = build_match_explanation_user_payload(result_df.iloc[0])
example_payload = make_json_safe(example_payload)

print(json.dumps(example_payload, ensure_ascii=False, indent=2)[:3000])


# In[33]:


# Cell 32 | 누락 함수 보정 - to_cell_text

# to_cell_text 함수 누락 방지
def to_cell_text(value):
    import json
    import math
    import numpy as np
    import pandas as pd

    if value is None:
        return ""

    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass

    if isinstance(value, np.ndarray):
        value = value.tolist()

    if isinstance(value, (np.integer,)):
        return str(int(value))

    if isinstance(value, (np.floating,)):
        value = float(value)
        if math.isnan(value) or math.isinf(value):
            return ""
        return str(value)

    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return ""
        return str(value)

    if isinstance(value, (list, dict, tuple, set)):
        return json.dumps(value, ensure_ascii=False)

    return str(value)


# In[34]:


# Cell 33 | LLM 매칭 해석 생성

# ------------------
# 1. 기본 import 및 Gemini 설정값 준비
# ------------------

import os
import json
import re
import time
import math
from pathlib import Path

import numpy as np
import pandas as pd


if "GEMINI_API_KEY" not in globals():
    GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "")

if "MATCH_EXPLANATION_SYSTEM_PROMPT" not in globals():
    MATCH_EXPLANATION_SYSTEM_PROMPT = """
당신은 JD와 후보자 CV 매칭 결과를 설명하는 평가자입니다.
반드시 JSON 형식으로만 답하세요.

출력 JSON 키:
{
  "llm_match_summary": "JD와 후보자의 전반적 매칭 요약",
  "llm_strength": "후보자의 주요 강점",
  "llm_gap": "JD 대비 부족하거나 확인이 필요한 부분",
  "llm_risk": "채용 또는 배치 시 리스크",
  "llm_recommendation": "추천 여부와 간단한 판단"
}
"""


# ------------------
# 2. JSON 변환 안전 처리
# ------------------

def make_json_safe(obj):
    if obj is None:
        return None

    if isinstance(obj, dict):
        return {str(k): make_json_safe(v) for k, v in obj.items()}

    if isinstance(obj, list):
        return [make_json_safe(v) for v in obj]

    if isinstance(obj, tuple):
        return [make_json_safe(v) for v in obj]

    if isinstance(obj, set):
        return [make_json_safe(v) for v in obj]

    if isinstance(obj, np.ndarray):
        return make_json_safe(obj.tolist())

    if isinstance(obj, np.integer):
        return int(obj)

    if isinstance(obj, np.floating):
        value = float(obj)
        if math.isnan(value) or math.isinf(value):
            return None
        return value

    if isinstance(obj, np.bool_):
        return bool(obj)

    if isinstance(obj, pd.Timestamp):
        return obj.isoformat()

    if isinstance(obj, float):
        if math.isnan(obj) or math.isinf(obj):
            return None
        return obj

    try:
        if pd.isna(obj):
            return None
    except Exception:
        pass

    return obj


# ------------------
# 3. DataFrame 셀 저장용 텍스트 변환
# ------------------

def to_cell_text(value):
    if value is None:
        return ""

    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass

    if isinstance(value, np.ndarray):
        value = value.tolist()

    if isinstance(value, np.integer):
        return str(int(value))

    if isinstance(value, np.floating):
        value = float(value)
        if math.isnan(value) or math.isinf(value):
            return ""
        return str(value)

    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return ""
        return str(value)

    if isinstance(value, (list, dict, tuple, set)):
        return json.dumps(make_json_safe(value), ensure_ascii=False)

    return str(value)


# ------------------
# 4. Gemini 응답 JSON 추출
# ------------------

def extract_json_from_text(text):
    if not isinstance(text, str):
        return None

    s = text.strip()
    s = re.sub(r"^```json\s*", "", s)
    s = re.sub(r"^```\s*", "", s)
    s = re.sub(r"\s*```$", "", s)

    try:
        return json.loads(s)
    except Exception:
        pass

    m = re.search(r"\{.*\}", s, flags=re.DOTALL)
    if m:
        try:
            return json.loads(m.group(0))
        except Exception:
            pass

    return None


# ------------------
# 5. Gemini 사용 가능 모델 목록 생성
# ------------------

def get_gemini_candidate_models(api_key):
    preferred = [
        "gemini-2.5-flash",
        "gemini-2.5-flash-lite",
        "gemini-2.5-pro",
        "gemini-2.0-flash-lite",
        "gemini-1.5-flash-latest",
        "gemini-1.5-pro-latest",
    ]

    blocked = {
        "gemini-2.0-flash",
        "gemini-1.5-flash",
        "gemini-1.5-pro",
    }

    available = []

    try:
        import google.generativeai as genai

        genai.configure(api_key=api_key)

        for model in genai.list_models():
            model_name = getattr(model, "name", "")
            methods = getattr(model, "supported_generation_methods", [])

            model_name = model_name.replace("models/", "")

            if "generateContent" not in methods:
                continue

            if model_name in blocked:
                continue

            available.append(model_name)

    except Exception:
        available = []

    ordered = []

    for model_name in preferred:
        if model_name in available and model_name not in ordered:
            ordered.append(model_name)

    for model_name in available:
        if model_name not in ordered:
            ordered.append(model_name)

    if not ordered:
        ordered = [
            "gemini-2.5-flash",
            "gemini-2.5-flash-lite",
            "gemini-2.5-pro",
            "gemini-2.0-flash-lite",
        ]

    return ordered


# ------------------
# 6. Gemini JSON 호출 함수
# ------------------

def call_gemini_json(system_prompt, payload, model_name=None):
    api_key = globals().get("GEMINI_API_KEY", "") or os.getenv("GEMINI_API_KEY", "")

    if not api_key:
        return {"error": "GEMINI_API_KEY가 없습니다."}

    if model_name is None:
        candidate_models = get_gemini_candidate_models(api_key)
    else:
        candidate_models = [model_name] + [
            m for m in get_gemini_candidate_models(api_key)
            if m != model_name
        ]

    last_error = None

    for candidate_model in candidate_models:
        try:
            try:
                import google.generativeai as genai

                genai.configure(api_key=api_key)

                model = genai.GenerativeModel(
                    model_name=candidate_model,
                    system_instruction=system_prompt,
                )

                response = model.generate_content(
                    json.dumps(make_json_safe(payload), ensure_ascii=False)
                )

                text = getattr(response, "text", "")
                parsed = extract_json_from_text(text)

                if isinstance(parsed, dict):
                    parsed["_gemini_model_name"] = candidate_model
                    return parsed

                return {
                    "parse_error": "Gemini 응답을 JSON으로 파싱하지 못했습니다.",
                    "raw_text": text,
                    "_gemini_model_name": candidate_model,
                }

            except ImportError:
                from google import genai

                client = genai.Client(api_key=api_key)

                response = client.models.generate_content(
                    model=candidate_model,
                    contents=[
                        system_prompt,
                        json.dumps(make_json_safe(payload), ensure_ascii=False),
                    ],
                )

                text = getattr(response, "text", "")
                parsed = extract_json_from_text(text)

                if isinstance(parsed, dict):
                    parsed["_gemini_model_name"] = candidate_model
                    return parsed

                return {
                    "parse_error": "Gemini 응답을 JSON으로 파싱하지 못했습니다.",
                    "raw_text": text,
                    "_gemini_model_name": candidate_model,
                }

        except Exception as e:
            last_error = repr(e)

            if (
                "NotFound" in last_error
                or "not found" in last_error.lower()
                or "no longer available" in last_error.lower()
                or "not supported" in last_error.lower()
            ):
                continue

            return {
                "error": last_error,
                "_gemini_model_name": candidate_model,
            }

    return {
        "error": last_error or "사용 가능한 Gemini 모델을 찾지 못했습니다.",
        "_gemini_model_name": "",
    }


# ------------------
# 7. LLM 입력 payload 생성
# ------------------

def build_match_explanation_user_payload(row):
    def get_value(key, default=""):
        try:
            value = row.get(key, default)
        except Exception:
            value = default

        if value is None:
            return ""

        try:
            if pd.isna(value):
                return ""
        except Exception:
            pass

        return value

    return {
        "jd": {
            "jd_id": str(get_value("jd_id", "")),
            "jd_title": str(get_value("jd_title", get_value("title", ""))),
            "jd_position": str(get_value("jd_position", "")),
            "jd_required_skills": get_value("jd_required_skills", get_value("required_skills", "")),
            "jd_tasks": get_value("jd_tasks", get_value("tasks", "")),
            "jd_domain": get_value("jd_domain", get_value("domain", "")),
            "jd_text": str(get_value("jd_text", get_value("jd_full_text", "")))[:4000],
        },
        "candidate": {
            "candidate_id": str(get_value("candidate_id", "")),
            "cv_name": str(get_value("cv_name", get_value("name", ""))),
            "cv_position": str(get_value("cv_position", "")),
            "cv_skills": get_value("cv_skills", get_value("skills", "")),
            "cv_experience": get_value("cv_experience", get_value("experience", "")),
            "cv_text": str(get_value("cv_text", get_value("cv_full_text", "")))[:4000],
        },
        "scores": {
            "rank": get_value("rank", ""),
            "final_score_100": get_value("final_score_100", ""),
            "final_score": get_value("final_score", ""),
            "retrieval_score_norm": get_value("retrieval_score_norm", ""),
            "keyword_baseline_score_norm": get_value("keyword_baseline_score_norm", ""),
            "column_score": get_value("column_score", ""),
            "skill_score": get_value("skill_score", ""),
            "task_score": get_value("task_score", ""),
            "domain_score": get_value("domain_score", ""),
            "role_score": get_value("role_score", ""),
            "ai_skill_overlap_score": get_value("ai_skill_overlap_score", ""),
            "cross_encoder_score": get_value("cross_encoder_score", ""),
            "embedding_score": get_value("embedding_score", ""),
            "retrieval_bm25_score": get_value("retrieval_bm25_score", ""),
            "retrieval_tfidf_score": get_value("retrieval_tfidf_score", ""),
        },
        "rule_based_reason": str(get_value("match_reason_rule_based", ""))[:2000],
    }


# ------------------
# 8. Gemini 응답 표준 컬럼 매핑
# ------------------

def normalize_llm_output(out):
    if isinstance(out, dict) and "error" in out:
        return {
            "llm_match_summary": "",
            "llm_strength": "",
            "llm_gap": "",
            "llm_risk": "",
            "llm_recommendation": "",
            "llm_error": str(out.get("error", "")),
        }

    if isinstance(out, str):
        parsed = extract_json_from_text(out)
        if isinstance(parsed, dict):
            out = parsed
        else:
            return {
                "llm_match_summary": out[:1500],
                "llm_strength": "",
                "llm_gap": "",
                "llm_risk": "",
                "llm_recommendation": "",
                "llm_error": "",
            }

    if not isinstance(out, dict):
        return {
            "llm_match_summary": str(out)[:1500],
            "llm_strength": "",
            "llm_gap": "",
            "llm_risk": "",
            "llm_recommendation": "",
            "llm_error": f"응답 타입 오류: {type(out)}",
        }

    for key in ["text", "response", "content", "output", "answer", "message", "result"]:
        if key in out and isinstance(out[key], str):
            parsed = extract_json_from_text(out[key])
            if isinstance(parsed, dict):
                out = parsed
                break

    for key in ["data", "parsed", "json", "arguments", "function_call", "tool_call", "payload"]:
        if key in out and isinstance(out[key], dict):
            out = out[key]
            break

    def pick(keys):
        for k in keys:
            if k in out and out[k] not in [None, ""]:
                return out[k]
        return ""

    summary = pick([
        "llm_match_summary", "match_summary", "matching_summary",
        "overall_summary", "summary", "fit_summary",
        "candidate_summary", "종합의견", "요약", "매칭요약", "매칭_요약",
    ])

    strength = pick([
        "llm_strength", "strength", "strengths", "fit_strength",
        "matched_strength", "positive_points", "강점", "적합점",
    ])

    gap = pick([
        "llm_gap", "gap", "gaps", "weakness", "weaknesses",
        "mismatch", "missing_points", "부족점", "갭", "차이점",
    ])

    risk = pick([
        "llm_risk", "risk", "risks", "concern", "concerns",
        "주의사항", "리스크",
    ])

    recommendation = pick([
        "llm_recommendation", "recommendation", "decision",
        "hiring_recommendation", "fit_decision", "추천의견", "판단", "추천",
    ])

    if summary == "" and "raw_text" in out:
        summary = str(out.get("raw_text", ""))[:1500]

    if summary == "":
        summary = json.dumps(make_json_safe(out), ensure_ascii=False)[:1500]

    return {
        "llm_match_summary": summary,
        "llm_strength": strength,
        "llm_gap": gap,
        "llm_risk": risk,
        "llm_recommendation": recommendation,
        "llm_error": "",
    }


# ------------------
# 9. result_df 확인 및 LLM 결과 컬럼 초기화
# ------------------

if "result_df" not in globals():
    raise NameError("result_df가 없습니다. 이전 매칭 결과 셀을 먼저 실행하세요.")

if result_df.empty:
    raise ValueError("result_df가 비어 있습니다. 이전 매칭 결과 셀을 확인하세요.")

llm_cols = [
    "llm_match_summary",
    "llm_strength",
    "llm_gap",
    "llm_risk",
    "llm_recommendation",
    "llm_error",
]

for col in llm_cols:
    if col not in result_df.columns:
        result_df[col] = ""
    else:
        result_df[col] = ""


# ------------------
# 10. LLM 해석 대상 선정
# ------------------

LLM_EXPLAIN_TOP_N = globals().get("LLM_EXPLAIN_TOP_N", 10)

target_df = (
    result_df
    .sort_values(["jd_id", "rank"], ascending=[True, True])
    .groupby("jd_id", group_keys=False)
    .head(LLM_EXPLAIN_TOP_N)
)

explain_indices = target_df.index.tolist()

candidate_models = get_gemini_candidate_models(GEMINI_API_KEY) if GEMINI_API_KEY else []

print(f"LLM 해석 대상 건수: {len(explain_indices)}")
print(f"JD별 Top {LLM_EXPLAIN_TOP_N} 기준")
print(f"Gemini 후보 모델: {candidate_models[:5] if candidate_models else 'API_KEY 없음'}")
print()


# ------------------
# 11. Gemini 호출 및 result_df 반영
# ------------------

success_count = 0
fail_count = 0
first_raw_output_printed = False
selected_model_name = ""

for n, idx in enumerate(explain_indices, start=1):
    row = result_df.loc[idx]

    jd_id = row.get("jd_id", "")
    candidate_id = row.get("candidate_id", "")
    rank = row.get("rank", "")

    print(f"[{n}/{len(explain_indices)}] jd_id={jd_id}, candidate_id={candidate_id}, rank={rank}")

    try:
        payload = build_match_explanation_user_payload(row)
        payload = make_json_safe(payload)

        out = call_gemini_json(
            MATCH_EXPLANATION_SYSTEM_PROMPT,
            payload,
            model_name=selected_model_name if selected_model_name else None,
        )

        if isinstance(out, dict) and out.get("_gemini_model_name", ""):
            selected_model_name = out.get("_gemini_model_name", "")

        if not first_raw_output_printed:
            print()
            print("첫 번째 Gemini 원본 응답 미리보기:")
            print(str(out)[:1500])
            print()
            first_raw_output_printed = True

        parsed = normalize_llm_output(out)

        result_df.loc[idx, "llm_match_summary"] = to_cell_text(parsed.get("llm_match_summary", ""))
        result_df.loc[idx, "llm_strength"] = to_cell_text(parsed.get("llm_strength", ""))
        result_df.loc[idx, "llm_gap"] = to_cell_text(parsed.get("llm_gap", ""))
        result_df.loc[idx, "llm_risk"] = to_cell_text(parsed.get("llm_risk", ""))
        result_df.loc[idx, "llm_recommendation"] = to_cell_text(parsed.get("llm_recommendation", ""))
        result_df.loc[idx, "llm_error"] = to_cell_text(parsed.get("llm_error", ""))

        if result_df.loc[idx, "llm_error"] != "":
            fail_count += 1
            print("  실패:", result_df.loc[idx, "llm_error"][:300])
        elif result_df.loc[idx, "llm_match_summary"] != "":
            success_count += 1
            print("  성공")
        else:
            fail_count += 1
            result_df.loc[idx, "llm_error"] = "summary 생성 실패"
            print("  실패: summary 생성 실패")

        time.sleep(0.2)

    except Exception as e:
        fail_count += 1
        result_df.loc[idx, "llm_error"] = repr(e)
        print("  실패:", repr(e))


# ------------------
# 12. 실행 결과 확인
# ------------------

print()
print("LLM 해석 완료")
print("성공:", success_count)
print("실패:", fail_count)
print("사용 모델:", selected_model_name)

display_cols = [
    "jd_id",
    "candidate_id",
    "rank",
    "final_score_100",
    "llm_match_summary",
    "llm_gap",
    "llm_error",
]

display(
    result_df
    .sort_values(["jd_id", "rank"])
    [display_cols]
    .head(30)
)


# In[35]:


# Cell 34 | 상단에 추가: 누락 함수 보정

def compact_text(value, max_len=120):
    import re
    import json
    import math
    import numpy as np
    import pandas as pd

    if value is None:
        return ""

    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass

    if isinstance(value, np.ndarray):
        value = value.tolist()

    if isinstance(value, (list, dict, tuple, set)):
        value = json.dumps(value, ensure_ascii=False)

    if isinstance(value, np.integer):
        value = int(value)

    if isinstance(value, np.floating):
        value = float(value)
        if math.isnan(value) or math.isinf(value):
            return ""

    text = str(value)
    text = re.sub(r"\s+", " ", text).strip()

    if max_len is not None and len(text) > max_len:
        return text[:max_len].rstrip() + "..."

    return text


# In[36]:


# Cell 35 | 누락 함수 보정 - clean_value

import math
import json
import numpy as np
import pandas as pd


def clean_value(value):
    """
    Excel/CSV 저장 전에 object 값을 안전하게 문자열 또는 기본 Python 타입으로 변환합니다.
    - None / NaN / inf: 빈 문자열
    - ndarray: list로 변환 후 JSON 문자열화
    - list / dict / tuple / set: JSON 문자열화
    - numpy integer / floating: Python 기본 타입으로 변환
    """

    if value is None:
        return ""

    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass

    if isinstance(value, np.ndarray):
        value = value.tolist()

    if isinstance(value, (list, dict, tuple, set)):
        return json.dumps(value, ensure_ascii=False)

    if isinstance(value, np.integer):
        return int(value)

    if isinstance(value, np.floating):
        value = float(value)
        if math.isnan(value) or math.isinf(value):
            return ""
        return value

    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return ""
        return value

    return value


print("✅ clean_value 함수 보정 완료")


# In[37]:


# Cell 36 | Top3 후보 커리어 패스 타임라인 요약 optional
# ------------------------------------------------------------
# 목적:
# - JD별 Top3 후보자의 Career / Career_Description / Raw_Text를 근거로
#   커리어 패스를 타임라인 형태로 요약합니다.
# - 대시보드에서는 candidate 상세 영역에서 이 컬럼을 사용합니다.
# - LLM 실패 시에도 rule-based fallback을 생성해 스키마를 유지합니다.
# ------------------------------------------------------------

import os
import json
import re
import time
import math

import numpy as np
import pandas as pd


# ------------------------------------------------------------
# 1. 실행 옵션
# ------------------------------------------------------------
# 기본값은 LLM을 사용하지 않습니다.
# 대시보드용 커리어 패스 컬럼을 빠르게 만들기 위해 rule-based 방식으로 먼저 생성합니다.
# LLM을 사용하려면 아래 값을 True로 바꿔 1~3건만 먼저 테스트하세요.

USE_CAREER_PATH_LLM = globals().get("USE_CAREER_PATH_LLM", True)
CAREER_PATH_MODEL_NAME = globals().get("CAREER_PATH_MODEL_NAME", "gemini-2.5-flash")
CAREER_PATH_TOP_N_PER_JD = globals().get("CAREER_PATH_TOP_N_PER_JD", 3)
CAREER_PATH_LLM_TIMEOUT = globals().get("CAREER_PATH_LLM_TIMEOUT", 20)


# ------------------------------------------------------------
# 2. 기본 보조 함수
# ------------------------------------------------------------

def clean_value(value, default=""):
    if value is None:
        return default

    try:
        if pd.isna(value):
            return default
    except Exception:
        pass

    text = str(value).strip()

    if text.lower() in ["nan", "none", "null"]:
        return default

    return text


def clean_text(value):
    text = clean_value(value, "")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def compact_text(value, max_len=160):
    text = clean_text(value)

    if len(text) <= max_len:
        return text

    return text[:max_len].rstrip() + "..."


def make_json_safe(obj):
    if obj is None:
        return None

    if isinstance(obj, dict):
        return {str(k): make_json_safe(v) for k, v in obj.items()}

    if isinstance(obj, list):
        return [make_json_safe(v) for v in obj]

    if isinstance(obj, tuple):
        return [make_json_safe(v) for v in obj]

    if isinstance(obj, set):
        return [make_json_safe(v) for v in obj]

    if isinstance(obj, np.ndarray):
        return make_json_safe(obj.tolist())

    if isinstance(obj, np.integer):
        return int(obj)

    if isinstance(obj, np.floating):
        value = float(obj)
        if math.isnan(value) or math.isinf(value):
            return None
        return value

    if isinstance(obj, np.bool_):
        return bool(obj)

    if isinstance(obj, pd.Timestamp):
        return obj.isoformat()

    if isinstance(obj, float):
        if math.isnan(obj) or math.isinf(obj):
            return None
        return obj

    try:
        if pd.isna(obj):
            return None
    except Exception:
        pass

    return obj


def extract_json_from_text(text):
    if not isinstance(text, str):
        return None

    s = text.strip()
    s = re.sub(r"^```json\s*", "", s)
    s = re.sub(r"^```\s*", "", s)
    s = re.sub(r"\s*```$", "", s)

    try:
        return json.loads(s)
    except Exception:
        pass

    m = re.search(r"\{.*\}", s, flags=re.DOTALL)
    if m:
        try:
            return json.loads(m.group(0))
        except Exception:
            pass

    return None


# ------------------------------------------------------------
# 3. Career Timeline Prompt
# ------------------------------------------------------------

CAREER_TIMELINE_SYSTEM_PROMPT = """
당신은 채용 매칭 대시보드의 후보자 커리어 분석가입니다.
입력된 후보자의 경력 텍스트를 근거로, JD 적합도 관점에서 읽기 쉬운 커리어 패스 타임라인을 한국어 JSON으로 작성하세요.

반드시 아래 JSON schema만 반환하세요.
{
  "career_path_summary": "후보자의 커리어 흐름을 2~3문장으로 요약",
  "career_fit_comment": "이번 JD와 연결되는 커리어상의 강점/주의점을 1~2문장으로 요약",
  "timeline": [
    {
      "period": "연도/기간을 알 수 없으면 '기간 미상'",
      "title": "회사/역할/프로젝트를 압축한 제목",
      "role": "담당 역할",
      "evidence": "원문 경력에서 확인되는 근거 요약",
      "jd_relevance": "JD와 연결되는 지점"
    }
  ]
}

규칙:
- 원문에 없는 회사명, 기간, 성과 수치를 지어내지 마세요.
- 기간 정보가 없으면 '기간 미상'이라고 쓰세요.
- timeline은 최대 5개로 제한하세요.
- 경력의 흐름이 잘 드러나도록 오래된 경험에서 최근 경험 순서로 정리하세요.
"""


# ------------------------------------------------------------
# 4. LLM 입력 payload 생성
# ------------------------------------------------------------

def build_career_path_payload(row):
    return {
        "jd": {
            "company": row.get("jd_company", ""),
            "title": row.get("jd_title", ""),
            "required_skill": row.get("jd_required_skill", ""),
            "main_task": row.get("jd_main_task", ""),
            "domain": row.get("jd_domain", ""),
            "role_signal": row.get("jd_role_signal", ""),
        },
        "candidate": {
            "candidate_id": row.get("candidate_id", ""),
            "job_role": row.get("cv_job_role", ""),
            "position": row.get("cv_position", ""),
            "total_career_years": row.get("cv_total_career_years", ""),
            "career": row.get("cv_career", ""),
            "career_description": row.get("cv_career_description", ""),
            "skill": row.get("cv_skill", ""),
            "raw_text_excerpt": str(row.get("cv_raw_text", ""))[:5000],
        },
        "match_context": {
            "rank": row.get("rank", ""),
            "final_score_100": row.get("final_score_100", ""),
            "skill_score_100": row.get("skill_score_100", ""),
            "task_score_100": row.get("task_score_100", ""),
            "role_score_100": row.get("role_score_100", ""),
            "rule_based_reason": row.get("match_reason_rule_based", ""),
        }
    }


# ------------------------------------------------------------
# 5. Career Path 전용 Gemini 호출
# ------------------------------------------------------------
# 이 함수는 genai.list_models()를 호출하지 않습니다.
# 모델 목록 자동 조회에서 멈추는 문제를 피하기 위해 고정 모델명만 사용합니다.

def call_career_path_json(system_prompt, payload, model_name=None):
    if not globals().get("USE_CAREER_PATH_LLM", False):
        return {"error": "CAREER_PATH_LLM_DISABLED"}

    api_key = globals().get("GEMINI_API_KEY", "") or os.getenv("GEMINI_API_KEY", "")

    if not api_key:
        return {"error": "GEMINI_API_KEY가 없습니다."}

    if model_name is None:
        model_name = globals().get("CAREER_PATH_MODEL_NAME", "gemini-2.5-flash")

    try:
        import google.generativeai as genai

        genai.configure(api_key=api_key)

        model = genai.GenerativeModel(
            model_name=model_name,
            system_instruction=system_prompt,
        )

        response = model.generate_content(
            json.dumps(make_json_safe(payload), ensure_ascii=False),
            request_options={"timeout": globals().get("CAREER_PATH_LLM_TIMEOUT", 20)},
        )

        text = getattr(response, "text", "")
        parsed = extract_json_from_text(text)

        if isinstance(parsed, dict):
            parsed["_gemini_model_name"] = model_name
            return parsed

        return {
            "parse_error": "Gemini 응답을 JSON으로 파싱하지 못했습니다.",
            "raw_text": text,
            "_gemini_model_name": model_name,
        }

    except Exception as e:
        return {
            "error": repr(e),
            "_gemini_model_name": model_name,
        }


# ------------------------------------------------------------
# 6. Rule-based Career Timeline 생성
# ------------------------------------------------------------

def build_rule_based_career_timeline(row):
    career = clean_text(row.get("cv_career", ""))
    desc = clean_text(row.get("cv_career_description", ""))
    raw_text = clean_text(row.get("cv_raw_text", ""))

    source = career if career else desc
    if not source:
        source = raw_text

    chunks = [
        c.strip(" -•\t")
        for c in re.split(r"\n+|(?<=다\.)\s+|(?<=요\.)\s+|(?<=\.)\s+", source)
        if c.strip()
    ]

    timeline = []

    for chunk in chunks[:5]:
        period_match = re.search(
            r"(20\d{2}\s*[.~\-]\s*(?:20\d{2}|현재|Present|present)?|\d+\s*년\s*\d*\s*개월?)",
            chunk
        )

        period = period_match.group(1) if period_match else "기간 미상"

        timeline.append({
            "period": period,
            "title": compact_text(chunk, 70),
            "role": clean_value(row.get("cv_position", "")),
            "evidence": compact_text(chunk, 160),
            "jd_relevance": compact_text(row.get("match_reason_rule_based", ""), 120),
        })

    if not timeline:
        timeline = [{
            "period": "기간 미상",
            "title": clean_value(row.get("cv_job_role", "후보 경력")),
            "role": clean_value(row.get("cv_position", "")),
            "evidence": compact_text(desc or career or raw_text or row.get("cv_skill", ""), 180),
            "jd_relevance": compact_text(row.get("match_reason_rule_based", ""), 120),
        }]

    career_years = clean_value(row.get("cv_total_career_years", ""))
    job_role = clean_value(row.get("cv_job_role", ""))
    candidate_id = clean_value(row.get("candidate_id", ""))

    if career_years or job_role:
        summary = f"{candidate_id}는 {career_years} 경력의 {job_role} 후보입니다. 경력 원문 기준으로 JD와 연결되는 수행 경험을 타임라인으로 정리했습니다."
    else:
        summary = f"{candidate_id} 후보의 경력 원문을 기준으로 JD와 연결되는 수행 경험을 타임라인으로 정리했습니다."

    return {
        "career_path_summary": summary,
        "career_fit_comment": compact_text(row.get("match_reason_rule_based", ""), 220),
        "timeline": timeline,
    }


# ------------------------------------------------------------
# 7. LLM 출력 표준화
# ------------------------------------------------------------

def normalize_career_timeline_output(out, row):
    if not isinstance(out, dict):
        return build_rule_based_career_timeline(row)

    if out.get("error") or out.get("parse_error"):
        return build_rule_based_career_timeline(row)

    summary = clean_value(out.get("career_path_summary", ""))
    comment = clean_value(out.get("career_fit_comment", ""))
    timeline = out.get("timeline", [])

    if not isinstance(timeline, list):
        timeline = []

    cleaned_timeline = []

    for item in timeline[:5]:
        if not isinstance(item, dict):
            continue

        cleaned_timeline.append({
            "period": clean_value(item.get("period", "기간 미상"), "기간 미상"),
            "title": compact_text(item.get("title", ""), 100),
            "role": compact_text(item.get("role", ""), 100),
            "evidence": compact_text(item.get("evidence", ""), 220),
            "jd_relevance": compact_text(item.get("jd_relevance", ""), 180),
        })

    if not summary or not cleaned_timeline:
        return build_rule_based_career_timeline(row)

    return {
        "career_path_summary": summary,
        "career_fit_comment": comment,
        "timeline": cleaned_timeline,
    }


# ------------------------------------------------------------
# 8. result_df 확인 및 결과 컬럼 초기화
# ------------------------------------------------------------

if "result_df" not in globals():
    raise NameError("result_df가 없습니다. 이전 매칭 결과 셀을 먼저 실행하세요.")

if result_df.empty:
    raise ValueError("result_df가 비어 있습니다. 이전 매칭 결과 셀을 확인하세요.")

career_timeline_cols = [
    "career_path_summary",
    "career_fit_comment",
    "career_timeline_json",
]

for col in career_timeline_cols:
    if col not in result_df.columns:
        result_df[col] = ""
    else:
        result_df[col] = ""


# ------------------------------------------------------------
# 9. 커리어 패스 타임라인 대상 선정
# ------------------------------------------------------------

career_indices = (
    result_df[result_df["rank"] <= CAREER_PATH_TOP_N_PER_JD]
    .sort_values(["jd_id", "rank"], ascending=[True, True])
    .index
    .tolist()
)

print("커리어 패스 타임라인 대상:", len(career_indices))
print("JD별 Top N:", CAREER_PATH_TOP_N_PER_JD)
print("LLM 사용:", USE_CAREER_PATH_LLM)
print("LLM 모델:", CAREER_PATH_MODEL_NAME if USE_CAREER_PATH_LLM else "사용 안 함")
print()


# ------------------------------------------------------------
# 10. 커리어 패스 타임라인 생성
# ------------------------------------------------------------

success_count = 0
fallback_count = 0
fail_count = 0

for n, idx in enumerate(career_indices, start=1):
    row = result_df.loc[idx]

    jd_id = row.get("jd_id", "")
    candidate_id = row.get("candidate_id", "")
    rank = row.get("rank", "")

    print(f"[{n}/{len(career_indices)}] jd_id={jd_id}, candidate_id={candidate_id}, rank={rank}")

    try:
        out = {}

        if USE_CAREER_PATH_LLM and GEMINI_API_KEY:
            payload = build_career_path_payload(row)
            payload = make_json_safe(payload)

            out = call_career_path_json(
                CAREER_TIMELINE_SYSTEM_PROMPT,
                payload,
                model_name=CAREER_PATH_MODEL_NAME,
            )

        normalized = normalize_career_timeline_output(out, row)

        if not out or out.get("error") or out.get("parse_error"):
            fallback_count += 1
            status = "fallback"
        else:
            success_count += 1
            status = "success"

        result_df.loc[idx, "career_path_summary"] = clean_value(normalized.get("career_path_summary", ""))
        result_df.loc[idx, "career_fit_comment"] = clean_value(normalized.get("career_fit_comment", ""))
        result_df.loc[idx, "career_timeline_json"] = json.dumps(
            make_json_safe(normalized.get("timeline", [])),
            ensure_ascii=False,
        )

        print("  완료:", status)

        if USE_CAREER_PATH_LLM and GEMINI_API_KEY:
            time.sleep(0.5)

    except Exception as e:
        fail_count += 1

        fallback = build_rule_based_career_timeline(row)

        result_df.loc[idx, "career_path_summary"] = clean_value(fallback.get("career_path_summary", ""))
        result_df.loc[idx, "career_fit_comment"] = clean_value(fallback.get("career_fit_comment", ""))
        result_df.loc[idx, "career_timeline_json"] = json.dumps(
            make_json_safe(fallback.get("timeline", [])),
            ensure_ascii=False,
        )

        print("  실패 후 fallback:", repr(e)[:300])


# ------------------------------------------------------------
# 11. 실행 결과 확인
# ------------------------------------------------------------

print()
print("커리어 패스 타임라인 생성 완료")
print("LLM 성공:", success_count)
print("Fallback 생성:", fallback_count)
print("실패 후 Fallback:", fail_count)

display(
    result_df[
        [
            "jd_id",
            "candidate_id",
            "rank",
            "career_path_summary",
            "career_fit_comment",
        ]
    ]
    .sort_values(["jd_id", "rank"])
    .head(15)
)


# In[38]:


# Cell 37 | 하드스킬 근거 JSON 생성 optional
# ------------------------------------------------------------
# 목적:
# - JD별 Top 후보자의 JD 요구 스킬과 후보자 스킬/경력 텍스트를 근거로
#   하드스킬 매칭 근거를 JSON 형태로 생성합니다.
# - 대시보드에서는 candidate 상세 영역의 skill evidence 컬럼으로 사용합니다.
# - LLM 실패 시에도 rule-based fallback을 생성해 스키마를 유지합니다.
# ------------------------------------------------------------

import os
import json
import re
import time
import math

import numpy as np
import pandas as pd


# ------------------------------------------------------------
# 1. 실행 옵션
# ------------------------------------------------------------

USE_HARDSKILL_EVIDENCE_LLM = globals().get("USE_HARDSKILL_EVIDENCE_LLM", True)         # 기본값 : False, 사용 시 True
HARDSKILL_EVIDENCE_MODEL_NAME = globals().get("HARDSKILL_EVIDENCE_MODEL_NAME", "gemini-2.5-flash")
LLM_EXPLAIN_TOP_N_PER_JD = globals().get("LLM_EXPLAIN_TOP_N_PER_JD", 3)
HARDSKILL_EVIDENCE_LLM_TIMEOUT = globals().get("HARDSKILL_EVIDENCE_LLM_TIMEOUT", 20)


# ------------------------------------------------------------
# 2. 기본 보조 함수
# ------------------------------------------------------------

def clean_value(value, default=""):
    if value is None:
        return default

    try:
        if pd.isna(value):
            return default
    except Exception:
        pass

    text = str(value).strip()

    if text.lower() in ["nan", "none", "null"]:
        return default

    return text


def clean_text(value):
    text = clean_value(value, "")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def compact_text(value, max_len=180):
    text = clean_text(value)

    if len(text) <= max_len:
        return text

    return text[:max_len].rstrip() + "..."


def make_json_safe(obj):
    if obj is None:
        return None

    if isinstance(obj, dict):
        return {str(k): make_json_safe(v) for k, v in obj.items()}

    if isinstance(obj, list):
        return [make_json_safe(v) for v in obj]

    if isinstance(obj, tuple):
        return [make_json_safe(v) for v in obj]

    if isinstance(obj, set):
        return [make_json_safe(v) for v in obj]

    if isinstance(obj, np.ndarray):
        return make_json_safe(obj.tolist())

    if isinstance(obj, np.integer):
        return int(obj)

    if isinstance(obj, np.floating):
        value = float(obj)
        if math.isnan(value) or math.isinf(value):
            return None
        return value

    if isinstance(obj, np.bool_):
        return bool(obj)

    if isinstance(obj, pd.Timestamp):
        return obj.isoformat()

    if isinstance(obj, float):
        if math.isnan(obj) or math.isinf(obj):
            return None
        return obj

    try:
        if pd.isna(obj):
            return None
    except Exception:
        pass

    return obj


def extract_json_from_text(text):
    if not isinstance(text, str):
        return None

    s = text.strip()
    s = re.sub(r"^```json\s*", "", s)
    s = re.sub(r"^```\s*", "", s)
    s = re.sub(r"\s*```$", "", s)

    try:
        return json.loads(s)
    except Exception:
        pass

    m = re.search(r"\{.*\}", s, flags=re.DOTALL)
    if m:
        try:
            return json.loads(m.group(0))
        except Exception:
            pass

    return None


def split_terms(value):
    text = clean_text(value)
    if not text:
        return []

    parts = re.split(r"[,;/|·•\n]+", text)
    terms = []

    for part in parts:
        term = clean_text(part)
        term = re.sub(r"^[\-\*\d\.\)\s]+", "", term).strip()

        if len(term) < 2:
            continue

        if term.lower() in ["nan", "none", "null"]:
            continue

        if term not in terms:
            terms.append(term)

    return terms


# ------------------------------------------------------------
# 3. Hard Skill Evidence Prompt
# ------------------------------------------------------------

HARDSKILL_EVIDENCE_SYSTEM_PROMPT = """
당신은 채용 매칭 대시보드의 하드스킬 근거 분석가입니다.
JD 요구 스킬과 후보자 이력서 텍스트를 근거로, 하드스킬 매칭 근거를 한국어 JSON으로 작성하세요.

반드시 아래 JSON schema만 반환하세요.
{
  "matched_skills": [
    {
      "skill": "매칭된 하드스킬명",
      "evidence": "후보자 원문에서 확인되는 근거",
      "jd_relevance": "JD 요구사항과 연결되는 이유"
    }
  ],
  "missing_or_weak_skills": [
    {
      "skill": "부족하거나 근거가 약한 스킬명",
      "reason": "부족하다고 판단한 이유"
    }
  ],
  "hardskill_summary": "하드스킬 관점 종합 요약"
}

규칙:
- 원문에 없는 기술 경험을 지어내지 마세요.
- 근거가 약하면 matched_skills가 아니라 missing_or_weak_skills에 넣으세요.
- matched_skills는 최대 8개, missing_or_weak_skills는 최대 5개로 제한하세요.
"""


# ------------------------------------------------------------
# 4. LLM 입력 payload 생성
# ------------------------------------------------------------

def build_hardskill_evidence_user_payload(row):
    return {
        "jd": {
            "jd_id": row.get("jd_id", ""),
            "company": row.get("jd_company", ""),
            "title": row.get("jd_title", ""),
            "required_skill": row.get("jd_required_skill", ""),
            "main_task": row.get("jd_main_task", ""),
            "domain": row.get("jd_domain", ""),
        },
        "candidate": {
            "candidate_id": row.get("candidate_id", ""),
            "job_role": row.get("cv_job_role", ""),
            "position": row.get("cv_position", ""),
            "skill": row.get("cv_skill", ""),
            "career": row.get("cv_career", ""),
            "career_description": row.get("cv_career_description", ""),
            "raw_text_excerpt": str(row.get("cv_raw_text", ""))[:5000],
        },
        "match_context": {
            "rank": row.get("rank", ""),
            "final_score_100": row.get("final_score_100", ""),
            "skill_score_100": row.get("skill_score_100", ""),
            "match_reason_rule_based": row.get("match_reason_rule_based", ""),
        },
    }


# ------------------------------------------------------------
# 5. Hard Skill 전용 Gemini 호출
# ------------------------------------------------------------

def call_hardskill_evidence_json(system_prompt, payload, model_name=None):
    if not globals().get("USE_HARDSKILL_EVIDENCE_LLM", False):
        return {"error": "HARDSKILL_EVIDENCE_LLM_DISABLED"}

    api_key = globals().get("GEMINI_API_KEY", "") or os.getenv("GEMINI_API_KEY", "")

    if not api_key:
        return {"error": "GEMINI_API_KEY가 없습니다."}

    if model_name is None:
        model_name = globals().get("HARDSKILL_EVIDENCE_MODEL_NAME", "gemini-2.5-flash")

    try:
        import google.generativeai as genai

        genai.configure(api_key=api_key)

        model = genai.GenerativeModel(
            model_name=model_name,
            system_instruction=system_prompt,
        )

        response = model.generate_content(
            json.dumps(make_json_safe(payload), ensure_ascii=False),
            request_options={"timeout": globals().get("HARDSKILL_EVIDENCE_LLM_TIMEOUT", 20)},
        )

        text = getattr(response, "text", "")
        parsed = extract_json_from_text(text)

        if isinstance(parsed, dict):
            parsed["_gemini_model_name"] = model_name
            return parsed

        return {
            "parse_error": "Gemini 응답을 JSON으로 파싱하지 못했습니다.",
            "raw_text": text,
            "_gemini_model_name": model_name,
        }

    except Exception as e:
        return {
            "error": repr(e),
            "_gemini_model_name": model_name,
        }


# ------------------------------------------------------------
# 6. Rule-based Hard Skill Evidence 생성
# ------------------------------------------------------------

def build_rule_based_hardskill_evidence(row):
    jd_terms = split_terms(row.get("jd_required_skill", ""))
    cv_terms = split_terms(row.get("cv_skill", ""))

    cv_text = clean_text(
        " ".join([
            clean_value(row.get("cv_skill", "")),
            clean_value(row.get("cv_career", "")),
            clean_value(row.get("cv_career_description", "")),
            clean_value(row.get("cv_raw_text", "")),
        ])
    )

    matched_skills = []
    missing_or_weak_skills = []

    for skill in jd_terms[:20]:
        skill_text = clean_text(skill)

        if not skill_text:
            continue

        if skill_text.lower() in cv_text.lower():
            matched_skills.append({
                "skill": skill_text,
                "evidence": compact_text(cv_text, 220),
                "jd_relevance": f"JD 요구 스킬 '{skill_text}'이 후보자 스킬/경력 텍스트에서 확인됩니다.",
            })
        else:
            missing_or_weak_skills.append({
                "skill": skill_text,
                "reason": f"후보자 스킬/경력 텍스트에서 '{skill_text}'에 대한 직접 근거가 약합니다.",
            })

    if not matched_skills and cv_terms:
        for skill in cv_terms[:5]:
            matched_skills.append({
                "skill": skill,
                "evidence": compact_text(cv_text, 220),
                "jd_relevance": "후보자 보유 스킬로 확인되며 JD 요구 역량과의 연결 검토가 필요합니다.",
            })

    matched_skills = matched_skills[:8]
    missing_or_weak_skills = missing_or_weak_skills[:5]

    if matched_skills:
        summary = f"후보자는 {', '.join([x['skill'] for x in matched_skills[:5]])} 등의 하드스킬 근거를 보유합니다."
    else:
        summary = "후보자의 하드스킬 근거는 원문 기준으로 제한적으로 확인됩니다."

    return {
        "matched_skills": matched_skills,
        "missing_or_weak_skills": missing_or_weak_skills,
        "hardskill_summary": summary,
    }


# ------------------------------------------------------------
# 7. LLM 출력 표준화
# ------------------------------------------------------------

def normalize_hardskill_evidence_output(out, row):
    if not isinstance(out, dict):
        return build_rule_based_hardskill_evidence(row)

    if out.get("error") or out.get("parse_error"):
        return build_rule_based_hardskill_evidence(row)

    matched_skills = out.get("matched_skills", [])
    missing_or_weak_skills = out.get("missing_or_weak_skills", [])
    summary = clean_value(out.get("hardskill_summary", ""))

    if not isinstance(matched_skills, list):
        matched_skills = []

    if not isinstance(missing_or_weak_skills, list):
        missing_or_weak_skills = []

    cleaned_matched = []

    for item in matched_skills[:8]:
        if not isinstance(item, dict):
            continue

        cleaned_matched.append({
            "skill": compact_text(item.get("skill", ""), 80),
            "evidence": compact_text(item.get("evidence", ""), 260),
            "jd_relevance": compact_text(item.get("jd_relevance", ""), 220),
        })

    cleaned_missing = []

    for item in missing_or_weak_skills[:5]:
        if not isinstance(item, dict):
            continue

        cleaned_missing.append({
            "skill": compact_text(item.get("skill", ""), 80),
            "reason": compact_text(item.get("reason", ""), 220),
        })

    if not summary:
        summary = build_rule_based_hardskill_evidence(row).get("hardskill_summary", "")

    return {
        "matched_skills": cleaned_matched,
        "missing_or_weak_skills": cleaned_missing,
        "hardskill_summary": summary,
    }


# ------------------------------------------------------------
# 8. result_df 확인 및 결과 컬럼 초기화
# ------------------------------------------------------------

if "result_df" not in globals():
    raise NameError("result_df가 없습니다. 이전 매칭 결과 셀을 먼저 실행하세요.")

if result_df.empty:
    raise ValueError("result_df가 비어 있습니다. 이전 매칭 결과 셀을 확인하세요.")

hardskill_cols = [
    "hardskill_evidence_json",
    "hardskill_evidence_summary",
]

for col in hardskill_cols:
    if col not in result_df.columns:
        result_df[col] = ""
    else:
        result_df[col] = ""


# ------------------------------------------------------------
# 9. Hard Skill Evidence 대상 선정
# ------------------------------------------------------------

target_indices = (
    result_df[result_df["rank"] <= LLM_EXPLAIN_TOP_N_PER_JD]
    .sort_values(["jd_id", "rank"], ascending=[True, True])
    .index
    .tolist()
)

print("하드스킬 근거 생성 대상:", len(target_indices))
print("JD별 Top N:", LLM_EXPLAIN_TOP_N_PER_JD)
print("LLM 사용:", USE_HARDSKILL_EVIDENCE_LLM)
print("LLM 모델:", HARDSKILL_EVIDENCE_MODEL_NAME if USE_HARDSKILL_EVIDENCE_LLM else "사용 안 함")
print()


# ------------------------------------------------------------
# 10. Hard Skill Evidence 생성
# ------------------------------------------------------------

success_count = 0
fallback_count = 0
fail_count = 0

for n, idx in enumerate(target_indices, start=1):
    row = result_df.loc[idx]

    jd_id = row.get("jd_id", "")
    candidate_id = row.get("candidate_id", "")
    rank = row.get("rank", "")

    print(f"[{n}/{len(target_indices)}] jd_id={jd_id}, candidate_id={candidate_id}, rank={rank}")

    try:
        out = {}

        if USE_HARDSKILL_EVIDENCE_LLM and GEMINI_API_KEY:
            payload = build_hardskill_evidence_user_payload(row)
            payload = make_json_safe(payload)

            out = call_hardskill_evidence_json(
                HARDSKILL_EVIDENCE_SYSTEM_PROMPT,
                payload,
                model_name=HARDSKILL_EVIDENCE_MODEL_NAME,
            )

        normalized = normalize_hardskill_evidence_output(out, row)

        if not out or out.get("error") or out.get("parse_error"):
            fallback_count += 1
            status = "fallback"
        else:
            success_count += 1
            status = "success"

        result_df.loc[idx, "hardskill_evidence_json"] = json.dumps(
            make_json_safe({
                "matched_skills": normalized.get("matched_skills", []),
                "missing_or_weak_skills": normalized.get("missing_or_weak_skills", []),
            }),
            ensure_ascii=False,
        )
        result_df.loc[idx, "hardskill_evidence_summary"] = clean_value(
            normalized.get("hardskill_summary", "")
        )

        print("  완료:", status)

        if USE_HARDSKILL_EVIDENCE_LLM and GEMINI_API_KEY:
            time.sleep(0.5)

    except Exception as e:
        fail_count += 1

        fallback = build_rule_based_hardskill_evidence(row)

        result_df.loc[idx, "hardskill_evidence_json"] = json.dumps(
            make_json_safe({
                "matched_skills": fallback.get("matched_skills", []),
                "missing_or_weak_skills": fallback.get("missing_or_weak_skills", []),
            }),
            ensure_ascii=False,
        )
        result_df.loc[idx, "hardskill_evidence_summary"] = clean_value(
            fallback.get("hardskill_summary", "")
        )

        print("  실패 후 fallback:", repr(e)[:300])


# ------------------------------------------------------------
# 11. 실행 결과 확인
# ------------------------------------------------------------

print()
print("하드스킬 근거 생성 완료")
print("LLM 성공:", success_count)
print("Fallback 생성:", fallback_count)
print("실패 후 Fallback:", fail_count)

display(
    result_df[
        [
            "jd_id",
            "candidate_id",
            "rank",
            "hardskill_evidence_summary",
            "hardskill_evidence_json",
        ]
    ]
    .sort_values(["jd_id", "rank"])
    .head(15)
)


# In[39]:


# Cell 38 | Soft Skill 근거 JSON 생성 optional
# ------------------------------------------------------------
# 목적:
# - JD별 Top 후보자의 경력/자기소개/Raw_Text를 근거로
#   협업, 커뮤니케이션, 문제해결 등 Soft Skill 근거를 JSON 형태로 생성합니다.
# - 대시보드에서는 candidate 상세 영역의 soft skill evidence 컬럼으로 사용합니다.
# - LLM 실패 시에도 rule-based fallback을 생성해 스키마를 유지합니다.
# ------------------------------------------------------------

import os
import json
import re
import time
import math

import numpy as np
import pandas as pd


# ------------------------------------------------------------
# 1. 실행 옵션
# ------------------------------------------------------------

USE_SOFTSKILL_LLM = globals().get("USE_SOFTSKILL_LLM", True)           # 기본값 False, 사용 시 True로 변경
SOFTSKILL_MODEL_NAME = globals().get("SOFTSKILL_MODEL_NAME", "gemini-2.5-flash")
LLM_EXPLAIN_TOP_N_PER_JD = globals().get("LLM_EXPLAIN_TOP_N_PER_JD", 3)
SOFTSKILL_LLM_TIMEOUT = globals().get("SOFTSKILL_LLM_TIMEOUT", 20)


# ------------------------------------------------------------
# 2. 기본 보조 함수
# ------------------------------------------------------------

def clean_value(value, default=""):
    if value is None:
        return default

    try:
        if pd.isna(value):
            return default
    except Exception:
        pass

    text = str(value).strip()

    if text.lower() in ["nan", "none", "null"]:
        return default

    return text


def clean_text(value):
    text = clean_value(value, "")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def compact_text(value, max_len=180):
    text = clean_text(value)

    if len(text) <= max_len:
        return text

    return text[:max_len].rstrip() + "..."


def make_json_safe(obj):
    if obj is None:
        return None

    if isinstance(obj, dict):
        return {str(k): make_json_safe(v) for k, v in obj.items()}

    if isinstance(obj, list):
        return [make_json_safe(v) for v in obj]

    if isinstance(obj, tuple):
        return [make_json_safe(v) for v in obj]

    if isinstance(obj, set):
        return [make_json_safe(v) for v in obj]

    if isinstance(obj, np.ndarray):
        return make_json_safe(obj.tolist())

    if isinstance(obj, np.integer):
        return int(obj)

    if isinstance(obj, np.floating):
        value = float(obj)
        if math.isnan(value) or math.isinf(value):
            return None
        return value

    if isinstance(obj, np.bool_):
        return bool(obj)

    if isinstance(obj, pd.Timestamp):
        return obj.isoformat()

    if isinstance(obj, float):
        if math.isnan(obj) or math.isinf(obj):
            return None
        return obj

    try:
        if pd.isna(obj):
            return None
    except Exception:
        pass

    return obj


def extract_json_from_text(text):
    if not isinstance(text, str):
        return None

    s = text.strip()
    s = re.sub(r"^```json\s*", "", s)
    s = re.sub(r"^```\s*", "", s)
    s = re.sub(r"\s*```$", "", s)

    try:
        return json.loads(s)
    except Exception:
        pass

    m = re.search(r"\{.*\}", s, flags=re.DOTALL)
    if m:
        try:
            return json.loads(m.group(0))
        except Exception:
            pass

    return None


# ------------------------------------------------------------
# 3. Soft Skill Prompt
# ------------------------------------------------------------

SOFTSKILL_SYSTEM_PROMPT = """
당신은 채용 매칭 대시보드의 소프트스킬 근거 분석가입니다.
후보자의 경력/자기소개/Raw_Text를 근거로 협업, 커뮤니케이션, 문제해결, 리더십, 자기주도성 등의 소프트스킬 근거를 한국어 JSON으로 작성하세요.

반드시 아래 JSON schema만 반환하세요.
{
  "softskills": [
    {
      "skill": "소프트스킬명",
      "evidence": "후보자 원문에서 확인되는 근거",
      "jd_relevance": "JD와 연결되는 이유"
    }
  ],
  "softskill_summary": "소프트스킬 관점 종합 요약"
}

규칙:
- 원문에 없는 성격, 태도, 리더십을 지어내지 마세요.
- 근거가 약하면 단정하지 말고 '근거 제한적'이라고 표현하세요.
- softskills는 최대 6개로 제한하세요.
"""


# ------------------------------------------------------------
# 4. LLM 입력 payload 생성
# ------------------------------------------------------------

def build_softskill_user_payload(row):
    return {
        "jd": {
            "jd_id": row.get("jd_id", ""),
            "company": row.get("jd_company", ""),
            "title": row.get("jd_title", ""),
            "main_task": row.get("jd_main_task", ""),
            "domain": row.get("jd_domain", ""),
            "role_signal": row.get("jd_role_signal", ""),
        },
        "candidate": {
            "candidate_id": row.get("candidate_id", ""),
            "job_role": row.get("cv_job_role", ""),
            "position": row.get("cv_position", ""),
            "career": row.get("cv_career", ""),
            "career_description": row.get("cv_career_description", ""),
            "intro": row.get("cv_intro", ""),
            "raw_text_excerpt": str(row.get("cv_raw_text", ""))[:5000],
        },
        "match_context": {
            "rank": row.get("rank", ""),
            "final_score_100": row.get("final_score_100", ""),
            "task_score_100": row.get("task_score_100", ""),
            "role_score_100": row.get("role_score_100", ""),
            "match_reason_rule_based": row.get("match_reason_rule_based", ""),
        },
    }


# ------------------------------------------------------------
# 5. Soft Skill 전용 Gemini 호출
# ------------------------------------------------------------

def call_softskill_json(system_prompt, payload, model_name=None):
    if not globals().get("USE_SOFTSKILL_LLM", False):
        return {"error": "SOFTSKILL_LLM_DISABLED"}

    api_key = globals().get("GEMINI_API_KEY", "") or os.getenv("GEMINI_API_KEY", "")

    if not api_key:
        return {"error": "GEMINI_API_KEY가 없습니다."}

    if model_name is None:
        model_name = globals().get("SOFTSKILL_MODEL_NAME", "gemini-2.5-flash")

    try:
        import google.generativeai as genai

        genai.configure(api_key=api_key)

        model = genai.GenerativeModel(
            model_name=model_name,
            system_instruction=system_prompt,
        )

        response = model.generate_content(
            json.dumps(make_json_safe(payload), ensure_ascii=False),
            request_options={"timeout": globals().get("SOFTSKILL_LLM_TIMEOUT", 20)},
        )

        text = getattr(response, "text", "")
        parsed = extract_json_from_text(text)

        if isinstance(parsed, dict):
            parsed["_gemini_model_name"] = model_name
            return parsed

        return {
            "parse_error": "Gemini 응답을 JSON으로 파싱하지 못했습니다.",
            "raw_text": text,
            "_gemini_model_name": model_name,
        }

    except Exception as e:
        return {
            "error": repr(e),
            "_gemini_model_name": model_name,
        }


# ------------------------------------------------------------
# 6. Rule-based Soft Skill Evidence 생성
# ------------------------------------------------------------

def build_rule_based_softskill(row):
    text = clean_text(
        " ".join([
            clean_value(row.get("cv_career", "")),
            clean_value(row.get("cv_career_description", "")),
            clean_value(row.get("cv_intro", "")),
            clean_value(row.get("cv_raw_text", "")),
        ])
    )

    reason = clean_text(row.get("match_reason_rule_based", ""))

    skill_patterns = [
        ("협업", ["협업", "협력", "팀", "cross-functional", "stakeholder", "커뮤니케이션"]),
        ("커뮤니케이션", ["커뮤니케이션", "소통", "보고", "발표", "문서화", "협의"]),
        ("문제해결", ["문제 해결", "이슈", "개선", "최적화", "해결", "트러블슈팅"]),
        ("리더십", ["리드", "관리", "PM", "프로젝트 관리", "멘토링", "조율"]),
        ("자기주도성", ["주도", "기획", "설계", "도입", "구축", "자동화"]),
        ("분석적 사고", ["분석", "데이터", "지표", "모델링", "검증", "평가"]),
    ]

    softskills = []

    lower_text = text.lower()

    for skill, keywords in skill_patterns:
        matched_keywords = [
            kw for kw in keywords
            if kw.lower() in lower_text
        ]

        if matched_keywords:
            softskills.append({
                "skill": skill,
                "evidence": compact_text(text, 260),
                "jd_relevance": compact_text(reason or f"후보자 원문에서 {', '.join(matched_keywords[:3])} 관련 표현이 확인됩니다.", 220),
            })

    if not softskills:
        softskills = [{
            "skill": "근거 제한적",
            "evidence": compact_text(text, 260),
            "jd_relevance": compact_text(reason or "원문 기준으로 명확한 소프트스킬 근거는 제한적입니다.", 220),
        }]

    softskills = softskills[:6]

    if softskills and softskills[0].get("skill") != "근거 제한적":
        summary = f"후보자는 {', '.join([x['skill'] for x in softskills[:4]])} 관련 소프트스킬 근거가 확인됩니다."
    else:
        summary = "후보자의 소프트스킬 근거는 원문 기준으로 제한적으로 확인됩니다."

    return {
        "softskills": softskills,
        "softskill_summary": summary,
    }


# ------------------------------------------------------------
# 7. LLM 출력 표준화
# ------------------------------------------------------------

def normalize_softskill_output(out, row):
    if not isinstance(out, dict):
        return build_rule_based_softskill(row)

    if out.get("error") or out.get("parse_error"):
        return build_rule_based_softskill(row)

    softskills = out.get("softskills", [])
    summary = clean_value(out.get("softskill_summary", ""))

    if not isinstance(softskills, list):
        softskills = []

    cleaned_softskills = []

    for item in softskills[:6]:
        if not isinstance(item, dict):
            continue

        cleaned_softskills.append({
            "skill": compact_text(item.get("skill", ""), 80),
            "evidence": compact_text(item.get("evidence", ""), 260),
            "jd_relevance": compact_text(item.get("jd_relevance", ""), 220),
        })

    if not summary:
        summary = build_rule_based_softskill(row).get("softskill_summary", "")

    if not cleaned_softskills:
        return build_rule_based_softskill(row)

    return {
        "softskills": cleaned_softskills,
        "softskill_summary": summary,
    }


# ------------------------------------------------------------
# 8. result_df 확인 및 결과 컬럼 초기화
# ------------------------------------------------------------

if "result_df" not in globals():
    raise NameError("result_df가 없습니다. 이전 매칭 결과 셀을 먼저 실행하세요.")

if result_df.empty:
    raise ValueError("result_df가 비어 있습니다. 이전 매칭 결과 셀을 확인하세요.")

softskill_cols = [
    "softskill_json",
    "softskill_summary",
]

for col in softskill_cols:
    if col not in result_df.columns:
        result_df[col] = ""
    else:
        result_df[col] = ""


# ------------------------------------------------------------
# 9. Soft Skill 대상 선정
# ------------------------------------------------------------

target_indices = (
    result_df[result_df["rank"] <= LLM_EXPLAIN_TOP_N_PER_JD]
    .sort_values(["jd_id", "rank"], ascending=[True, True])
    .index
    .tolist()
)

print("소프트스킬 근거 생성 대상:", len(target_indices))
print("JD별 Top N:", LLM_EXPLAIN_TOP_N_PER_JD)
print("LLM 사용:", USE_SOFTSKILL_LLM)
print("LLM 모델:", SOFTSKILL_MODEL_NAME if USE_SOFTSKILL_LLM else "사용 안 함")
print()


# ------------------------------------------------------------
# 10. Soft Skill Evidence 생성
# ------------------------------------------------------------

success_count = 0
fallback_count = 0
fail_count = 0

for n, idx in enumerate(target_indices, start=1):
    row = result_df.loc[idx]

    jd_id = row.get("jd_id", "")
    candidate_id = row.get("candidate_id", "")
    rank = row.get("rank", "")

    print(f"[{n}/{len(target_indices)}] jd_id={jd_id}, candidate_id={candidate_id}, rank={rank}")

    try:
        out = {}

        if USE_SOFTSKILL_LLM and GEMINI_API_KEY:
            payload = build_softskill_user_payload(row)
            payload = make_json_safe(payload)

            out = call_softskill_json(
                SOFTSKILL_SYSTEM_PROMPT,
                payload,
                model_name=SOFTSKILL_MODEL_NAME,
            )

        normalized = normalize_softskill_output(out, row)

        if not out or out.get("error") or out.get("parse_error"):
            fallback_count += 1
            status = "fallback"
        else:
            success_count += 1
            status = "success"

        result_df.loc[idx, "softskill_json"] = json.dumps(
            make_json_safe({
                "softskills": normalized.get("softskills", []),
            }),
            ensure_ascii=False,
        )
        result_df.loc[idx, "softskill_summary"] = clean_value(
            normalized.get("softskill_summary", "")
        )

        print("  완료:", status)

        if USE_SOFTSKILL_LLM and GEMINI_API_KEY:
            time.sleep(0.5)

    except Exception as e:
        fail_count += 1

        fallback = build_rule_based_softskill(row)

        result_df.loc[idx, "softskill_json"] = json.dumps(
            make_json_safe({
                "softskills": fallback.get("softskills", []),
            }),
            ensure_ascii=False,
        )
        result_df.loc[idx, "softskill_summary"] = clean_value(
            fallback.get("softskill_summary", "")
        )

        print("  실패 후 fallback:", repr(e)[:300])


# ------------------------------------------------------------
# 11. 실행 결과 확인
# ------------------------------------------------------------

print()
print("소프트스킬 근거 생성 완료")
print("LLM 성공:", success_count)
print("Fallback 생성:", fallback_count)
print("실패 후 Fallback:", fail_count)

display(
    result_df[
        [
            "jd_id",
            "candidate_id",
            "rank",
            "softskill_summary",
            "softskill_json",
        ]
    ]
    .sort_values(["jd_id", "rank"])
    .head(15)
)


# In[40]:


# Cell 39 | 하드스킬 Evidence / 소프트스킬 Evidence 생성
# ------------------------------------------------------------
# 목적:
# - JD별 Top 후보자의 하드스킬/소프트스킬 근거를 대시보드용 JSON 컬럼으로 생성합니다.
# - Gemini / LLM 호출 없이 rule-based fallback 기준으로 즉시 생성합니다.
# - 기존 call_gemini_json()을 호출하지 않아 gRPC 대기/무응답 문제를 피합니다.
# ------------------------------------------------------------

import json
import re

import pandas as pd


# ------------------------------------------------------------
# 1. 기본 보조 함수
# ------------------------------------------------------------

def clean_value(value, default=""):
    if value is None:
        return default

    try:
        if pd.isna(value):
            return default
    except Exception:
        pass

    text = str(value).strip()

    if text.lower() in ["nan", "none", "null"]:
        return default

    return text


def clean_text(value):
    text = clean_value(value, "")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def compact_text(value, max_len=220):
    text = clean_text(value)

    if len(text) <= max_len:
        return text

    return text[:max_len].rstrip() + "..."


def split_terms(value):
    text = clean_text(value)

    if not text:
        return []

    parts = re.split(r"[,;/|·•\n]+", text)
    terms = []

    for part in parts:
        term = clean_text(part)
        term = re.sub(r"^[\-\*\d\.\)\s]+", "", term).strip()

        if len(term) < 2:
            continue

        if term.lower() in ["nan", "none", "null"]:
            continue

        if term not in terms:
            terms.append(term)

    return terms


# ------------------------------------------------------------
# 2. 하드스킬 Evidence 생성
# ------------------------------------------------------------

def build_rule_based_hardskill_evidence(row):
    jd_terms = split_terms(row.get("jd_required_skill", ""))
    cv_terms = split_terms(row.get("cv_skill", ""))

    cv_text = clean_text(
        " ".join([
            clean_value(row.get("cv_skill", "")),
            clean_value(row.get("cv_career", "")),
            clean_value(row.get("cv_career_description", "")),
            clean_value(row.get("cv_raw_text", "")),
        ])
    )

    matched_skills = []
    missing_or_weak_skills = []

    for skill in jd_terms[:20]:
        skill_text = clean_text(skill)

        if not skill_text:
            continue

        if skill_text.lower() in cv_text.lower():
            matched_skills.append({
                "skill": skill_text,
                "evidence": compact_text(cv_text, 260),
                "jd_relevance": f"JD 요구 스킬 '{skill_text}'이 후보자 스킬/경력 텍스트에서 확인됩니다.",
            })
        else:
            missing_or_weak_skills.append({
                "skill": skill_text,
                "reason": f"후보자 스킬/경력 텍스트에서 '{skill_text}'에 대한 직접 근거가 약합니다.",
            })

    if not matched_skills and cv_terms:
        for skill in cv_terms[:5]:
            matched_skills.append({
                "skill": skill,
                "evidence": compact_text(cv_text, 260),
                "jd_relevance": "후보자 보유 스킬로 확인되며 JD 요구 역량과의 연결 검토가 필요합니다.",
            })

    matched_skills = matched_skills[:8]
    missing_or_weak_skills = missing_or_weak_skills[:5]

    if matched_skills:
        summary = f"후보자는 {', '.join([x['skill'] for x in matched_skills[:5]])} 등의 하드스킬 근거를 보유합니다."
    else:
        summary = "후보자의 하드스킬 근거는 원문 기준으로 제한적으로 확인됩니다."

    return {
        "matched_skills": matched_skills,
        "missing_or_weak_skills": missing_or_weak_skills,
        "hardskill_summary": summary,
    }


# ------------------------------------------------------------
# 3. 소프트스킬 Evidence 생성
# ------------------------------------------------------------

def build_rule_based_softskill(row):
    text = clean_text(
        " ".join([
            clean_value(row.get("cv_career", "")),
            clean_value(row.get("cv_career_description", "")),
            clean_value(row.get("cv_intro", "")),
            clean_value(row.get("cv_raw_text", "")),
        ])
    )

    reason = clean_text(row.get("match_reason_rule_based", ""))

    skill_patterns = [
        ("협업", ["협업", "협력", "팀", "cross-functional", "stakeholder", "커뮤니케이션"]),
        ("커뮤니케이션", ["커뮤니케이션", "소통", "보고", "발표", "문서화", "협의"]),
        ("문제해결", ["문제 해결", "이슈", "개선", "최적화", "해결", "트러블슈팅"]),
        ("리더십", ["리드", "관리", "PM", "프로젝트 관리", "멘토링", "조율"]),
        ("자기주도성", ["주도", "기획", "설계", "도입", "구축", "자동화"]),
        ("분석적 사고", ["분석", "데이터", "지표", "모델링", "검증", "평가"]),
    ]

    softskills = []
    lower_text = text.lower()

    for skill, keywords in skill_patterns:
        matched_keywords = [
            kw for kw in keywords
            if kw.lower() in lower_text
        ]

        if matched_keywords:
            softskills.append({
                "skill": skill,
                "evidence": compact_text(text, 260),
                "jd_relevance": compact_text(
                    reason or f"후보자 원문에서 {', '.join(matched_keywords[:3])} 관련 표현이 확인됩니다.",
                    220,
                ),
            })

    if not softskills:
        softskills = [{
            "skill": "근거 제한적",
            "evidence": compact_text(text, 260),
            "jd_relevance": compact_text(
                reason or "원문 기준으로 명확한 소프트스킬 근거는 제한적입니다.",
                220,
            ),
        }]

    softskills = softskills[:6]

    if softskills and softskills[0].get("skill") != "근거 제한적":
        summary = f"후보자는 {', '.join([x['skill'] for x in softskills[:4]])} 관련 소프트스킬 근거가 확인됩니다."
    else:
        summary = "후보자의 소프트스킬 근거는 원문 기준으로 제한적으로 확인됩니다."

    return {
        "softskills": softskills,
        "softskill_summary": summary,
    }


# ------------------------------------------------------------
# 4. result_df 확인 및 컬럼 준비
# ------------------------------------------------------------

if "result_df" not in globals():
    raise NameError("result_df가 없습니다. 이전 매칭 결과 셀을 먼저 실행하세요.")

if result_df.empty:
    raise ValueError("result_df가 비어 있습니다. 이전 매칭 결과 셀을 확인하세요.")

for col in [
    "hardskill_evidence_json",
    "hardskill_evidence_summary",
    "softskill_json",
    "softskill_summary",
]:
    if col not in result_df.columns:
        result_df[col] = ""


# ------------------------------------------------------------
# 5. 대상 선정
# ------------------------------------------------------------

LLM_EXPLAIN_TOP_N_PER_JD = globals().get("LLM_EXPLAIN_TOP_N_PER_JD", 3)

target_indices = (
    result_df[result_df["rank"] <= LLM_EXPLAIN_TOP_N_PER_JD]
    .sort_values(["jd_id", "rank"], ascending=[True, True])
    .index
    .tolist()
)

print("하드스킬/소프트스킬 Evidence 생성 대상:", len(target_indices))
print("JD별 Top N:", LLM_EXPLAIN_TOP_N_PER_JD)
print("LLM 사용: False")
print()


# ------------------------------------------------------------
# 6. Evidence 생성
# ------------------------------------------------------------

for n, idx in enumerate(target_indices, start=1):
    row = result_df.loc[idx]

    jd_id = row.get("jd_id", "")
    candidate_id = row.get("candidate_id", "")
    rank = row.get("rank", "")

    print(f"[{n}/{len(target_indices)}] jd_id={jd_id}, candidate_id={candidate_id}, rank={rank}")

    hard = build_rule_based_hardskill_evidence(row)
    soft = build_rule_based_softskill(row)

    result_df.loc[idx, "hardskill_evidence_json"] = json.dumps(
        {
            "matched_skills": hard.get("matched_skills", []),
            "missing_or_weak_skills": hard.get("missing_or_weak_skills", []),
        },
        ensure_ascii=False,
    )

    result_df.loc[idx, "hardskill_evidence_summary"] = clean_value(
        hard.get("hardskill_summary", "")
    )

    result_df.loc[idx, "softskill_json"] = json.dumps(
        {
            "softskills": soft.get("softskills", []),
        },
        ensure_ascii=False,
    )

    result_df.loc[idx, "softskill_summary"] = clean_value(
        soft.get("softskill_summary", "")
    )

    print("  완료")


# ------------------------------------------------------------
# 7. 실행 결과 확인
# ------------------------------------------------------------

print()
print("하드스킬/소프트스킬 Evidence 생성 완료")

display(
    result_df[
        [
            "jd_id",
            "candidate_id",
            "rank",
            "hardskill_evidence_summary",
            "softskill_summary",
        ]
    ]
    .sort_values(["jd_id", "rank"])
    .head(15)
)


# In[41]:


# Cell 40 | 대시보드 Summary Table 생성
# Cell 25 | 대시보드 Summary Table 생성
# ------------------------------------------------------------
# 대시보드 전용 노트북이 읽을 요약 테이블입니다.
# 1행 = JD 1개 기준 Top 후보 1명
# ------------------------------------------------------------

dashboard_base_cols = [
    "jd_id", "jd_company", "jd_title", "jd_required_skill", "jd_main_task",
    "jd_domain", "jd_role_signal",
    "candidate_id", "cv_job_role", "cv_total_career_years", "cv_position",
    "cv_education", "cv_gpa", "cv_certificates", "cv_languages", "cv_paper",
    "cv_overseas_experience", "cv_training", "cv_awards", "cv_employment_status", "cv_salary",
    "rank", "final_score_100", "retrieval_score_norm_100", "column_score_100",
    "skill_score_100", "ai_skill_overlap_score_100", "task_score_100",
    "domain_score_100", "role_score_100",
    "experience_score_100", "soft_skill_score_100",
    "radar_skill_score_100", "radar_task_score_100", "radar_domain_score_100",
    "radar_role_score_100", "radar_experience_score_100", "radar_soft_skill_score_100",
    "skill_risk_flag", "jd_skill_standard", "cv_skill_standard",
    "jd_ai_skill_tags", "cv_ai_skill_tags", "jd_task_standard", "cv_task_standard",
    "match_reason_rule_based", "match_gap_rule_based", "interview_question_rule_based",
    "llm_match_summary", "llm_strength", "llm_gap", "llm_risk",
    "llm_interview_question", "llm_evidence",
    "career_path_summary", "career_fit_comment", "career_timeline_json",
    "hardskill_evidence_json", "softskill_json",
]

# 없을 수 있는 컬럼은 빈 문자열로 생성해서 스키마를 고정합니다.
for col in dashboard_base_cols:
    if col not in result_df.columns:
        result_df[col] = ""

dashboard_summary_df = result_df[dashboard_base_cols].copy()

print("dashboard_summary_df:", dashboard_summary_df.shape)
display(dashboard_summary_df.head(15))



# In[42]:


# Cell 41 | final_score 기준 전체 후보 랭킹 생성 - 855+10 / 855only v5
# ------------------------------------------------------------
# 목적:
# - 기존 latest_matching_result_all은 JD별 Top 10만 저장하므로 나머지 후보를 확인하기 어렵습니다.
# - 이 셀은 Cell 19에서 계산된 pair_df의 final_score를 사용하여 JD별 전체 후보를 다시 정렬합니다.
# - 저장 기준은 universe_embedding_score 단독 점수가 아니라 latest_matching_result_all과 동일한 final_score입니다.
# - 855+10 v5 파일: 실제 CV 855명 + 가상 CV 10명 = 865명 전체 후보 저장
# - 855only v5 파일: 가상 CV를 제외한 실제 CV 855명만 저장
#
# 중요:
# - 이 셀은 점수를 새로 계산하지 않습니다.
# - 전체 후보 점수를 얻으려면 Cell 01의 TOP_N_RETRIEVAL이 전체 후보 수 이상이어야 합니다.
# - CrossEncoder를 사용하는 설정이면, Cell 16에서 전체 후보에 대해 CrossEncoder 점수가 계산된 뒤 이 셀을 실행해야 합니다.
# ------------------------------------------------------------

import numpy as np
import pandas as pd


def _ensure_numeric_score(df, col):
    """점수 컬럼을 숫자형으로 변환하고 변환 불가 값은 NaN으로 둡니다."""
    if col in df.columns:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    return df


def _add_score_100_columns(df):
    """주요 0~1 점수 컬럼의 100점 환산 컬럼을 생성합니다."""
    score_cols_for_100 = [
        "final_score",
        "final_score_before_role_penalty",
        "retrieval_score",
        "retrieval_score_norm",
        "keyword_baseline_score",
        "keyword_baseline_score_norm",
        "column_score",
        "column_score_before_skill_penalty",
        "skill_score",
        "task_score",
        "domain_score",
        "role_score",
        "ai_skill_overlap_score",
    ]
    for col in score_cols_for_100:
        if col in df.columns:
            df[f"{col}_100"] = (pd.to_numeric(df[col], errors="coerce") * 100).round(2)
    return df


def _add_candidate_type(df):
    """candidate_id 기준으로 가상 CV / 실제 CV 구분값을 보정합니다."""
    if "candidate_type" not in df.columns:
        df["candidate_type"] = np.where(
            df["candidate_id"].astype(str).str.startswith("G_JD_"),
            "virtual_cv",
            "real_cv",
        )
    else:
        df["candidate_type"] = df["candidate_type"].fillna(
            np.where(
                df["candidate_id"].astype(str).str.startswith("G_JD_"),
                "virtual_cv",
                "real_cv",
            )
        )
    return df


def _sort_columns_for_export(df):
    """제출/검수용으로 주요 컬럼을 앞으로 배치하고 나머지 컬럼은 뒤에 보존합니다."""
    preferred_cols = [
        "jd_id",
        "jd_title",
        "jd_name",
        "company",
        "job_title",
        "candidate_id",
        "resume_no",
        "file_name",
        "candidate_type",
        "rank",
        "final_score",
        "final_score_100",
        "final_score_before_role_penalty",
        "final_score_before_role_penalty_100",
        "retrieval_score",
        "retrieval_score_100",
        "retrieval_score_norm",
        "retrieval_score_norm_100",
        "keyword_baseline_score",
        "keyword_baseline_score_100",
        "keyword_baseline_score_norm",
        "keyword_baseline_score_norm_100",
        "column_score",
        "column_score_100",
        "column_score_before_skill_penalty",
        "column_score_before_skill_penalty_100",
        "skill_score",
        "skill_score_100",
        "task_score",
        "task_score_100",
        "domain_score",
        "domain_score_100",
        "role_score",
        "role_score_100",
        "ai_skill_overlap_score",
        "ai_skill_overlap_score_100",
        "skill_risk_flag",
        "skill_risk_penalty",
        "role_mismatch_penalty",
        "scoring_method",
    ]
    save_cols = [col for col in preferred_cols if col in df.columns]
    extra_cols = [col for col in df.columns if col not in save_cols]
    return df[save_cols + extra_cols].copy()


if "pair_df" not in globals():
    raise NameError(
        "pair_df가 메모리에 없습니다. Cell 13~19를 먼저 실행하여 final_score를 계산해 주세요."
    )

if not isinstance(pair_df, pd.DataFrame):
    raise TypeError(f"pair_df가 pandas DataFrame이 아닙니다. 현재 타입: {type(pair_df)}")

required_cols = ["jd_id", "candidate_id", "final_score"]
missing_cols = [col for col in required_cols if col not in pair_df.columns]
if missing_cols:
    raise KeyError(
        f"전체 후보 랭킹 생성에 필요한 필수 컬럼이 없습니다: {missing_cols}\n"
        "Cell 19 | 점수 정규화 및 final_score 계산 셀까지 실행한 뒤 다시 실행하세요."
    )

full_rank_source_df = pair_df.copy()
full_rank_source_df["jd_id"] = full_rank_source_df["jd_id"].astype(str)
full_rank_source_df["candidate_id"] = full_rank_source_df["candidate_id"].astype(str)
full_rank_source_df = _ensure_numeric_score(full_rank_source_df, "final_score")

before_rows = len(full_rank_source_df)
full_rank_source_df = full_rank_source_df.dropna(subset=["final_score"]).copy()
after_rows = len(full_rank_source_df)
if after_rows < before_rows:
    print(f"주의: final_score 결측 row 제거: {before_rows - after_rows:,}건")

# 1) 855+10: 가상 CV 포함 전체 후보 final_score 랭킹
full_ranking_855plus10_df = (
    full_rank_source_df
    .sort_values(["jd_id", "final_score"], ascending=[True, False])
    .reset_index(drop=True)
)
full_ranking_855plus10_df["rank"] = full_ranking_855plus10_df.groupby("jd_id").cumcount() + 1
full_ranking_855plus10_df = _add_candidate_type(full_ranking_855plus10_df)
full_ranking_855plus10_df = _add_score_100_columns(full_ranking_855plus10_df)
full_ranking_855plus10_df = _sort_columns_for_export(full_ranking_855plus10_df)

# 2) 855only: 가상 CV 제외 실제 CV만 final_score 랭킹
full_ranking_855only_df = full_ranking_855plus10_df[
    full_ranking_855plus10_df["candidate_type"].astype(str).ne("virtual_cv")
].copy()
full_ranking_855only_df = (
    full_ranking_855only_df
    .sort_values(["jd_id", "final_score"], ascending=[True, False])
    .reset_index(drop=True)
)
full_ranking_855only_df["rank"] = full_ranking_855only_df.groupby("jd_id").cumcount() + 1
full_ranking_855only_df = _sort_columns_for_export(full_ranking_855only_df)

# 검증 요약
plus10_counts = full_ranking_855plus10_df.groupby("jd_id")["candidate_id"].nunique()
only_counts = full_ranking_855only_df.groupby("jd_id")["candidate_id"].nunique()

print("final_score 기준 전체 후보 랭킹 생성 완료")
print("[855+10] JD 수:", full_ranking_855plus10_df["jd_id"].nunique())
print("[855+10] JD별 후보 수 min/max:", int(plus10_counts.min()), "/", int(plus10_counts.max()))
print("[855+10] 전체 row 수:", f"{len(full_ranking_855plus10_df):,}")
print("[855only] JD 수:", full_ranking_855only_df["jd_id"].nunique())
print("[855only] JD별 후보 수 min/max:", int(only_counts.min()), "/", int(only_counts.max()))
print("[855only] 전체 row 수:", f"{len(full_ranking_855only_df):,}")

if int(plus10_counts.max()) < 865:
    print(
        "주의: 855+10 후보 수가 865보다 작습니다. "
        "TOP_N_RETRIEVAL 또는 후보 생성 단계에서 pool이 제한되었을 수 있습니다."
    )
if int(only_counts.max()) < 855:
    print(
        "주의: 855only 후보 수가 855보다 작습니다. "
        "실제 CV 후보가 일부 제외되었거나 candidate_type 판정/embedding alignment를 확인해야 합니다."
    )

display_cols = [
    "jd_id", "candidate_id", "candidate_type", "rank", "final_score", "final_score_100"
]
display_cols = [c for c in display_cols if c in full_ranking_855plus10_df.columns]
print("[855+10] JD별 Top 3 미리보기")
display(full_ranking_855plus10_df.groupby("jd_id").head(3)[display_cols])
print("[855only] JD별 Top 3 미리보기")
display(full_ranking_855only_df.groupby("jd_id").head(3)[display_cols])


# In[43]:


# Cell 42 | latest_matching_result_all_semantics 생성
# ------------------------------------------------------------
# 목적:
# - latest_matching_result_all과 동일한 매칭/점수 체계를 사용합니다.
# - 단, Gold/가상 CV는 제외합니다.
# - JD별 상위 N 제한 없이 JD별 전원 후보에 대해 맥락 해석을 작성합니다.
# - full_ranking_855only_df가 있으면 이를 최우선으로 사용합니다.
# - full_ranking_855only_df가 없으면 full_ranking_855plus10_df 또는 latest_matching_result_all.csv를 fallback으로 사용합니다.
# ------------------------------------------------------------

from datetime import datetime
import numpy as np
import pandas as pd
import re

SEMANTICS_CREATED_AT = datetime.now().isoformat(timespec="seconds")
SEMANTICS_OUTPUT_CSV_LATEST = MATCHING_OUTPUT_DIR / f"{SEMANTICS_OUTPUT_BASENAME}.csv"
SEMANTICS_OUTPUT_XLSX_LATEST = MATCHING_OUTPUT_DIR / f"{SEMANTICS_OUTPUT_BASENAME}.xlsx"


def _read_latest_matching_result_all_fallback():
    """운영 경로 또는 임시 업로드 경로에서 latest_matching_result_all CSV를 찾습니다."""
    candidates = [
        MATCHING_OUTPUT_DIR / "latest_matching_result_all.csv",
        Path("/mnt/data/latest_matching_result_all.csv"),
        Path("/mnt/data/latest_matching_result_all_5.csv"),
    ]
    sandbox_matches = sorted(Path("/mnt/data").glob("latest_matching_result_all*.csv")) if Path("/mnt/data").exists() else []
    candidates.extend(sandbox_matches)

    for path in candidates:
        path = Path(path)
        if path.exists():
            return pd.read_csv(path), str(path)
    raise FileNotFoundError(
        "full_ranking_855only_df/full_ranking_855plus10_df가 없고 latest_matching_result_all CSV도 찾지 못했습니다."
    )


def _gold_cv_mask(df):
    """Gold/가상 CV 후보를 보수적으로 식별합니다."""
    mask = pd.Series(False, index=df.index)

    if "candidate_type" in df.columns:
        candidate_type = df["candidate_type"].fillna("").astype(str).str.lower()
        mask |= candidate_type.isin(["virtual_cv", "gold_cv", "gold", "virtual"])

    if "candidate_id" in df.columns:
        candidate_id = df["candidate_id"].fillna("").astype(str)
        mask |= candidate_id.str.startswith("G_JD_")
        mask |= candidate_id.str.contains(r"\bgold\b|virtual|가상", case=False, regex=True, na=False)

    text_cols = [
        col for col in ["cv_file_name", "file_name", "resume_no", "cv_job_role", "candidate_name"]
        if col in df.columns
    ]
    for col in text_cols:
        text = df[col].fillna("").astype(str)
        mask |= text.str.contains(r"\bgold\b|virtual cv|가상\s*cv", case=False, regex=True, na=False)

    return mask


def _split_terms_for_semantics(value):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return []
    return [term.strip() for term in str(value).split(",") if term.strip()]


def _overlap_for_semantics(row, jd_col, cv_col):
    jd_terms = set(_split_terms_for_semantics(row.get(jd_col, "")))
    cv_terms = set(_split_terms_for_semantics(row.get(cv_col, "")))
    return sorted(jd_terms.intersection(cv_terms))


def _missing_for_semantics(row, jd_col, cv_col):
    jd_terms = set(_split_terms_for_semantics(row.get(jd_col, "")))
    cv_terms = set(_split_terms_for_semantics(row.get(cv_col, "")))
    return sorted(jd_terms - cv_terms)


def _clip_text(value, limit=260):
    text = "" if value is None or (isinstance(value, float) and pd.isna(value)) else str(value)
    text = re.sub(r"\s+", " ", text).strip()
    return text[:limit]


def _score_100_value(row, col):
    if col in row.index and pd.notna(row.get(col)):
        return round(float(row.get(col)) * 100, 1)
    col100 = f"{col}_100"
    if col100 in row.index and pd.notna(row.get(col100)):
        return round(float(row.get(col100)), 1)
    return None


def build_semantic_strength_context(row):
    parts = []
    skill_overlap = _overlap_for_semantics(row, "jd_skill_standard", "cv_skill_standard")
    task_overlap = _overlap_for_semantics(row, "jd_task_standard", "cv_task_standard")
    domain_overlap = _overlap_for_semantics(row, "jd_domain_standard", "cv_domain_standard")
    role_overlap = _overlap_for_semantics(row, "jd_role_standard", "cv_role_standard")
    ai_overlap = _overlap_for_semantics(row, "jd_ai_skill_tags", "cv_ai_skill_tags")

    if skill_overlap:
        parts.append("Skill 직접 일치: " + ", ".join(skill_overlap[:12]))
    if ai_overlap:
        parts.append("AI 핵심 태그 일치: " + ", ".join(ai_overlap[:12]))
    if task_overlap:
        parts.append("Task 맥락 일치: " + ", ".join(task_overlap[:12]))
    if domain_overlap:
        parts.append("Domain 맥락 일치: " + ", ".join(domain_overlap[:8]))
    if role_overlap:
        parts.append("Role 신호 일치: " + ", ".join(role_overlap[:8]))

    career_evidence = _clip_text(row.get("cv_career_description", ""), 220)
    if career_evidence:
        parts.append("CV 경력 근거: " + career_evidence)

    if not parts and "match_reason_rule_based" in row.index and pd.notna(row.get("match_reason_rule_based")):
        parts.append(_clip_text(row.get("match_reason_rule_based"), 500))

    return "\n".join(parts) if parts else "명시적 강점 근거가 제한적이므로 CV 원문 확인이 필요합니다."


def build_semantic_gap_context(row):
    parts = []
    missing_skill = _missing_for_semantics(row, "jd_skill_standard", "cv_skill_standard")
    missing_ai = _missing_for_semantics(row, "jd_ai_skill_tags", "cv_ai_skill_tags")

    if missing_skill:
        parts.append("JD 요구 스킬 중 CV에서 약하게 보이는 항목: " + ", ".join(missing_skill[:12]))
    if missing_ai:
        parts.append("AI 핵심 태그 중 추가 확인 필요: " + ", ".join(missing_ai[:12]))

    domain_score = _score_100_value(row, "domain_score")
    role_score = _score_100_value(row, "role_score")
    task_score = _score_100_value(row, "task_score")

    if domain_score is not None and domain_score < 50:
        parts.append(f"Domain score {domain_score}점으로 산업/서비스 맥락 적합성 확인 필요")
    if role_score is not None and role_score < 50:
        parts.append(f"Role score {role_score}점으로 직무 역할 일치도 확인 필요")
    if task_score is not None and task_score < 50:
        parts.append(f"Task score {task_score}점으로 실제 수행 과제 유사성 확인 필요")

    if not parts and "match_gap_rule_based" in row.index and pd.notna(row.get("match_gap_rule_based")):
        gap_text = _clip_text(row.get("match_gap_rule_based"), 500)
        if gap_text and gap_text != "큰 Gap은 규칙 기반에서 확인되지 않음":
            parts.append(gap_text)

    return "\n".join(parts) if parts else "규칙 기반 기준에서 큰 Gap은 확인되지 않았습니다."


def build_semantic_fit_level(row):
    final_score_100 = _score_100_value(row, "final_score")
    if final_score_100 is None:
        return "판정 보류"
    if final_score_100 >= 75:
        return "높음"
    if final_score_100 >= 60:
        return "중상"
    if final_score_100 >= 45:
        return "보통"
    return "낮음"


def build_semantic_interview_focus(row):
    if "interview_question_rule_based" in row.index and pd.notna(row.get("interview_question_rule_based")):
        existing = _clip_text(row.get("interview_question_rule_based"), 700)
        if existing:
            return existing

    questions = []
    if _score_100_value(row, "skill_score") is not None and _score_100_value(row, "skill_score") < 60:
        questions.append("JD 핵심 스킬을 실제 프로젝트에서 어떤 산출물로 구현했는지 확인")
    if _score_100_value(row, "task_score") is not None and _score_100_value(row, "task_score") < 60:
        questions.append("JD 주요업무와 유사한 과제 수행 경험 및 본인 기여도 확인")
    if _score_100_value(row, "domain_score") is not None and _score_100_value(row, "domain_score") < 60:
        questions.append("해당 산업/서비스 도메인 이해도와 전환 가능성 확인")
    if not questions:
        questions.append("가장 유사한 프로젝트에서 본인 역할, 사용 기술, 정량/정성 성과 확인")
    return "\n".join([f"- {q}" for q in questions])


def build_semantic_context_interpretation(row):
    jd_title = _clip_text(row.get("jd_title", row.get("job_title", "")), 80)
    candidate_id = _clip_text(row.get("candidate_id", ""), 80)
    final_score_100 = _score_100_value(row, "final_score")
    skill_score_100 = _score_100_value(row, "skill_score")
    task_score_100 = _score_100_value(row, "task_score")
    domain_score_100 = _score_100_value(row, "domain_score")
    role_score_100 = _score_100_value(row, "role_score")

    score_parts = []
    for label, score in [
        ("Final", final_score_100),
        ("Skill", skill_score_100),
        ("Task", task_score_100),
        ("Domain", domain_score_100),
        ("Role", role_score_100),
    ]:
        if score is not None:
            score_parts.append(f"{label} {score}점")

    score_sentence = ", ".join(score_parts) if score_parts else "점수 정보 제한"
    strength = build_semantic_strength_context(row)
    gap = build_semantic_gap_context(row)

    return (
        f"[{build_semantic_fit_level(row)}] JD '{jd_title}' 기준 후보 '{candidate_id}'의 매칭 맥락입니다. "
        f"{score_sentence}.\n"
        f"강점 맥락:\n{strength}\n"
        f"확인 필요 맥락:\n{gap}"
    )


# 1) semantics source 결정: JD별 전원 후보를 위해 full_ranking_855only_df를 최우선 사용합니다.
if "full_ranking_855only_df" in globals() and isinstance(full_ranking_855only_df, pd.DataFrame) and not full_ranking_855only_df.empty:
    semantics_source_df = full_ranking_855only_df.copy()
    semantics_source_name = "full_ranking_855only_df"
elif "full_ranking_855plus10_df" in globals() and isinstance(full_ranking_855plus10_df, pd.DataFrame) and not full_ranking_855plus10_df.empty:
    semantics_source_df = full_ranking_855plus10_df.copy()
    semantics_source_name = "full_ranking_855plus10_df"
else:
    semantics_source_df, semantics_source_name = _read_latest_matching_result_all_fallback()

semantics_source_df["jd_id"] = semantics_source_df["jd_id"].astype(str)
semantics_source_df["candidate_id"] = semantics_source_df["candidate_id"].astype(str)

source_row_count = len(semantics_source_df)
source_jd_count = semantics_source_df["jd_id"].nunique()
source_pair_count = semantics_source_df[["jd_id", "candidate_id"]].drop_duplicates().shape[0]

# 2) Gold/가상 CV 제외
semantics_source_df["is_gold_cv_excluded"] = _gold_cv_mask(semantics_source_df)
excluded_gold_count = int(semantics_source_df["is_gold_cv_excluded"].sum())
semantics_df = semantics_source_df[~semantics_source_df["is_gold_cv_excluded"]].copy()

# 3) JD-candidate 중복 제거 및 JD별 전원 재랭킹
sort_cols = ["jd_id"]
ascending = [True]
if "final_score" in semantics_df.columns:
    semantics_df["final_score"] = pd.to_numeric(semantics_df["final_score"], errors="coerce")
    sort_cols.append("final_score")
    ascending.append(False)
elif "rank" in semantics_df.columns:
    semantics_df["rank"] = pd.to_numeric(semantics_df["rank"], errors="coerce")
    sort_cols.append("rank")
    ascending.append(True)

semantics_df = semantics_df.sort_values(sort_cols, ascending=ascending).drop_duplicates(["jd_id", "candidate_id"], keep="first").reset_index(drop=True)
semantics_df["rank"] = semantics_df.groupby("jd_id").cumcount() + 1

# 4) 맥락 해석 컬럼 생성
semantics_df["semantic_fit_level"] = semantics_df.apply(build_semantic_fit_level, axis=1)
semantics_df["semantic_strength_context"] = semantics_df.apply(build_semantic_strength_context, axis=1)
semantics_df["semantic_gap_context"] = semantics_df.apply(build_semantic_gap_context, axis=1)
semantics_df["semantic_interview_focus"] = semantics_df.apply(build_semantic_interview_focus, axis=1)
semantics_df["semantic_context_interpretation"] = semantics_df.apply(build_semantic_context_interpretation, axis=1)
semantics_df["semantics_source"] = semantics_source_name
semantics_df["semantics_created_at"] = SEMANTICS_CREATED_AT

# 5) 컬럼 순서 정리
semantics_front_cols = [
    "jd_id", "jd_title", "jd_company", "candidate_id", "cv_file_name", "rank",
    "final_score", "final_score_100", "semantic_fit_level",
    "semantic_context_interpretation", "semantic_strength_context", "semantic_gap_context",
    "semantic_interview_focus", "semantics_source", "semantics_created_at",
]
semantics_front_cols = [col for col in semantics_front_cols if col in semantics_df.columns]
semantics_rest_cols = [col for col in semantics_df.columns if col not in semantics_front_cols]
semantics_df = semantics_df[semantics_front_cols + semantics_rest_cols].copy()

# 6) 누락 검토 audit
semantics_source_counts = semantics_source_df.groupby("jd_id")["candidate_id"].nunique().rename("source_unique_candidates")
semantics_gold_counts = semantics_source_df[semantics_source_df["is_gold_cv_excluded"]].groupby("jd_id")["candidate_id"].nunique().rename("excluded_gold_candidates")
semantics_final_counts = semantics_df.groupby("jd_id")["candidate_id"].nunique().rename("semantics_unique_candidates")
semantics_audit_df = pd.concat([semantics_source_counts, semantics_gold_counts, semantics_final_counts], axis=1).fillna(0).reset_index()
semantics_audit_df[["source_unique_candidates", "excluded_gold_candidates", "semantics_unique_candidates"]] = semantics_audit_df[["source_unique_candidates", "excluded_gold_candidates", "semantics_unique_candidates"]].astype(int)
semantics_audit_df["expected_after_gold_exclusion"] = semantics_audit_df["source_unique_candidates"] - semantics_audit_df["excluded_gold_candidates"]
semantics_audit_df["missing_after_semantics"] = semantics_audit_df["expected_after_gold_exclusion"] - semantics_audit_df["semantics_unique_candidates"]

missing_rows = semantics_df["semantic_context_interpretation"].fillna("").astype(str).str.len().eq(0).sum()
if int(semantics_audit_df["missing_after_semantics"].sum()) != 0:
    raise ValueError("semantics 결과에 누락이 있습니다. semantics_audit_df를 확인하세요.")
if missing_rows:
    raise ValueError(f"semantic_context_interpretation 빈 값이 있습니다: {missing_rows}건")

# 7) 저장
semantics_df.to_csv(SEMANTICS_OUTPUT_CSV_LATEST, index=False, encoding="utf-8-sig")
semantics_df.to_excel(SEMANTICS_OUTPUT_XLSX_LATEST, index=False)

print("latest_matching_result_all_semantics 생성 완료")
print("source:", semantics_source_name)
print("source rows:", f"{source_row_count:,}")
print("source JD count:", source_jd_count)
print("source JD-candidate pairs:", f"{source_pair_count:,}")
print("excluded Gold/virtual CV rows:", f"{excluded_gold_count:,}")
print("semantics rows:", f"{len(semantics_df):,}")
print("semantics CSV:", SEMANTICS_OUTPUT_CSV_LATEST)
print("semantics XLSX:", SEMANTICS_OUTPUT_XLSX_LATEST)
display(semantics_audit_df)
display(semantics_df.head(10))


# In[44]:


# Cell 43 | 채용공고별 결과 저장
# ------------------------------------------------------------
# 산출물 정리 정책:
# - 매 실행마다 timestamp 파일을 새로 만들지 않습니다.
# - 전체 결과와 JD별 결과는 latest 고정 파일만 갱신합니다.
# - 실행 이력은 작은 index CSV 하나만 유지합니다.
#
# 저장 구조:
# data/08_matching/latest_matching_result_all.csv
# data/08_matching/latest_matching_result_all.xlsx
# data/08_matching/latest_matching_result_855plus10_v5.csv
# data/08_matching/latest_matching_result_855plus10_v5.xlsx
# data/08_matching/latest_matching_result_855only_v5.csv
# data/08_matching/latest_matching_result_855only_v5.xlsx
# data/09_dashboard/latest_dashboard_summary_all.csv
# data/09_dashboard/latest_dashboard_summary_all.xlsx
# data/08_matching/by_jd/jd_191/latest_matching_result_jd_191.csv
# data/08_matching/by_jd/jd_191/latest_matching_result_jd_191.xlsx
# data/09_dashboard/by_jd/jd_191/latest_dashboard_summary_jd_191.csv
# data/09_dashboard/by_jd/jd_191/latest_dashboard_summary_jd_191.xlsx
# ------------------------------------------------------------

from datetime import datetime
import json

RUN_CREATED_AT = datetime.now().isoformat(timespec="seconds")

MATCHING_BY_JD_DIR = MATCHING_OUTPUT_DIR / "by_jd"
DASHBOARD_BY_JD_DIR = DASHBOARD_OUTPUT_DIR / "by_jd"
MATCHING_BY_JD_DIR.mkdir(parents=True, exist_ok=True)
DASHBOARD_BY_JD_DIR.mkdir(parents=True, exist_ok=True)

run_index_rows = []

# 전체 결과는 latest 고정 파일만 갱신합니다.
OUTPUT_MATCHING_CSV_LATEST = MATCHING_OUTPUT_DIR / "latest_matching_result_all.csv"
OUTPUT_MATCHING_XLSX_LATEST = MATCHING_OUTPUT_DIR / "latest_matching_result_all.xlsx"
OUTPUT_DASHBOARD_SUMMARY_CSV_LATEST = DASHBOARD_OUTPUT_DIR / "latest_dashboard_summary_all.csv"
OUTPUT_DASHBOARD_SUMMARY_XLSX_LATEST = DASHBOARD_OUTPUT_DIR / "latest_dashboard_summary_all.xlsx"
OUTPUT_FULL_RANKING_855PLUS10_CSV_LATEST = MATCHING_OUTPUT_DIR / f"{FULL_RANKING_PLUS10_OUTPUT_BASENAME}.csv"
OUTPUT_FULL_RANKING_855PLUS10_XLSX_LATEST = MATCHING_OUTPUT_DIR / f"{FULL_RANKING_PLUS10_OUTPUT_BASENAME}.xlsx"
OUTPUT_FULL_RANKING_855ONLY_CSV_LATEST = MATCHING_OUTPUT_DIR / f"{FULL_RANKING_855ONLY_OUTPUT_BASENAME}.csv"
OUTPUT_FULL_RANKING_855ONLY_XLSX_LATEST = MATCHING_OUTPUT_DIR / f"{FULL_RANKING_855ONLY_OUTPUT_BASENAME}.xlsx"

# Cell 26이 실행되지 않았거나 result_df가 없는 경우를 방지합니다.
# full_ranking_855plus10_df에서 JD별 Top-N을 재구성하여 저장 흐름이 끊기지 않게 합니다.
if "result_df" not in globals() or not isinstance(result_df, pd.DataFrame):
    if "full_ranking_855plus10_df" in globals():
        print("주의: result_df가 없어 full_ranking_855plus10_df에서 Top-N result_df를 재생성합니다.")
        result_df = full_ranking_855plus10_df[
            pd.to_numeric(full_ranking_855plus10_df["rank"], errors="coerce") <= TOP_N_FINAL
        ].copy()
    else:
        raise NameError("result_df가 없고 full_ranking_855plus10_df도 없습니다. Cell 26 또는 Cell 41 실행 상태를 확인하세요.")

if "dashboard_summary_df" not in globals() or not isinstance(dashboard_summary_df, pd.DataFrame):
    print("주의: dashboard_summary_df가 없어 result_df 기반 최소 summary를 생성합니다.")
    dashboard_summary_df = result_df.copy()

result_df.to_csv(OUTPUT_MATCHING_CSV_LATEST, index=False, encoding="utf-8-sig")
result_df.to_excel(OUTPUT_MATCHING_XLSX_LATEST, index=False)
dashboard_summary_df.to_csv(OUTPUT_DASHBOARD_SUMMARY_CSV_LATEST, index=False, encoding="utf-8-sig")
dashboard_summary_df.to_excel(OUTPUT_DASHBOARD_SUMMARY_XLSX_LATEST, index=False)

if "full_ranking_855plus10_df" in globals():
    full_ranking_855plus10_df.to_csv(OUTPUT_FULL_RANKING_855PLUS10_CSV_LATEST, index=False, encoding="utf-8-sig")
    full_ranking_855plus10_df.to_excel(OUTPUT_FULL_RANKING_855PLUS10_XLSX_LATEST, index=False)
else:
    print("주의: full_ranking_855plus10_df가 없어 855+10 final_score 랭킹 저장을 건너뜁니다.")

if "full_ranking_855only_df" in globals():
    full_ranking_855only_df.to_csv(OUTPUT_FULL_RANKING_855ONLY_CSV_LATEST, index=False, encoding="utf-8-sig")
    full_ranking_855only_df.to_excel(OUTPUT_FULL_RANKING_855ONLY_XLSX_LATEST, index=False)
else:
    print("주의: full_ranking_855only_df가 없어 855only final_score 랭킹 저장을 건너뜁니다.")

for jd_id in sorted(result_df["jd_id"].astype(str).unique()):
    safe_jd_id = re.sub(r"[^0-9A-Za-z가-힣_-]+", "_", str(jd_id))
    jd_result_df = result_df[result_df["jd_id"].astype(str) == str(jd_id)].copy()
    jd_summary_df = dashboard_summary_df[dashboard_summary_df["jd_id"].astype(str) == str(jd_id)].copy()

    jd_match_dir = MATCHING_BY_JD_DIR / f"jd_{safe_jd_id}"
    jd_dash_dir = DASHBOARD_BY_JD_DIR / f"jd_{safe_jd_id}"
    jd_match_dir.mkdir(parents=True, exist_ok=True)
    jd_dash_dir.mkdir(parents=True, exist_ok=True)

    latest_result_csv = jd_match_dir / f"latest_matching_result_jd_{safe_jd_id}.csv"
    latest_result_xlsx = jd_match_dir / f"latest_matching_result_jd_{safe_jd_id}.xlsx"
    latest_summary_csv = jd_dash_dir / f"latest_dashboard_summary_jd_{safe_jd_id}.csv"
    latest_summary_xlsx = jd_dash_dir / f"latest_dashboard_summary_jd_{safe_jd_id}.xlsx"

    jd_result_df.to_csv(latest_result_csv, index=False, encoding="utf-8-sig")
    jd_result_df.to_excel(latest_result_xlsx, index=False)
    jd_summary_df.to_csv(latest_summary_csv, index=False, encoding="utf-8-sig")
    jd_summary_df.to_excel(latest_summary_xlsx, index=False)

    jd_title = ""
    jd_company = ""
    if not jd_summary_df.empty:
        jd_title = str(jd_summary_df["jd_title"].iloc[0]) if "jd_title" in jd_summary_df.columns else ""
        jd_company = str(jd_summary_df["jd_company"].iloc[0]) if "jd_company" in jd_summary_df.columns else ""

    run_index_rows.append({
        "created_at": RUN_CREATED_AT,
        "jd_id": jd_id,
        "jd_company": jd_company,
        "jd_title": jd_title,
        "n_candidates": len(jd_result_df),
        "latest_result_csv": str(latest_result_csv),
        "latest_result_xlsx": str(latest_result_xlsx),
        "latest_summary_csv": str(latest_summary_csv),
        "latest_summary_xlsx": str(latest_summary_xlsx),
    })

run_index_new = pd.DataFrame(run_index_rows)
run_index_path = MATCHING_OUTPUT_DIR / "latest_matching_run_index.csv"
run_index_new.to_csv(run_index_path, index=False, encoding="utf-8-sig")

print("저장 완료")
print("생성 정책: latest 고정 파일만 갱신")
print("전체 result:", OUTPUT_MATCHING_XLSX_LATEST)
print("전체 dashboard summary:", OUTPUT_DASHBOARD_SUMMARY_XLSX_LATEST)
if "full_ranking_855plus10_df" in globals():
    print("전체 후보 final_score ranking 855+10:", OUTPUT_FULL_RANKING_855PLUS10_XLSX_LATEST)
if "full_ranking_855only_df" in globals():
    print("실제 CV only final_score ranking 855only:", OUTPUT_FULL_RANKING_855ONLY_XLSX_LATEST)
if "semantics_df" in globals():
    print("Gold CV 제외 JD별 전원 semantics:", SEMANTICS_OUTPUT_XLSX_LATEST)
print("JD별 latest index:", run_index_path)
display(run_index_new)


# In[45]:


# Cell 44 | Pipeline 연결 검증
# ------------------------------------------------------------
# Taxonomy → Retrieval → 정밀 점수 → 가중치 → 대시보드 저장까지
# 실제 사용된 방법과 주요 컬럼 상태를 한 번에 점검합니다.
# ------------------------------------------------------------

validation_report = {
    "taxonomy_source": TAXONOMY_SOURCE,
    "taxonomy_terms": {
        "skill": len(SKILL_TAXONOMY),
        "task": len(TASK_TAXONOMY),
        "domain": len(DOMAIN_TAXONOMY),
        "role": len(ROLE_TAXONOMY),
    },
    "cv_standard_non_empty": {
        col: int(cv_df[col].fillna("").astype(str).str.len().gt(0).sum())
        for col in ["cv_skill_standard", "cv_task_standard", "cv_domain_standard", "cv_role_standard"]
    },
    "jd_standard_non_empty": {
        col: int(jd_df[col].fillna("").astype(str).str.len().gt(0).sum())
        for col in ["jd_skill_standard", "jd_task_standard", "jd_domain_standard", "jd_role_standard"]
    },
    "retrieval_method_counts": result_df["retrieval_method"].value_counts(dropna=False).to_dict(),
    "scoring_method_counts": result_df["scoring_method"].value_counts(dropna=False).to_dict(),
    "cross_encoder_requested": bool(globals().get("USE_CROSS_ENCODER_REQUESTED", globals().get("USE_CROSS_ENCODER", False))),
    "cross_encoder_used_in_cell16": bool(globals().get("USE_CROSS_ENCODER", False)),
    "retrieval_column_weights": globals().get("RETRIEVAL_COLUMN_WEIGHTS", None),
    "weights": {
        "retrieval_score_norm": RETRIEVAL_SCORE_WEIGHT,
        "column_score": COLUMN_SCORE_WEIGHT,
        **WEIGHTS,
    },
    "radar_columns_available": [
        col for col in [
            "radar_skill_score_100",
            "radar_task_score_100",
            "radar_domain_score_100",
            "radar_role_score_100",
            "radar_experience_score_100",
            "radar_soft_skill_score_100",
        ] if col in dashboard_summary_df.columns
    ],
    "evaluation_summary": globals().get("EVALUATION_SUMMARY_DF", pd.DataFrame()).to_dict("records") if "EVALUATION_SUMMARY_DF" in globals() else None,
    "weight_config_validation_top": globals().get("WEIGHT_CONFIG_VALIDATION_DF", pd.DataFrame()).head(10).to_dict("records") if "WEIGHT_CONFIG_VALIDATION_DF" in globals() else None,
    "result_shape": result_df.shape,
    "dashboard_shape": dashboard_summary_df.shape,
    "full_universe_ranking_shape": full_universe_ranking_df.shape if "full_universe_ranking_df" in globals() else None,
    "full_universe_ranking_file": str(MATCHING_OUTPUT_DIR / f"{FULL_RANKING_PLUS10_OUTPUT_BASENAME}.xlsx"),
    "semantics_output_file": str(MATCHING_OUTPUT_DIR / f"{SEMANTICS_OUTPUT_BASENAME}.xlsx") if "SEMANTICS_OUTPUT_BASENAME" in globals() else None,
    "semantics_shape": semantics_df.shape if "semantics_df" in globals() else None,
    "semantics_audit_total_missing": int(semantics_audit_df["missing_after_semantics"].sum()) if "semantics_audit_df" in globals() else None,
}

print(json.dumps(validation_report, ensure_ascii=False, indent=2))





# In[46]:


# Cell 45 | 완료 메시지
print("09_matching_universe_v5 완료")
print("결과는 채용공고별로 latest 고정 파일에 저장됩니다.")
print("- JD별 Top 10: data/08_matching/latest_matching_result_all.xlsx")
print("- 대시보드 summary: data/09_dashboard/latest_dashboard_summary_all.xlsx")
print("- JD별 전체 후보 final_score ranking 855+10: data/08_matching/latest_matching_result_855plus10_v5.xlsx")
print("- JD별 실제 CV only final_score ranking 855only: data/08_matching/latest_matching_result_855only_v5.xlsx")
print("- Gold CV 제외 JD별 전원 semantics: data/08_matching/latest_matching_result_all_semantics.xlsx")
print("- 대시보드는 02_dashboard_builder.ipynb에서 생성하세요.")


# In[48]:


# Cell 46 | latest_matching_result_all_semantics2 생성
# 목적:
# - latest_matching_result_all 파일에서 Gold CV를 제외
# - Gold CV 제외 후 JD별 기존 순서/점수 기준 상위 10명 유지
# - 컬럼 CG 이후를 포함한 모든 원본 컬럼 보존
# - latest_matching_result_all_semantics2.csv / audit.csv 생성

from pathlib import Path
import pandas as pd
import numpy as np

# ------------------------------------------------------------
# 1. 입력/출력 경로 설정
# ------------------------------------------------------------
BASE_DIR = Path(".")
INPUT_CANDIDATES = [
    BASE_DIR / "latest_matching_result_all.csv",
    BASE_DIR / "latest_matching_result_all(6).csv",
    BASE_DIR / "latest_matching_result_all(5).csv",
    Path("/mnt/data/latest_matching_result_all.csv"),
    Path("/mnt/data/latest_matching_result_all(6).csv"),
    Path("/mnt/data/latest_matching_result_all(5).csv"),
]

input_path = next((p for p in INPUT_CANDIDATES if p.exists()), None)

if input_path is None:
    raise FileNotFoundError(
        "latest_matching_result_all 입력 파일을 찾지 못했습니다. "
        "현재 폴더 또는 /mnt/data에 latest_matching_result_all.csv 파일을 두고 다시 실행하세요."
    )

output_csv_path = BASE_DIR / "latest_matching_result_all_semantics2.csv"
audit_csv_path = BASE_DIR / "latest_matching_result_all_semantics2_audit.csv"

print(f"[INFO] 입력 파일: {input_path}")
print(f"[INFO] 출력 파일: {output_csv_path}")
print(f"[INFO] Audit 파일: {audit_csv_path}")

# ------------------------------------------------------------
# 2. 파일 읽기
# ------------------------------------------------------------
df = pd.read_csv(input_path, encoding="utf-8-sig")
original_columns = list(df.columns)

print(f"[INFO] 입력 행 수: {len(df):,}")
print(f"[INFO] 입력 컬럼 수: {len(original_columns):,}")

# Excel 컬럼 CG는 1-based 85번째, Python index 84
cg_index = 84
if len(original_columns) > cg_index:
    cg_column = original_columns[cg_index]
    print(f"[INFO] Excel 기준 CG 컬럼명: {cg_column}")
    print(f"[INFO] CG 이후 포함 컬럼 수: {len(original_columns[cg_index:]):,}")
else:
    cg_column = None
    print("[WARN] 입력 파일의 컬럼 수가 CG 위치보다 적습니다. 그래도 모든 컬럼을 보존해 출력합니다.")

# ------------------------------------------------------------
# 3. 주요 컬럼 자동 탐지
# ------------------------------------------------------------
def find_col(columns, candidates, contains_any=None):
    """
    candidates: 정확히 일치 우선 후보
    contains_any: 소문자 기준 부분 문자열 후보
    """
    col_set = set(columns)

    for c in candidates:
        if c in col_set:
            return c

    lowered = {str(c).lower(): c for c in columns}

    for c in candidates:
        key = str(c).lower()
        if key in lowered:
            return lowered[key]

    if contains_any:
        for col in columns:
            col_l = str(col).lower()
            if any(token.lower() in col_l for token in contains_any):
                return col

    return None


jd_col = find_col(
    original_columns,
    candidates=[
        "jd_id", "JD_ID", "jd", "JD", "job_id", "Job_ID",
        "jd_no", "JD_NO", "공고ID", "채용공고ID"
    ],
    contains_any=["jd_id", "job_id", "jd"]
)

rank_col = find_col(
    original_columns,
    candidates=[
        "rank", "Rank", "ranking", "Ranking", "match_rank",
        "top_rank", "순위", "랭킹"
    ],
    contains_any=["rank", "순위"]
)

score_col = find_col(
    original_columns,
    candidates=[
        "score", "Score", "final_score", "matching_score",
        "match_score", "total_score", "similarity", "cosine_similarity"
    ],
    contains_any=["score", "similarity"]
)

candidate_id_col = find_col(
    original_columns,
    candidates=[
        "candidate_id", "Candidate_ID", "cv_id", "CV_ID",
        "resume_id", "Resume_ID", "applicant_id", "Applicant_ID",
        "id", "ID"
    ],
    contains_any=["candidate", "cv_id", "resume", "applicant"]
)

candidate_type_col = find_col(
    original_columns,
    candidates=[
        "candidate_type", "Candidate_Type", "cv_type", "CV_Type",
        "type", "Type"
    ],
    contains_any=["candidate_type", "cv_type"]
)

print("[INFO] 자동 탐지 컬럼")
print(f"  - JD 컬럼: {jd_col}")
print(f"  - Rank 컬럼: {rank_col}")
print(f"  - Score 컬럼: {score_col}")
print(f"  - Candidate ID 컬럼: {candidate_id_col}")
print(f"  - Candidate Type 컬럼: {candidate_type_col}")

if jd_col is None:
    raise ValueError(
        "JD 식별 컬럼을 찾지 못했습니다. "
        "jd_id, JD_ID, jd, job_id 등 JD 식별 컬럼명을 확인하세요."
    )

# ------------------------------------------------------------
# 4. Gold CV 제외 마스크 생성
# ------------------------------------------------------------
gold_mask = pd.Series(False, index=df.index)

# candidate_type 기반 제외
if candidate_type_col is not None:
    candidate_type_s = df[candidate_type_col].astype(str).str.lower()
    gold_mask |= candidate_type_s.str.contains(
        "gold|virtual|가상|gold_cv|virtual_cv",
        regex=True,
        na=False
    )

# candidate_id 기반 제외
if candidate_id_col is not None:
    candidate_id_s = df[candidate_id_col].astype(str)
    gold_mask |= candidate_id_s.str.startswith("G_JD_", na=False)
    gold_mask |= candidate_id_s.str.contains(
        r"(^|[^A-Za-z0-9])G_JD_|gold|virtual|가상",
        regex=True,
        case=False,
        na=False
    )

# 모든 object 컬럼에서 보조 탐지
object_cols = df.select_dtypes(include=["object"]).columns.tolist()
for col in object_cols:
    s = df[col].astype(str)
    gold_mask |= s.str.startswith("G_JD_", na=False)
    gold_mask |= s.str.contains(
        r"(^|[^A-Za-z0-9])G_JD_|gold cv|gold_cv|virtual cv|virtual_cv|가상 cv|가상_cv",
        regex=True,
        case=False,
        na=False
    )

df_no_gold = df.loc[~gold_mask].copy()

print(f"[INFO] Gold CV 제외 건수: {int(gold_mask.sum()):,}")
print(f"[INFO] Gold CV 제외 후 행 수: {len(df_no_gold):,}")

# ------------------------------------------------------------
# 5. JD별 상위 10명 선별
# ------------------------------------------------------------
# 원칙:
# - rank 컬럼이 있으면 rank 오름차순
# - rank가 없고 score 컬럼이 있으면 score 내림차순
# - 둘 다 없으면 원본 행 순서 유지
df_no_gold["_original_order_for_semantics2"] = np.arange(len(df_no_gold))

sort_cols = [jd_col]
ascending = [True]

if rank_col is not None:
    sort_cols.append(rank_col)
    ascending.append(True)
elif score_col is not None:
    sort_cols.append(score_col)
    ascending.append(False)

sort_cols.append("_original_order_for_semantics2")
ascending.append(True)

df_sorted = df_no_gold.sort_values(sort_cols, ascending=ascending).copy()

semantics2_df = (
    df_sorted
    .groupby(jd_col, group_keys=False, dropna=False)
    .head(10)
    .copy()
)

# 보조 컬럼 제거 후 원본 컬럼 순서 100% 보존
semantics2_df = semantics2_df.drop(columns=["_original_order_for_semantics2"], errors="ignore")
semantics2_df = semantics2_df[original_columns]

# ------------------------------------------------------------
# 6. Audit 생성
# ------------------------------------------------------------
input_count_by_jd = df.groupby(jd_col, dropna=False).size().rename("input_count")
gold_count_by_jd = df.loc[gold_mask].groupby(jd_col, dropna=False).size().rename("gold_excluded_count")
no_gold_count_by_jd = df_no_gold.groupby(jd_col, dropna=False).size().rename("after_gold_exclusion_count")
output_count_by_jd = semantics2_df.groupby(jd_col, dropna=False).size().rename("output_count")

audit_df = (
    pd.concat(
        [input_count_by_jd, gold_count_by_jd, no_gold_count_by_jd, output_count_by_jd],
        axis=1
    )
    .fillna(0)
    .reset_index()
)

for col in ["input_count", "gold_excluded_count", "after_gold_exclusion_count", "output_count"]:
    audit_df[col] = audit_df[col].astype(int)

audit_df["expected_output_count"] = audit_df["after_gold_exclusion_count"].clip(upper=10)
audit_df["audit_status"] = np.where(
    audit_df["output_count"] == audit_df["expected_output_count"],
    "OK",
    "CHECK"
)

# ------------------------------------------------------------
# 7. 저장
# ------------------------------------------------------------
semantics2_df.to_csv(output_csv_path, index=False, encoding="utf-8-sig")
audit_df.to_csv(audit_csv_path, index=False, encoding="utf-8-sig")

print("\n[DONE] latest_matching_result_all_semantics2 생성 완료")
print(f"  - 결과 CSV: {output_csv_path}")
print(f"  - Audit CSV: {audit_csv_path}")
print(f"  - 최종 행 수: {len(semantics2_df):,}")
print(f"  - 최종 컬럼 수: {len(semantics2_df.columns):,}")
print(f"  - 원본 컬럼 전체 보존: {list(semantics2_df.columns) == original_columns}")

if cg_column is not None:
    cg_after_missing = [c for c in original_columns[cg_index:] if c not in semantics2_df.columns]
    print(f"  - CG 이후 컬럼 누락 수: {len(cg_after_missing):,}")
    if cg_after_missing:
        print(f"  - CG 이후 누락 컬럼: {cg_after_missing}")

remaining_gold_like = 0
if candidate_id_col is not None:
    remaining_gold_like += semantics2_df[candidate_id_col].astype(str).str.contains(
        r"^G_JD_|gold|virtual|가상",
        regex=True,
        case=False,
        na=False
    ).sum()

print(f"  - 결과 내 Gold 의심 후보 수: {int(remaining_gold_like):,}")

print("\n[JD별 Audit]")
display(audit_df)

# In[57]:


# Cell 51 COMPLETE FIX FINAL | latest_matching_result_all_semantics7 생성
# 목적:
# - latest_matching_result_all_semantics6 결과의 후보/순위/전체 컬럼 양식 유지
# - CG 컬럼 match_reason_rule_based 이후 내용 채우기
# - 기존 semantics2/latest_matching_result_all에 있는 CG 이후 내용은 먼저 병합
# - 그래도 빈 값이면 rule-based 내용 생성
# - pandas LossySetitemError / float64 문자열 대입 오류 방지
# - llm_error는 전부 비어 있어도 정상 허용
# - 생성 파일: latest_matching_result_all_semantics7.xlsx / .csv

from pathlib import Path
import pandas as pd
import numpy as np
from IPython.display import display, FileLink

BASE_DIR = Path(".")

# ------------------------------------------------------------
# 1. 입력/출력 파일 설정
# ------------------------------------------------------------
target_candidates = [
    BASE_DIR / "latest_matching_result_all_semantics6.csv",
    BASE_DIR / "latest_matching_result_all_semantics6.xlsx",
    Path("/mnt/data/latest_matching_result_all_semantics6.csv"),
    Path("/mnt/data/latest_matching_result_all_semantics6.xlsx"),
]

semantic_source_candidates = [
    BASE_DIR / "latest_matching_result_all_semantics2.csv",
    BASE_DIR / "latest_matching_result_all_semantics2(2).csv",
    BASE_DIR / "latest_matching_result_all.csv",
    BASE_DIR / "latest_matching_result_all(6).csv",
    BASE_DIR / "latest_matching_result_all(5).csv",
    Path("/mnt/data/latest_matching_result_all_semantics2.csv"),
    Path("/mnt/data/latest_matching_result_all_semantics2(2).csv"),
    Path("/mnt/data/latest_matching_result_all.csv"),
    Path("/mnt/data/latest_matching_result_all(6).csv"),
    Path("/mnt/data/latest_matching_result_all(5).csv"),
]

target_path = next((p for p in target_candidates if p.exists()), None)
semantic_source_path = next((p for p in semantic_source_candidates if p.exists()), None)

if target_path is None:
    raise FileNotFoundError(
        "latest_matching_result_all_semantics6 파일을 찾지 못했습니다. "
        "먼저 semantics6 생성 셀을 실행하세요."
    )

if semantic_source_path is None:
    raise FileNotFoundError(
        "CG 이후 내용을 가져올 semantics2/latest_matching_result_all 파일을 찾지 못했습니다."
    )

output_excel_path = BASE_DIR / "latest_matching_result_all_semantics7.xlsx"
output_csv_path = BASE_DIR / "latest_matching_result_all_semantics7.csv"
audit_csv_path = BASE_DIR / "latest_matching_result_all_semantics7_cg_audit.csv"

print(f"[INFO] target 파일: {target_path}")
print(f"[INFO] CG 내용 source 파일: {semantic_source_path}")
print(f"[INFO] 출력 Excel 파일: {output_excel_path}")
print(f"[INFO] 출력 CSV 파일: {output_csv_path}")

# ------------------------------------------------------------
# 2. 파일 읽기
# ------------------------------------------------------------
def read_table(path):
    path = Path(path)
    if path.suffix.lower() == ".xlsx":
        return pd.read_excel(path)
    return pd.read_csv(path, encoding="utf-8-sig")

target_df = read_table(target_path)
source_df = read_table(semantic_source_path)

original_columns = list(target_df.columns)

print(f"[INFO] target 행 수: {len(target_df):,}")
print(f"[INFO] target 컬럼 수: {len(target_df.columns):,}")
print(f"[INFO] source 행 수: {len(source_df):,}")
print(f"[INFO] source 컬럼 수: {len(source_df.columns):,}")

# ------------------------------------------------------------
# 3. CG 이후 컬럼 확인
# ------------------------------------------------------------
cg_col = "match_reason_rule_based"

if cg_col not in target_df.columns:
    raise ValueError(f"target 파일에 CG 기준 컬럼 '{cg_col}'이 없습니다.")

cg_idx = target_df.columns.get_loc(cg_col)
cg_after_cols = list(target_df.columns[cg_idx:])

print(f"[INFO] CG 기준 컬럼: {cg_col}")
print(f"[INFO] CG 이후 컬럼 수: {len(cg_after_cols):,}")

print("\n[CG 이후 컬럼 목록]")
for i, col in enumerate(cg_after_cols, start=1):
    print(f"{i:02d}. {col}")

# ------------------------------------------------------------
# 4. key 컬럼 확인
# ------------------------------------------------------------
required_keys = ["jd_id", "candidate_id"]

for key in required_keys:
    if key not in target_df.columns:
        raise ValueError(f"target 파일에 key 컬럼이 없습니다: {key}")
    if key not in source_df.columns:
        raise ValueError(f"source 파일에 key 컬럼이 없습니다: {key}")

# ------------------------------------------------------------
# 5. CG 이후 컬럼 dtype 강제 object 변환
# ------------------------------------------------------------
for col in cg_after_cols:
    if col not in target_df.columns:
        target_df[col] = ""
    if col not in source_df.columns:
        source_df[col] = ""

    target_df[col] = target_df[col].astype("object")
    source_df[col] = source_df[col].astype("object")

target_df["_merge_jd_id"] = target_df["jd_id"].astype(str)
target_df["_merge_candidate_id"] = target_df["candidate_id"].astype(str)

source_df["_merge_jd_id"] = source_df["jd_id"].astype(str)
source_df["_merge_candidate_id"] = source_df["candidate_id"].astype(str)

source_semantic_df = (
    source_df[["_merge_jd_id", "_merge_candidate_id"] + cg_after_cols]
    .drop_duplicates(["_merge_jd_id", "_merge_candidate_id"], keep="first")
    .copy()
)

for col in cg_after_cols:
    source_semantic_df[col] = source_semantic_df[col].astype("object")

# ------------------------------------------------------------
# 6. 안전 함수
# ------------------------------------------------------------
def is_blank_value(x):
    if x is None:
        return True
    try:
        if pd.isna(x):
            return True
    except Exception:
        pass
    s = str(x).strip()
    return s == "" or s.lower() in ["nan", "none", "null", "<na>"]

def clean_cell_value(x):
    if x is None:
        return ""

    try:
        if pd.isna(x):
            return ""
    except Exception:
        pass

    if isinstance(x, (list, tuple, set)):
        out = []
        for v in x:
            try:
                if pd.isna(v):
                    continue
            except Exception:
                pass
            out.append(str(v))
        return " | ".join(out)

    if isinstance(x, dict):
        return str(x)

    if hasattr(x, "tolist") and not isinstance(x, (str, bytes)):
        try:
            values = x.tolist()
            if isinstance(values, list):
                out = []
                for v in values:
                    try:
                        if pd.isna(v):
                            continue
                    except Exception:
                        pass
                    out.append(str(v))
                return " | ".join(out)
            return str(values)
        except Exception:
            return str(x)

    return str(x)

def safe_value(row, col, default=""):
    if col in row.index and not is_blank_value(row[col]):
        return str(row[col]).strip()
    return default

def score_band(score):
    try:
        s = float(score)
    except Exception:
        return "점수 확인 필요"
    if s >= 0.85:
        return "매우 높은 적합도"
    if s >= 0.75:
        return "높은 적합도"
    if s >= 0.65:
        return "중간 이상 적합도"
    if s >= 0.55:
        return "보완 검토 필요"
    return "낮은 적합도"

def make_semantic_dict(row):
    jd_id = safe_value(row, "jd_id", "해당 JD")
    candidate_id = safe_value(row, "candidate_id", "해당 후보")
    rank = safe_value(row, "rank", "순위 미확인")
    final_score = safe_value(row, "final_score", "")
    score_desc = score_band(final_score)

    matched_keywords = safe_value(row, "matched_keywords", "")
    missing_keywords = safe_value(row, "missing_keywords", "")
    cv_summary = safe_value(row, "cv_summary", "")
    jd_summary = safe_value(row, "jd_summary", "")
    job_title = safe_value(row, "title", "")
    company = safe_value(row, "company", "")
    current_position = safe_value(row, "current_position", "")

    reason = (
        f"JD {jd_id} 기준 Gold CV 제외 후 재산정된 상위 10명 중 "
        f"후보 {candidate_id}는 {rank}위입니다. "
        f"final_score 기준 {score_desc}로 판단됩니다."
    )

    if matched_keywords:
        reason += f"\n매칭 근거: {matched_keywords}"

    if cv_summary:
        reason += f"\nCV 근거 일부: {cv_summary[:500]}"

    gap = (
        f"보완 또는 확인 필요 항목: {missing_keywords}"
        if missing_keywords
        else "명시적 미충족 항목은 제한적으로 확인되며, 실제 프로젝트 깊이와 역할 범위는 인터뷰 확인이 필요합니다."
    )

    questions = (
        "1. JD 핵심 요구역량과 직접 연결되는 프로젝트를 구체적으로 설명해 주세요.\n"
        "2. 본인이 직접 담당한 역할, 사용 기술, 산출물, 성과 지표를 설명해 주세요.\n"
        "3. 부족하거나 경험이 약한 영역을 어떻게 보완할 수 있는지 설명해 주세요."
    )

    career_path = []
    if current_position:
        career_path.append(f"현재/최근 포지션: {current_position}.")
    if company:
        career_path.append(f"관련 회사/조직: {company}.")
    if job_title:
        career_path.append(f"비교 JD/직무: {job_title}.")
    if jd_summary:
        career_path.append(f"JD 요약: {jd_summary[:300]}")
    if not career_path:
        career_path.append("경력 경로 정보가 제한적이므로 원본 CV 기반 확인이 필요합니다.")

    hard_summary = (
        f"하드스킬 근거: {matched_keywords}"
        if matched_keywords
        else "하드스킬 근거는 final_score 및 상위 랭킹 포함 여부를 기준으로 보수적으로 판단했습니다."
    )

    soft_summary = (
        "소프트스킬은 정량 컬럼만으로 단정하기 어렵습니다. "
        "협업 방식, 커뮤니케이션, 문제 해결 방식은 인터뷰에서 확인해야 합니다."
    )

    evidence = (
        f"jd_id={jd_id}; candidate_id={candidate_id}; rank={rank}; "
        f"final_score={final_score}; matched_keywords={matched_keywords}; missing_keywords={missing_keywords}"
    )

    return {
        "match_reason_rule_based": reason,
        "match_gap_rule_based": gap,
        "interview_question_rule_based": questions,
        "llm_match_summary": f"후보 {candidate_id}는 JD {jd_id}에 대해 {score_desc} 후보로 분류됩니다.",
        "llm_strength": f"강점: {matched_keywords if matched_keywords else '상위 10명에 포함될 수준의 정량 점수와 관련 이력 보유'}",
        "llm_gap": f"갭: {missing_keywords if missing_keywords else '세부 역량 깊이와 최근 프로젝트 직접성 확인 필요'}",
        "llm_risk": "자동 생성 해석이므로 실제 수행 범위, 책임 수준, 최근성은 인터뷰에서 검증해야 합니다.",
        "llm_recommendation": "상위 후보군으로 면접 검토를 권장하며, JD 핵심역량 중심의 구조화 질문을 권장합니다.",
        "llm_error": "",
        "career_path_summary": " ".join(career_path),
        "career_fit_comment": f"경력 적합도는 {score_desc}입니다. JD와 후보 이력의 직접 연결성을 추가 확인하세요.",
        "career_timeline_json": "{}",
        "hardskill_evidence_json": "{}",
        "hardskill_evidence_summary": hard_summary,
        "softskill_json": "{}",
        "softskill_summary": soft_summary,
        "llm_interview_question": questions,
        "llm_evidence": evidence,
    }

# ------------------------------------------------------------
# 7. source CG 내용 병합
# ------------------------------------------------------------
merged_df = target_df.merge(
    source_semantic_df,
    on=["_merge_jd_id", "_merge_candidate_id"],
    how="left",
    suffixes=("", "_from_source")
)

for col in cg_after_cols:
    if col not in merged_df.columns:
        merged_df[col] = ""
    merged_df[col] = merged_df[col].astype("object")

    source_col = f"{col}_from_source"
    if source_col in merged_df.columns:
        merged_df[source_col] = merged_df[source_col].astype("object")

for col in cg_after_cols:
    source_col = f"{col}_from_source"

    if source_col not in merged_df.columns:
        continue

    blank_mask = merged_df[col].map(is_blank_value)
    source_not_blank_mask = ~merged_df[source_col].map(is_blank_value)
    fill_mask = blank_mask & source_not_blank_mask

    if fill_mask.any():
        merged_df.loc[fill_mask, col] = (
            merged_df.loc[fill_mask, source_col]
            .map(clean_cell_value)
            .astype("object")
        )

drop_source_cols = [
    f"{col}_from_source"
    for col in cg_after_cols
    if f"{col}_from_source" in merged_df.columns
]
merged_df = merged_df.drop(columns=drop_source_cols, errors="ignore")

for col in cg_after_cols:
    merged_df[col] = merged_df[col].astype("object")

# ------------------------------------------------------------
# 8. 그래도 빈 CG 이후 컬럼은 rule-based로 행 단위 생성
# ------------------------------------------------------------
for idx, row in merged_df.iterrows():
    semantic_dict = make_semantic_dict(row)

    for col in cg_after_cols:
        if col not in merged_df.columns:
            merged_df[col] = ""
            merged_df[col] = merged_df[col].astype("object")

        if is_blank_value(merged_df.at[idx, col]):
            merged_df.at[idx, col] = clean_cell_value(semantic_dict.get(col, ""))

# ------------------------------------------------------------
# 9. 원본 컬럼 순서 복구 및 Excel-safe 변환
# ------------------------------------------------------------
merged_df = merged_df.drop(columns=["_merge_jd_id", "_merge_candidate_id"], errors="ignore")

for col in original_columns:
    if col not in merged_df.columns:
        merged_df[col] = ""

final_df = merged_df[original_columns].copy()

for col in final_df.columns:
    if final_df[col].dtype == "object" or str(final_df[col].dtype).startswith("string"):
        final_df[col] = final_df[col].map(clean_cell_value).astype("object")

# ------------------------------------------------------------
# 10. 검증
# ------------------------------------------------------------
def non_empty_count(series):
    return int(series.map(lambda x: not is_blank_value(x)).sum())

cg_audit_df = pd.DataFrame({
    "cg_after_column": cg_after_cols,
    "non_empty_count": [non_empty_count(final_df[c]) for c in cg_after_cols],
    "total_rows": len(final_df),
})

jd_191_count = int((final_df["jd_id"].astype(str) == "191").sum()) if "jd_id" in final_df.columns else None

gold_mask = pd.Series(False, index=final_df.index)
if "candidate_id" in final_df.columns:
    gold_mask |= final_df["candidate_id"].astype(str).str.contains(
        r"^G_JD_|G_JD_|gold|virtual|가상",
        regex=True,
        case=False,
        na=False
    )
gold_remaining_count = int(gold_mask.sum())

print("\n[검증 결과]")
print(f"  - 최종 행 수: {len(final_df):,}")
print(f"  - 최종 컬럼 수: {len(final_df.columns):,}")
print(f"  - 원본 컬럼 전체 보존: {list(final_df.columns) == original_columns}")
print(f"  - 191번 JD 후보자 수: {jd_191_count}")
print(f"  - Gold CV 잔여 의심 행 수: {gold_remaining_count}")

print("\n[CG 이후 컬럼별 채움 현황]")
display(cg_audit_df)

print("\n[191번 JD 결과 확인]")
if "jd_id" in final_df.columns:
    display(final_df.loc[final_df["jd_id"].astype(str) == "191"])

print("\n[CG 이후 컬럼 내용 미리보기]")
preview_cols = [c for c in ["jd_id", "rank", "candidate_id", "final_score"] if c in final_df.columns] + cg_after_cols
display(final_df[preview_cols].head(20))

if gold_remaining_count > 0:
    print("\n[WARN] Gold CV 의심 행")
    display(final_df.loc[gold_mask])
    raise RuntimeError("Gold CV 의심 행이 남아 있어 저장을 중단합니다.")

if jd_191_count != 10:
    raise RuntimeError(f"191번 JD 후보자 수가 10명이 아닙니다. 현재 {jd_191_count}명입니다.")

allowed_empty_cg_cols = ["llm_error"]

empty_cg_cols = cg_audit_df.loc[
    (cg_audit_df["non_empty_count"] == 0)
    & (~cg_audit_df["cg_after_column"].isin(allowed_empty_cg_cols)),
    "cg_after_column"
].tolist()

if empty_cg_cols:
    raise RuntimeError(f"아직 전부 빈 CG 이후 컬럼이 있습니다: {empty_cg_cols}")

print(f"[INFO] 전부 비어 있어도 정상 허용한 컬럼: {allowed_empty_cg_cols}")

# ------------------------------------------------------------
# 11. 저장
# ------------------------------------------------------------
final_df.to_excel(output_excel_path, index=False, engine="openpyxl")
final_df.to_csv(output_csv_path, index=False, encoding="utf-8-sig")
cg_audit_df.to_csv(audit_csv_path, index=False, encoding="utf-8-sig")

print("\n[DONE] CG 이후 내용 채움 완료")
print(f"  - Excel 파일: {output_excel_path}")
print(f"  - CSV 파일: {output_csv_path}")
print(f"  - Audit 파일: {audit_csv_path}")
print(f"  - 191번 JD 후보자 수: {jd_191_count}")
print(f"  - Gold CV 잔여 의심 행 수: {gold_remaining_count}")

display(FileLink(str(output_excel_path)))
display(FileLink(str(output_csv_path)))
display(FileLink(str(audit_csv_path)))
